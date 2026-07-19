
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import psycopg2
import psycopg2.extras
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import io
import base64
import sqlite3
import hashlib
import time
import calendar
from sqlalchemy import create_engine, text
from urllib.parse import quote_plus

from pathlib import Path
import json
from datetime import datetime, date, timedelta
import unicodedata
import re
import traceback

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_EXPORT_OK = True
    EXCEL_EXPORT_ERROR = ""
except Exception as exc:
    Workbook = None
    EXCEL_EXPORT_OK = False
    EXCEL_EXPORT_ERROR = str(exc)

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    PDF_EXPORT_OK = True
    PDF_EXPORT_ERROR = ""
except Exception as exc:
    PDF_EXPORT_OK = False
    PDF_EXPORT_ERROR = str(exc)

st.set_page_config(
    page_title="Rede Economize | KPI Comercial",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    .valor-negativo {
        color: #ff4d4f !important;
        font-weight: 700 !important;
    }
    .status-card-kpi {
        border: 1px solid #254b6b;
        border-radius: 14px;
        padding: 14px 16px;
        min-height: 108px;
        background: #0d2032;
    }
    .status-card-kpi .titulo {
        font-size: 13px;
        font-weight: 700;
        color: #ffffff;
    }
    .status-card-kpi .valor {
        font-size: 29px;
        line-height: 1.2;
        margin-top: 8px;
        color: #ffffff;
        font-variant-numeric: tabular-nums;
    }
    .status-card-kpi .status {
        display: inline-block;
        margin-top: 8px;
        padding: 3px 8px;
        border-radius: 999px;
        background: rgba(46, 204, 113, .18);
        color: #2ecc71;
        font-size: 12px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# PERFORMANCE / CACHE
# =========================================================

def _arquivo_token(*caminhos):
    """Token leve para invalidar caches quando arquivos persistentes mudarem."""
    partes = []
    for caminho in caminhos:
        try:
            path = Path(caminho)
            stat = path.stat()
            partes.append(f"{path}:{stat.st_mtime_ns}:{stat.st_size}")
        except Exception:
            partes.append(f"{caminho}:0:0")
    return "|".join(partes)

def _limpar_cache_dados():
    try:
        st.cache_data.clear()
    except Exception:
        pass


# =========================================================
# DADOS DEMONSTRATIVOS
# =========================================================

COMPRADORES = ["Paulo", "Francieli", "Sebastião"]

REALIZADOS = pd.DataFrame([
    ["Paulo", 1083974.52, 25.7, 961221.38, 88.7, 856101.24, 1.00, 69104.34, 8.1, 269908.06, 31.5, 460514.54, 53.8, 56574.30, 6.6, 13281.02, 1.2, 1003613.79, 104.4],
    ["Francieli", 1037338.03, 24.6, 708983.83, 68.3, 1155715.96, 2.00, 104317.28, 9.0, 278891.20, 24.1, 595751.47, 51.5, 176756.01, 15.3, 32132.06, 3.1, 807929.11, 114.0],
    ["Sebastião", 2102632.93, 49.8, 1239845.83, 59.0, 1951037.63, 2.00, 687151.14, 35.2, 661165.12, 33.9, 508011.09, 26.0, 94710.28, 4.9, 25934.34, 1.2, 1615224.91, 130.3],
], columns=[
    "Comprador", "Faturamento Total Atual", "Rep. Faturamento",
    "CMV mês Atual", "Rep. CMV", "Estoque Total", "Fator Cobertura",
    "Estoque Curva A", "Rep. Curva A", "Estoque Curva B", "Rep. Curva B",
    "Estoque Curva C", "Rep. Curva C", "Estoque Curva D", "Rep. Curva D",
    "Ruptura Ativa", "Ruptura %", "Entradas CUSTO", "Reposição CMV %"
])

METAS = pd.DataFrame([
    ["Paulo", 1093500.00, 27.0, 969571.19, 88.7, 1090767.59, 1.13, 218153.52, 20.0, 327230.28, 30.0, 436307.04, 40.0, 109076.76, 10.0, 54675.00, 5.0, 959875.48, 99.0],
    ["Francieli", 972000.00, 24.0, 664261.17, 68.3, 1252606.78, 1.89, 250521.36, 20.0, 375782.03, 30.0, 501042.71, 40.0, 125260.68, 10.0, 48600.00, 5.0, 657618.56, 99.0],
    ["Sebastião", 1984500.00, 49.0, 1170070.14, 59.0, 1880469.86, 1.61, 376093.97, 20.0, 564140.96, 30.0, 752187.95, 40.0, 188046.99, 10.0, 99225.00, 5.0, 1158369.44, 99.0],
], columns=[
    "Comprador", "Faturamento Total META", "Rep. Faturamento",
    "CMV mês META", "Rep. CMV", "Estoque Total META", "Fator Cobertura",
    "Estoque Curva A", "Rep. Curva A", "Estoque Curva B", "Rep. Curva B",
    "Estoque Curva C", "Rep. Curva C", "Estoque Curva D", "Rep. Curva D",
    "Ruptura Ativa", "Ruptura %", "Entradas CUSTO", "Reposição CMV %"
])

RESULTADO = pd.DataFrame([
    ["Paulo", 9525.48, 1.3, 8349.81, 0.0, 234666.36, 23.4, 149049.18, 11.9, 57322.22, -1.5, -24207.50, -13.8, 52502.46, 3.4, 41393.98, 3.8, -43738.31, -5.4],
    ["Francieli", -65338.03, -0.6, -44722.66, 0.0, 96890.82, 25.6, 146204.08, 11.0, 96890.83, 5.9, -94708.75, -11.5, -51495.33, -5.3, 16467.94, 1.9, -150310.55, -15.0],
    ["Sebastião", -118132.93, -0.8, -69775.69, 0.0, -70567.77, 3.4, -311057.17, -15.2, -97024.16, -3.9, 244176.85, 14.0, 93336.71, 5.1, 73290.66, 3.8, -456855.47, -31.3],
], columns=REALIZADOS.columns)

PREMIO = pd.DataFrame([
    ["Paulo", 28.51, 95.0, 30.00, 100.0, 95.66, 95.7, 96.65, 64.4, 149.61, 99.7, 88.11, 88.1, 88.50, 88.5, 86.01, 43.0, 139.58, 99.7],
    ["Francieli", 92.09, 102.3, 89.99, 100.0, 294.49, 98.2, 314.52, 69.9, 432.78, 96.2, 274.99, 91.7, 215.92, 72.0, 513.14, 85.5, 410.41, 97.7],
    ["Sebastião", 60.95, 101.6, 59.99, 100.0, 199.91, 100.0, 126.27, 42.1, 294.96, 98.3, 175.63, 87.8, 147.04, 73.5, 173.01, 43.3, 252.05, 90.0],
], columns=[
    "Comprador", "Faturamento Prêmio", "Faturamento Realizado",
    "CMV Prêmio", "CMV Realizado", "Estoque Total Prêmio", "Estoque Total Realizado",
    "Curva A Prêmio", "Curva A Realizado", "Curva B Prêmio", "Curva B Realizado",
    "Curva C Prêmio", "Curva C Realizado", "Curva D Prêmio", "Curva D Realizado",
    "Ruptura Prêmio", "Ruptura Realizado", "Entradas Prêmio", "Entradas Realizado"
])

PREMIO_KPI = pd.DataFrame([
    ["Faturamento", 3.0, 90.00, 102.3, 92.09],
    ["CMV", 3.0, 90.00, 100.0, 89.99],
    ["Fator Cobertura", 10.0, 300.00, 98.2, 294.49],
    ["Estoque Curva A", 15.0, 450.00, 69.9, 314.52],
    ["Estoque Curva B", 15.0, 450.00, 96.2, 432.78],
    ["Estoque Curva C", 10.0, 300.00, 91.7, 274.99],
    ["Estoque Curva D", 10.0, 300.00, 72.0, 215.92],
    ["Ruptura Ativa", 20.0, 600.00, 85.5, 513.14],
    ["Reposição CMV", 14.0, 420.00, 97.7, 410.41],
], columns=["KPI", "Peso sobre a meta", "Prêmio por KPI atingível", "Atingimento %", "Prêmio Atingido"])

# =========================================================
# FUNÇÕES
# =========================================================

def _numero_base(v, casas=2):
    """Formata número no padrão pt-BR."""
    if v is None or pd.isna(v):
        return ""
    try:
        numero = float(v)
    except (TypeError, ValueError):
        return str(v)

    formato = f"{numero:,.{int(casas)}f}"
    return (
        formato
        .replace(",", "\u0000")
        .replace(".", ",")
        .replace("\u0000", ".")
    )


def moeda(v):
    """Valor monetário sem o prefixo, sempre com duas casas."""
    if v is None or pd.isna(v):
        return ""
    return _numero_base(v, 2)


def moeda_real(v):
    """Valor monetário: R$ 100,00 ou R$ -100,00."""
    if v is None or pd.isna(v):
        return ""
    return f"R$ {_numero_base(v, 2)}"


def percentual(v):
    """Percentual: 100,00%."""
    if v is None or pd.isna(v):
        return ""
    return f"{_numero_base(v, 2)}%"


def numero_inteiro(v):
    """Número inteiro: 1.000."""
    if v is None or pd.isna(v):
        return ""
    try:
        return _numero_base(round(float(v)), 0)
    except (TypeError, ValueError):
        return str(v)


def numero_decimal(v, casas=2):
    """Número decimal sem símbolo monetário."""
    return _numero_base(v, casas)


def br_num(v, casas=2):
    """Compatibilidade com chamadas antigas."""
    return numero_decimal(v, casas)


def _nome_normalizado(nome):
    return str(nome).strip().casefold()


def _coluna_percentual(nome):
    texto = _nome_normalizado(nome)

    # Campos de valor que contêm as palavras CMV ou Margem não são percentuais.
    # Somente versões explicitamente percentuais, como "(%)", são tratadas
    # como percentual.
    campos_monetarios_exatos = {
        "cmv mês meta",
        "cmv mes meta",
        "cmv mês atual",
        "cmv mes atual",
        "margem bruta meta",
        "margem bruta atual",
        "meta margem bruta",
        "meta margem bruta (r$)",
        "margem bruta (r$)",
    }
    if texto in campos_monetarios_exatos:
        return False

    if "%" in texto or "percent" in texto:
        return True

    return any(chave in texto for chave in [
        "atingimento", "participação", "participacao",
        "reposição", "reposicao", "peso", "representatividade",
        "margem (%)", "cmv (%)", "cobertura (%)"
    ])


def _coluna_inteira(nome):
    texto = _nome_normalizado(nome)
    return any(chave in texto for chave in [
        "quantidade", "qtd", "itens", "item", "registros", "registro",
        "dias", "dia", "lojas", "loja", "produtos", "produto",
        "não mapeados", "nao mapeados", "código", "codigo",
        "parcela", "posição", "posicao"
    ])


def _coluna_monetaria(nome):
    texto = _nome_normalizado(nome)
    if _coluna_percentual(nome) or _coluna_inteira(nome):
        return False

    campos_monetarios_exatos = {
        "cmv mês meta",
        "cmv mes meta",
        "cmv mês atual",
        "cmv mes atual",
        "margem bruta meta",
        "margem bruta atual",
        "meta margem bruta",
        "meta margem bruta (r$)",
        "margem bruta (r$)",
    }
    if texto in campos_monetarios_exatos:
        return True

    return any(chave in texto for chave in [
        "venda", "faturamento", "custo", "estoque", "entrada", "compra",
        "lucro", "cmv", "margem bruta", "meta mês", "meta mes",
        "meta venda", "prêmio", "premio", "pagamento",
        "contas a pagar", "ruptura", "necessidade", "valor",
        "entrega (r$)", "realizado", "saldo", "documento"
    ])



def cor_valor(v):
    try:
        numero = float(v)
    except (TypeError, ValueError):
        return "#ffffff"
    return "#ff4d4f" if numero < 0 else "#ffffff"


def formatar_valor_grafico(v, tipo):
    if tipo == "percentual":
        return percentual(v)
    if tipo == "moeda":
        return moeda_real(v)
    if tipo == "inteiro":
        return numero_inteiro(v)
    return numero_decimal(v, 2)


def detectar_tipo_grafico(fig):
    textos = []
    try:
        textos.append(str(fig.layout.xaxis.title.text or ""))
    except Exception:
        pass
    try:
        textos.append(str(fig.layout.yaxis.title.text or ""))
    except Exception:
        pass
    try:
        textos.append(str(fig.layout.title.text or ""))
    except Exception:
        pass
    texto = " ".join(textos).casefold()

    if any(chave in texto for chave in [
        "percent", "%", "cmv", "margem", "reposição", "reposicao",
        "atingimento", "participação", "participacao"
    ]):
        return "percentual"
    if any(chave in texto for chave in [
        "r$", "valor", "venda", "faturamento", "custo", "estoque",
        "entrada", "lucro", "pagamento", "ruptura", "contas a pagar"
    ]):
        return "moeda"
    if any(chave in texto for chave in [
        "itens", "quantidade", "registros", "dias"
    ]):
        return "inteiro"
    return "numero"


def aplicar_formato_grafico(fig, tipo=None):
    """Padroniza eixo, rótulo, tooltip e negativos dos traces."""
    tipo = tipo or detectar_tipo_grafico(fig)
    try:
        fig.update_layout(separators=",.")

        if tipo == "percentual":
            fig.update_xaxes(ticksuffix="%", tickformat=",.2f")
            fig.update_yaxes(ticksuffix="%", tickformat=",.2f")
        elif tipo == "moeda":
            fig.update_xaxes(tickprefix="R$ ", tickformat=",.2f")
            fig.update_yaxes(tickprefix="R$ ", tickformat=",.2f")
        elif tipo == "inteiro":
            fig.update_xaxes(tickformat=",.0f")
            fig.update_yaxes(tickformat=",.0f")

        fig.update_xaxes(exponentformat="none", separatethousands=True)
        fig.update_yaxes(exponentformat="none", separatethousands=True)

        for trace in fig.data:
            valores = None
            if getattr(trace, "orientation", None) == "h":
                valores = getattr(trace, "x", None)
            else:
                valores = getattr(trace, "y", None)

            if valores is None:
                continue

            textos = []
            cores = []
            for valor in valores:
                try:
                    numero = float(valor)
                    textos.append(formatar_valor_grafico(numero, tipo))
                    cores.append("#ff4d4f" if numero < 0 else "#ffffff")
                except (TypeError, ValueError):
                    textos.append(str(valor))
                    cores.append("#ffffff")

            try:
                trace.text = textos
                trace.texttemplate = "%{text}"
                trace.textfont = dict(color=cores)
                trace.hovertemplate = "%{fullData.name}<br>%{text}<extra></extra>"
                trace.cliponaxis = False
            except Exception:
                pass
    except Exception:
        pass
    return fig


def card_status_base(titulo, registros):
    registrar_card_exportacao(titulo, numero_inteiro(registros))
    st.markdown(
        f"""
        <div class="status-card-kpi">
            <div class="titulo">{titulo}</div>
            <div class="valor">{numero_inteiro(registros)}</div>
            <div class="status">↑ registros salvos</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def contar_registros_cache(tabela, periodo=None):
    """Conta os registros reais no SQLite para o período selecionado."""
    try:
        with sqlite3.connect(CACHE_DB_FILE, timeout=30) as con:
            existe = con.execute(
                "SELECT 1 FROM sqlite_master "
                "WHERE type='table' AND name=?",
                (tabela,),
            ).fetchone()
            if not existe:
                return 0

            colunas = {
                linha[1]
                for linha in con.execute(
                    f'PRAGMA table_info("{tabela}")'
                ).fetchall()
            }
            if periodo and "periodo_referencia" in colunas:
                return int(
                    con.execute(
                        f'SELECT COUNT(*) FROM "{tabela}" '
                        'WHERE periodo_referencia=?',
                        (str(periodo),),
                    ).fetchone()[0]
                )
            return int(
                con.execute(
                    f'SELECT COUNT(*) FROM "{tabela}"'
                ).fetchone()[0]
            )
    except Exception:
        return 0



# =========================================================
# EXPORTAÇÃO UNIVERSAL DA TELA ATUAL
# =========================================================

def iniciar_contexto_exportacao(visao_atual, periodo_atual):
    st.session_state["_export_visao"] = str(visao_atual)
    st.session_state["_export_periodo"] = str(periodo_atual)
    st.session_state["_export_tabelas"] = []
    st.session_state["_export_cards"] = []
    st.session_state["_export_graficos"] = []
    st.session_state.pop("_export_pdf_bytes", None)
    st.session_state.pop("_export_excel_bytes", None)
    st.session_state.pop("_export_nome_base", None)


def registrar_tabela_exportacao(dados, titulo=None):
    if not isinstance(dados, pd.DataFrame):
        return
    try:
        numero = len(st.session_state.get("_export_tabelas", [])) + 1
        st.session_state.setdefault("_export_tabelas", []).append({
            "titulo": titulo or f"Tabela {numero}",
            "dados": dados.copy(),
        })
    except Exception:
        pass


def registrar_card_exportacao(titulo, valor):
    try:
        st.session_state.setdefault("_export_cards", []).append({
            "titulo": str(titulo),
            "valor": str(valor),
        })
    except Exception:
        pass


def _lista_plotly(valor):
    if valor is None:
        return []
    try:
        return list(valor)
    except Exception:
        return []


def registrar_grafico_exportacao(fig):
    try:
        titulo = "Gráfico"
        try:
            titulo = str(fig.layout.title.text or titulo)
        except Exception:
            pass
        registros = []
        for trace in fig.data:
            nome = str(getattr(trace, "name", "") or "Série")
            horizontal = getattr(trace, "orientation", None) == "h"
            categorias = _lista_plotly(
                getattr(trace, "y", None) if horizontal else getattr(trace, "x", None)
            )
            valores = _lista_plotly(
                getattr(trace, "x", None) if horizontal else getattr(trace, "y", None)
            )
            tamanho = max(len(categorias), len(valores))
            for i in range(tamanho):
                registros.append({
                    "Série": nome,
                    "Categoria": categorias[i] if i < len(categorias) else "",
                    "Valor": valores[i] if i < len(valores) else None,
                })
        if registros:
            st.session_state.setdefault("_export_graficos", []).append({
                "titulo": titulo,
                "dados": pd.DataFrame(registros),
            })
    except Exception:
        pass


def _sanitizar_nome_arquivo(texto):
    texto = unicodedata.normalize("NFKD", str(texto))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^A-Za-z0-9_-]+", "_", texto).strip("_")
    return texto or "exportacao"


def _nome_aba_excel(nome, usados):
    nome = re.sub(r'[:\\/?*\[\]]', " ", str(nome)).strip()
    nome = re.sub(r"\s+", " ", nome) or "Dados"
    base = nome[:31]
    candidato = base
    contador = 2
    while candidato.casefold() in usados:
        sufixo = f" {contador}"
        candidato = base[:31-len(sufixo)] + sufixo
        contador += 1
    usados.add(candidato.casefold())
    return candidato


def _tipo_coluna_exportacao(nome):
    if _coluna_percentual(nome):
        return "percentual"
    if _coluna_monetaria(nome):
        return "moeda"
    if _coluna_inteira(nome):
        return "inteiro"
    texto = _nome_normalizado(nome)
    if any(chave in texto for chave in ["data", "competência", "competencia", "período", "periodo"]):
        return "data"
    return "geral"


def gerar_excel_tela(visao_atual, periodo_atual, cards, tabelas, graficos):
    if not EXCEL_EXPORT_OK:
        raise RuntimeError(f"Biblioteca openpyxl não instalada: {EXCEL_EXPORT_ERROR}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    azul, ciano, branco, vermelho = "0D2032", "43D7E8", "FFFFFF", "FF4D4F"
    borda = Side(style="thin", color="35536D")
    ws["A1"] = "REDE ECONOMIZE - KPI COMERCIAL"
    ws["A2"] = str(visao_atual)
    ws["A3"] = f"Período: {periodo_atual}"
    ws["A4"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    for celula in ("A1", "A2"):
        ws[celula].fill = PatternFill("solid", fgColor=azul)
        ws[celula].font = Font(color=branco, bold=True, size=14)
        ws[celula].alignment = Alignment(horizontal="center")
    linha = 6
    if cards:
        ws.cell(linha, 1, "Indicador")
        ws.cell(linha, 2, "Valor")
        for celula in ws[linha]:
            celula.fill = PatternFill("solid", fgColor=ciano)
            celula.font = Font(bold=True)
        linha += 1
        for item in cards:
            ws.cell(linha, 1, item["titulo"])
            ws.cell(linha, 2, item["valor"])
            linha += 1
    usados = {ws.title.casefold()}
    for indice, item in enumerate(list(tabelas) + list(graficos), start=1):
        df = item.get("dados")
        if not isinstance(df, pd.DataFrame):
            continue
        aba = wb.create_sheet(_nome_aba_excel(item.get("titulo") or f"Dados {indice}", usados))
        aba.freeze_panes = "A2"
        for col_idx, coluna in enumerate(df.columns, start=1):
            celula = aba.cell(1, col_idx, str(coluna))
            celula.fill = PatternFill("solid", fgColor=azul)
            celula.font = Font(color=branco, bold=True)
            celula.alignment = Alignment(horizontal="center")
            celula.border = Border(left=borda, right=borda, top=borda, bottom=borda)
        for row_idx, valores in enumerate(df.itertuples(index=False, name=None), start=2):
            for col_idx, valor in enumerate(valores, start=1):
                coluna = df.columns[col_idx - 1]
                celula = aba.cell(row_idx, col_idx)
                if pd.isna(valor):
                    valor = None
                elif isinstance(valor, np.generic):
                    valor = valor.item()
                elif isinstance(valor, pd.Timestamp):
                    valor = valor.to_pydatetime()
                celula.value = valor
                tipo = _tipo_coluna_exportacao(coluna)
                if tipo == "moeda" and isinstance(valor, (int, float)):
                    celula.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
                elif tipo == "percentual" and isinstance(valor, (int, float)):
                    celula.number_format = '0.00"%";[Red]-0.00"%"'
                elif tipo == "inteiro" and isinstance(valor, (int, float)):
                    celula.number_format = '#,##0;[Red]-#,##0'
                elif isinstance(valor, float):
                    celula.number_format = '#,##0.00;[Red]-#,##0.00'
                elif tipo == "data" and isinstance(valor, (datetime, date)):
                    celula.number_format = "dd/mm/yyyy"
                if isinstance(valor, (int, float)) and valor < 0:
                    celula.font = Font(color=vermelho)
                celula.border = Border(left=borda, right=borda, top=borda, bottom=borda)
                celula.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, coluna in enumerate(df.columns, start=1):
            max_len = len(str(coluna))
            for valor in df.iloc[:500, col_idx - 1]:
                if not pd.isna(valor):
                    max_len = max(max_len, len(str(valor)))
            aba.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 38)
        aba.auto_filter.ref = aba.dimensions
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 25
    ws.sheet_view.showGridLines = False
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _texto_pdf(valor, coluna):
    if pd.isna(valor):
        return ""
    tipo = _tipo_coluna_exportacao(coluna)
    if isinstance(valor, (int, float, np.number)):
        if tipo == "moeda": return moeda_real(valor)
        if tipo == "percentual": return percentual(valor)
        if tipo == "inteiro": return numero_inteiro(valor)
        return numero_decimal(valor, 2)
    if isinstance(valor, (datetime, date, pd.Timestamp)):
        return valor.strftime("%d/%m/%Y")
    return str(valor)


def gerar_pdf_tela(visao_atual, periodo_atual, cards, tabelas, graficos):
    if not PDF_EXPORT_OK:
        raise RuntimeError(f"Biblioteca reportlab não instalada: {PDF_EXPORT_ERROR}")
    buffer = io.BytesIO()
    pagina = landscape(A4)
    doc = SimpleDocTemplate(buffer, pagesize=pagina, leftMargin=22, rightMargin=22,
        topMargin=22, bottomMargin=22, title=f"{visao_atual} - {periodo_atual}", author="Rede Economize")
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("TituloKPI", parent=estilos["Title"], fontName="Helvetica-Bold",
        fontSize=17, leading=21, textColor=colors.HexColor("#0D2032"), alignment=TA_CENTER, spaceAfter=8)
    estilo_sub = ParagraphStyle("SubKPI", parent=estilos["Normal"], fontSize=9, leading=12,
        alignment=TA_CENTER, textColor=colors.HexColor("#35536D"), spaceAfter=12)
    estilo_secao = ParagraphStyle("SecaoKPI", parent=estilos["Heading2"], fontName="Helvetica-Bold",
        fontSize=11, leading=14, textColor=colors.HexColor("#0D2032"), spaceBefore=8, spaceAfter=6)
    estilo_normal = ParagraphStyle("NormalKPI", parent=estilos["Normal"], fontSize=7, leading=9)
    estilo_negativo = ParagraphStyle("NegativoKPI", parent=estilo_normal, textColor=colors.HexColor("#FF4D4F"))
    elementos = [Paragraph("REDE ECONOMIZE - KPI COMERCIAL", estilo_titulo),
        Paragraph(f"{visao_atual} | Período: {periodo_atual} | Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", estilo_sub)]
    if cards:
        elementos.append(Paragraph("Indicadores", estilo_secao))
        dados_cards = [["Indicador", "Valor"]] + [[i["titulo"], i["valor"]] for i in cards]
        tabela_cards = Table(dados_cards, colWidths=[300, 180], repeatRows=1)
        tabela_cards.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0D2032")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 8), ("GRID", (0,0), (-1,-1), .35, colors.HexColor("#35536D")),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#EDF3F8")]),
            ("VALIGN", (0,0), (-1,-1), "TOP")]))
        elementos.extend([tabela_cards, Spacer(1,10)])
    conjuntos = list(tabelas) + list(graficos)
    if not conjuntos:
        elementos.append(Paragraph("Esta tela não possui tabelas ou séries de gráfico para exportação.", estilo_normal))
    limite_pdf = 2000
    for indice, item in enumerate(conjuntos, start=1):
        df = item.get("dados")
        if not isinstance(df, pd.DataFrame): continue
        elementos.append(Paragraph(str(item.get("titulo") or f"Dados {indice}"), estilo_secao))
        df_pdf = df.head(limite_pdf).copy()
        if len(df) > limite_pdf:
            elementos.extend([Paragraph(f"PDF limitado aos primeiros {numero_inteiro(limite_pdf)} de {numero_inteiro(len(df))} registros. O Excel contém a base completa.", estilo_normal), Spacer(1,4)])
        colunas = list(df_pdf.columns)
        if not colunas: continue
        largura = max(45, min(110, (pagina[0]-44)/max(len(colunas),1)))
        dados_pdf = [[Paragraph(str(c), estilo_normal) for c in colunas]]
        for valores in df_pdf.itertuples(index=False, name=None):
            linha_pdf = []
            for col_idx, valor in enumerate(valores):
                estilo = estilo_negativo if isinstance(valor, (int,float,np.number)) and valor < 0 else estilo_normal
                linha_pdf.append(Paragraph(_texto_pdf(valor, colunas[col_idx]), estilo))
            dados_pdf.append(linha_pdf)
        tabela = Table(dados_pdf, colWidths=[largura]*len(colunas), repeatRows=1, hAlign="LEFT")
        tabela.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0D2032")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID", (0,0), (-1,-1), .3, colors.HexColor("#35536D")),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#EDF3F8")]),
            ("VALIGN", (0,0), (-1,-1), "TOP"), ("LEFTPADDING", (0,0), (-1,-1), 3),
            ("RIGHTPADDING", (0,0), (-1,-1), 3), ("TOPPADDING", (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3)]))
        elementos.extend([tabela, Spacer(1,10)])
    def rodape(canvas, documento):
        canvas.saveState(); canvas.setFont("Helvetica",7); canvas.setFillColor(colors.HexColor("#35536D"))
        canvas.drawString(22,10,f"Rede Economize - {visao_atual}")
        canvas.drawRightString(pagina[0]-22,10,f"Página {documento.page}"); canvas.restoreState()
    doc.build(elementos, onFirstPage=rodape, onLaterPages=rodape)
    return buffer.getvalue()


def renderizar_botao_pdf_identico(visao_atual, periodo_atual):
    """Abre a impressão nativa preservando visual, cores, cards e gráficos."""
    nome_documento = (
        _sanitizar_nome_arquivo(visao_atual)
        + "_"
        + _sanitizar_nome_arquivo(periodo_atual)
    )
    html = f"""
    <html>
    <head>
      <style>
        html, body {{ margin: 0; padding: 0; background: transparent; }}
        button {{
          width: 100%; height: 42px; border: 1px solid #2c668b;
          border-radius: 8px; background: #0d2032; color: white;
          font-family: Arial, sans-serif; font-size: 14px; font-weight: 700;
          cursor: pointer;
        }}
        button:hover {{ background: #143653; border-color: #43d7e8; }}
      </style>
    </head>
    <body>
      <button onclick="exportarPDF()">🖨️ Exportar PDF idêntico à tela</button>
      <script>
        function exportarPDF() {{
          const w = window.parent;
          const d = w.document;
          const tituloAnterior = d.title;
          d.title = {nome_documento!r};

          let estilo = d.getElementById('rede-economize-pdf-visual');
          if (!estilo) {{
            estilo = d.createElement('style');
            estilo.id = 'rede-economize-pdf-visual';
            estilo.innerHTML = `
              @media print {{
                @page {{ size: A4 landscape; margin: 7mm; }}
                * {{
                  -webkit-print-color-adjust: exact !important;
                  print-color-adjust: exact !important;
                }}
                html, body, .stApp,
                [data-testid="stAppViewContainer"],
                [data-testid="stMain"] {{
                  background: #06111e !important;
                  overflow: visible !important;
                  height: auto !important;
                  min-height: 0 !important;
                }}
                [data-testid="stHeader"],
                [data-testid="stToolbar"],
                [data-testid="stDecoration"],
                [data-testid="stStatusWidget"],
                [data-testid="stBottomBlockContainer"],
                .stDeployButton,
                footer {{ display: none !important; }}

                [data-testid="stSidebar"] {{
                  position: static !important;
                  transform: none !important;
                  min-width: 225px !important;
                  width: 225px !important;
                  height: auto !important;
                  overflow: visible !important;
                }}
                [data-testid="stSidebarContent"] {{
                  height: auto !important;
                  overflow: visible !important;
                }}
                .main .block-container,
                [data-testid="stMainBlockContainer"] {{
                  max-width: none !important;
                  width: 100% !important;
                  padding: 5mm !important;
                  overflow: visible !important;
                }}
                [data-testid="stVerticalBlock"],
                [data-testid="stHorizontalBlock"] {{
                  overflow: visible !important;
                }}
                [data-testid="stDataFrame"],
                [data-testid="stPlotlyChart"],
                [data-testid="stMetric"],
                [data-testid="stExpanderDetails"],
                .status-card-kpi,
                .meta-card {{
                  break-inside: avoid !important;
                  page-break-inside: avoid !important;
                }}
                [data-testid="stDataFrame"] {{
                  height: auto !important;
                  max-height: none !important;
                  overflow: visible !important;
                }}
                [data-testid="stDataFrame"] > div,
                [data-testid="stDataFrame"] iframe {{
                  height: auto !important;
                  max-height: none !important;
                }}
                details {{ display: block !important; }}
                details > summary {{ display: none !important; }}
                iframe[title="streamlit_component"] {{ display: none !important; }}
                button, .stButton, .stDownloadButton {{ display: none !important; }}
                a {{ color: inherit !important; text-decoration: none !important; }}
              }}
            `;
            d.head.appendChild(estilo);
          }}

          const restaurar = () => {{
            d.title = tituloAnterior;
            w.removeEventListener('afterprint', restaurar);
          }};
          w.addEventListener('afterprint', restaurar);
          w.focus();
          setTimeout(() => w.print(), 350);
        }}
      </script>
    </body>
    </html>
    """
    components.html(html, height=48, scrolling=False)


def renderizar_exportacao_tela():
    """Módulo único de exportação, sem expanders aninhados."""
    visao_atual = st.session_state.get("_export_visao", "Tela")
    periodo_atual = st.session_state.get("_export_periodo", "")
    tabelas = st.session_state.get("_export_tabelas", [])
    cards = st.session_state.get("_export_cards", [])
    graficos = st.session_state.get("_export_graficos", [])

    st.markdown("---")
    st.markdown("### 📤 Exportar esta tela")
    st.caption(
        "Escolha o formato. A geração ocorre somente quando solicitada, "
        "preservando a performance da navegação."
    )

    aba_visual, aba_analitico, aba_excel = st.tabs([
        "🖨️ PDF visual",
        "📄 PDF analítico",
        "📊 Excel completo",
    ])

    nome_base_padrao = (
        _sanitizar_nome_arquivo(visao_atual)
        + "_"
        + _sanitizar_nome_arquivo(periodo_atual)
    )

    with aba_visual:
        st.markdown("#### PDF visual - igual à tela")
        st.caption(
            "Abre a impressão do navegador preservando o tema, cards, "
            "gráficos e tabelas exibidos."
        )
        componentes_html = """
        <script>
        function imprimirTelaKPI() {
            const topWindow = window.parent;
            const topDocument = topWindow.document;
            const styleId = 'kpi-print-style';
            let oldStyle = topDocument.getElementById(styleId);
            if (oldStyle) oldStyle.remove();
            const style = topDocument.createElement('style');
            style.id = styleId;
            style.innerHTML = `
                @page { size: A4 landscape; margin: 8mm; }
                @media print {
                    html, body, [data-testid="stAppViewContainer"],
                    [data-testid="stMain"], .stApp {
                        background: #020c16 !important;
                        color: #ffffff !important;
                        -webkit-print-color-adjust: exact !important;
                        print-color-adjust: exact !important;
                    }
                    [data-testid="stSidebar"],
                    [data-testid="stHeader"],
                    [data-testid="stToolbar"],
                    [data-testid="stDecoration"],
                    footer, .stDeployButton,
                    button[kind="header"],
                    [data-testid="collapsedControl"],
                    iframe[title="streamlit_components"] {
                        display: none !important;
                    }
                    [data-testid="stMainBlockContainer"] {
                        max-width: none !important;
                        padding: 0 !important;
                        margin: 0 !important;
                    }
                    .js-plotly-plot, .plot-container,
                    [data-testid="stDataFrame"],
                    [data-testid="stMetric"], .element-container {
                        break-inside: avoid !important;
                        page-break-inside: avoid !important;
                    }
                    [data-testid="stDataFrame"] { overflow: visible !important; }
                    canvas, svg { max-width: 100% !important; }
                }
            `;
            topDocument.head.appendChild(style);
            setTimeout(() => { topWindow.focus(); topWindow.print(); }, 350);
        }
        </script>
        <button onclick="imprimirTelaKPI()" style="
            width:100%; border:1px solid #35536d; border-radius:8px;
            padding:12px 16px; background:#0d2032; color:#ffffff;
            font-size:15px; font-weight:700; cursor:pointer;">
            🖨️ Abrir impressão para salvar como PDF
        </button>
        """
        components.html(componentes_html, height=58)
        st.info(
            "Na janela de impressão, selecione **Salvar como PDF**, "
            "orientação **Paisagem** e ative **Gráficos de plano de fundo**."
        )

    with aba_analitico:
        st.markdown("#### PDF analítico")
        st.caption(
            "Relatório estruturado com indicadores, tabelas e dados das "
            "séries dos gráficos."
        )
        if st.button(
            "Gerar PDF analítico", use_container_width=True,
            key=f"gerar_pdf_analitico_{_sanitizar_nome_arquivo(visao_atual)}",
        ):
            if not REPORTLAB_DISPONIVEL:
                st.error(
                    "A biblioteca ReportLab não está instalada. Execute o "
                    "iniciador do projeto novamente para instalar as dependências."
                )
            else:
                with st.spinner("Preparando PDF analítico..."):
                    st.session_state["_export_pdf_bytes"] = gerar_pdf_tela(
                        visao_atual, periodo_atual, cards, tabelas, graficos
                    )
                    st.session_state["_export_nome_base"] = nome_base_padrao
                st.success("PDF analítico preparado.")
        if st.session_state.get("_export_pdf_bytes"):
            st.download_button(
                "⬇️ Baixar PDF analítico",
                data=st.session_state["_export_pdf_bytes"],
                file_name=f"{st.session_state.get('_export_nome_base', nome_base_padrao)}.pdf",
                mime="application/pdf", use_container_width=True,
                key=f"download_pdf_analitico_{nome_base_padrao}",
            )
        else:
            st.button(
                "⬇️ Baixar PDF analítico", disabled=True,
                use_container_width=True,
                key=f"pdf_analitico_desabilitado_{nome_base_padrao}",
            )

    with aba_excel:
        st.markdown("#### Excel completo")
        st.caption("Contém os registros capturados na tela em abas separadas.")
        if st.button(
            "Gerar Excel desta tela", use_container_width=True,
            key=f"gerar_excel_{_sanitizar_nome_arquivo(visao_atual)}",
        ):
            with st.spinner("Preparando Excel..."):
                st.session_state["_export_excel_bytes"] = gerar_excel_tela(
                    visao_atual, periodo_atual, cards, tabelas, graficos
                )
                st.session_state["_export_nome_base"] = nome_base_padrao
            st.success("Excel preparado.")
        if st.session_state.get("_export_excel_bytes"):
            st.download_button(
                "⬇️ Baixar Excel",
                data=st.session_state["_export_excel_bytes"],
                file_name=f"{st.session_state.get('_export_nome_base', nome_base_padrao)}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"download_excel_{nome_base_padrao}",
            )
        else:
            st.button(
                "⬇️ Baixar Excel", disabled=True, use_container_width=True,
                key=f"excel_desabilitado_{nome_base_padrao}",
            )


def dataframe_br(dados, *args, **kwargs):
    """Exibe tabelas no padrão brasileiro sem alterar os dados salvos."""
    export_title = kwargs.pop("export_title", None)
    if isinstance(dados, pd.DataFrame):
        registrar_tabela_exportacao(dados, export_title)
    if not isinstance(dados, pd.DataFrame):
        return st.dataframe(dados, *args, **kwargs)

    exibicao = dados.copy()

    indices_percentuais = {
        "cmv geral", "margem contribuição", "margem contribuicao",
        "margem (%)", "cmv (%)", "reposição cmv", "reposicao cmv",
        "atingimento", "participação", "participacao"
    }
    indices_monetarios = {
        "venda geral", "custo médio geral", "custo medio geral",
        "entrada geral", "lucro bruto geral", "pagamento de fornecedor",
        "custo médio - pagamento de fornecedor (caixa)",
        "custo medio - pagamento de fornecedor (caixa)",
        "custo médio geral - entrada geral (competência)",
        "custo medio geral - entrada geral (competencia)",
        "contas a pagar fornecedor total", "estoque mês", "estoque mes",
        "estoque - contas a pagar"
    }
    indices_inteiros = {
        "itens", "quantidade", "registros", "dias"
    }

    # Matrizes com o indicador no índice.
    if len(exibicao.index):
        for indice in exibicao.index:
            nome_indice = _nome_normalizado(indice)
            if nome_indice in indices_percentuais:
                for coluna in exibicao.columns:
                    valor = exibicao.loc[indice, coluna]
                    if isinstance(valor, (int, float, np.number)) and not pd.isna(valor):
                        exibicao.loc[indice, coluna] = percentual(valor)
            elif nome_indice in indices_monetarios:
                for coluna in exibicao.columns:
                    valor = exibicao.loc[indice, coluna]
                    if isinstance(valor, (int, float, np.number)) and not pd.isna(valor):
                        exibicao.loc[indice, coluna] = moeda_real(valor)
            elif nome_indice in indices_inteiros:
                for coluna in exibicao.columns:
                    valor = exibicao.loc[indice, coluna]
                    if isinstance(valor, (int, float, np.number)) and not pd.isna(valor):
                        exibicao.loc[indice, coluna] = numero_inteiro(valor)

    # Tabelas convencionais com o indicador no nome da coluna.
    for coluna in exibicao.columns:
        if not pd.api.types.is_numeric_dtype(exibicao[coluna]):
            continue
        if _coluna_percentual(coluna):
            exibicao[coluna] = exibicao[coluna].map(percentual)
        elif _coluna_monetaria(coluna):
            exibicao[coluna] = exibicao[coluna].map(moeda_real)
        elif _coluna_inteira(coluna):
            exibicao[coluna] = exibicao[coluna].map(numero_inteiro)

    try:
        styler = exibicao.style
        for coluna in dados.columns:
            if coluna not in exibicao.columns:
                continue
            if not pd.api.types.is_numeric_dtype(dados[coluna]):
                continue
            valores_originais = pd.to_numeric(
                dados[coluna], errors="coerce"
            ).reset_index(drop=True)
            styler = styler.apply(
                lambda serie, valores=valores_originais: [
                    "color:#ff4d4f;font-weight:700"
                    if (
                        i < len(valores)
                        and pd.notna(valores.iloc[i])
                        and float(valores.iloc[i]) < 0
                    )
                    else ""
                    for i in range(len(serie))
                ],
                subset=[coluna],
            )
        return st.dataframe(styler, *args, **kwargs)
    except Exception:
        return st.dataframe(exibicao, *args, **kwargs)


def preparar_tabela(df):
    """Prepara tabelas de resultado no padrão visual brasileiro."""
    out = df.copy()
    for col in out.columns:
        if col in ["Comprador", "KPI"]:
            continue
        if not pd.api.types.is_numeric_dtype(out[col]):
            continue
        if _coluna_percentual(col):
            out[col] = out[col].map(percentual)
        elif _coluna_monetaria(col):
            out[col] = out[col].map(moeda_real)
        elif _coluna_inteira(col):
            out[col] = out[col].map(numero_inteiro)
        else:
            out[col] = out[col].map(lambda v: numero_decimal(v, 2))
    return out


def plotly_chart_br(fig, *args, **kwargs):
    """Exibe gráfico com padrão brasileiro em todos os elementos."""
    registrar_grafico_exportacao(fig)
    tipo = kwargs.pop("tipo", None)
    fig = aplicar_formato_grafico(fig, tipo=tipo)
    config = dict(kwargs.pop("config", {}) or {})
    config.setdefault("displayModeBar", False)
    config.setdefault("locale", "pt-BR")
    return st.plotly_chart(fig, *args, config=config, **kwargs)



def html_valor_negativo(valor, texto):
    try:
        negativo = float(valor) < 0
    except (TypeError, ValueError):
        negativo = False
    if negativo:
        return (
            '<span class="valor-negativo">'
            + str(texto)
            + '</span>'
        )
    return str(texto)


def card_meta(titulo, linhas, destaque=False):
    for rotulo, valor in linhas:
        registrar_card_exportacao(f"{titulo} - {rotulo}", valor)
    classe = "meta-card premium" if destaque else "meta-card"
    html = [f"<div class='{classe}'><div class='meta-card-title'>{titulo}</div>"]
    for rotulo, valor in linhas:
        html.append(f"<div class='meta-line'><span>{rotulo}</span><strong>{valor}</strong></div>")
    html.append("</div>")
    st.markdown("".join(html), unsafe_allow_html=True)

def section(texto, cls):
    st.markdown(f"<div class='section-title {cls}'>{texto}</div>", unsafe_allow_html=True)


# =========================================================
# CONFIGURAÇÃO, PERÍODO E HISTÓRICO DAS METAS
# =========================================================

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
LOG_DIR = DATA_DIR / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

METAS_FILE = DATA_DIR / "metas_gestor.json"
HISTORICO_FILE = DATA_DIR / "historico_metas.json"

METAS_PADRAO = {
    "id_meta": "META-2026-07",
    "periodo_referencia": "2026-05",
    "data_inicio": "2026-07-01",
    "data_fim": "2026-07-31",
    "descricao": "Metas comerciais do período",
    "status": "Ativa",
    "meta_venda_total_mes": 4050000.00,
    "meta_cmv_mes": 67.40,
    "fator_reducao_cmv": 0.01,
    "fator_cobertura": 1.50,
    "meta_ruptura": 5.00,
    "meta_reposicao": 99.00,
    "curva_a": 20.00,
    "curva_b": 30.00,
    "curva_c": 40.00,
    "curva_d": 10.00,
    "rep_paulo": 27.00,
    "rep_francieli": 24.00,
    "rep_sebastiao": 49.00,
    "peso_faturamento": 3.00,
    "peso_cmv": 3.00,
    "peso_fator_cobertura": 10.00,
    "peso_curva_a": 15.00,
    "peso_curva_b": 15.00,
    "peso_curva_c": 10.00,
    "peso_curva_d": 10.00,
    "peso_ruptura": 20.00,
    "peso_reposicao": 14.00,
    "valor_premio_total": 3000.00,
    "usuario_cadastro": "Gestor",
    "data_cadastro": "",
    "ultima_atualizacao": "",
}

@st.cache_data(show_spinner=False)
def carregar_historico():
    if HISTORICO_FILE.exists():
        try:
            dados = json.loads(HISTORICO_FILE.read_text(encoding="utf-8"))
            return dados if isinstance(dados, list) else []
        except Exception:
            return []
    return []

def salvar_historico(lista):
    HISTORICO_FILE.write_text(
        json.dumps(lista, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

@st.cache_data(show_spinner=False)
def carregar_metas():
    if METAS_FILE.exists():
        try:
            dados = json.loads(METAS_FILE.read_text(encoding="utf-8"))
            return {**METAS_PADRAO, **dados}
        except Exception:
            return METAS_PADRAO.copy()
    return METAS_PADRAO.copy()

def salvar_metas(dados, registrar_historico=True):
    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    if not dados.get("data_cadastro"):
        dados["data_cadastro"] = agora
    dados["ultima_atualizacao"] = agora

    METAS_FILE.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    if registrar_historico:
        historico = carregar_historico()
        historico = [h for h in historico if h.get("id_meta") != dados.get("id_meta")]
        historico.append(dados.copy())
        historico = sorted(historico, key=lambda x: x.get("periodo_referencia", ""), reverse=True)
        salvar_historico(historico)
    _limpar_cache_dados()


def data_br(valor):
    try:
        return datetime.strptime(valor, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return valor or "-"

METAS_GESTOR = carregar_metas()
HISTORICO_METAS = carregar_historico()


# =========================================================
# METAS DE LOJA: FATURAMENTO E MARGEM BRUTA
# =========================================================

METAS_LOJAS_FILE = DATA_DIR / "metas_lojas.json"
HISTORICO_METAS_LOJAS_FILE = DATA_DIR / "historico_metas_lojas.json"

METAS_LOJAS_PADRAO = [
    {
        "periodo_referencia": "2026-05",
        "regional_loja": "Filial 1 (Matriz)",
        "gerente": "Fábio",
        "meta_mes": 508547.37,
        "meta_margem_bruta_valor": 172906.11,
        "meta_margem_bruta_pct": 34.0,
        "representatividade_entrega_pct": 25.4,
        "representatividade_entrega_valor": 129251.71,
        "status": "Ativa",
    },
    {
        "periodo_referencia": "2026-05",
        "regional_loja": "Filial 3",
        "gerente": "Lanila",
        "meta_mes": 914607.20,
        "meta_margem_bruta_valor": 310966.45,
        "meta_margem_bruta_pct": 34.0,
        "representatividade_entrega_pct": 4.7,
        "representatividade_entrega_valor": 42718.64,
        "status": "Ativa",
    },
]

def carregar_metas_lojas():
    if METAS_LOJAS_FILE.exists():
        try:
            dados = json.loads(METAS_LOJAS_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, list):
                return dados
        except Exception:
            pass
    METAS_LOJAS_FILE.write_text(
        json.dumps(METAS_LOJAS_PADRAO, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return [dict(x) for x in METAS_LOJAS_PADRAO]

def carregar_historico_metas_lojas():
    if HISTORICO_METAS_LOJAS_FILE.exists():
        try:
            dados = json.loads(HISTORICO_METAS_LOJAS_FILE.read_text(encoding="utf-8"))
            return dados if isinstance(dados, list) else []
        except Exception:
            return []
    return []


def _replicar_lista_metas_maio_2026(arquivo, campo_chave):
    """Cria metas ausentes de 2026 usando maio/2026 como modelo.

    Nunca sobrescreve uma competência já existente.
    """
    if not arquivo.exists():
        return 0

    try:
        conteudo = json.loads(arquivo.read_text(encoding="utf-8"))
    except Exception:
        return 0

    lista = conteudo.get("metas", []) if isinstance(conteudo, dict) else conteudo
    if not isinstance(lista, list):
        return 0

    def periodo(item):
        return str(
            item.get("periodo_referencia")
            or item.get("competencia")
            or item.get("periodo")
            or ""
        )[:7]

    origem = [item for item in lista if periodo(item) == "2026-05"]
    if not origem:
        return 0

    existentes = {
        (periodo(item), str(item.get(campo_chave, "")).strip().upper())
        for item in lista
    }

    criadas = 0
    for mes in range(1, 13):
        competencia = f"2026-{mes:02d}"
        for modelo in origem:
            chave = (
                competencia,
                str(modelo.get(campo_chave, "")).strip().upper(),
            )
            if chave in existentes:
                continue

            nova = dict(modelo)
            nova["periodo_referencia"] = competencia
            nova["competencia"] = competencia

            primeiro_dia = date(2026, mes, 1)
            proximo = date(2027, 1, 1) if mes == 12 else date(2026, mes + 1, 1)
            ultimo_dia = proximo - timedelta(days=1)
            nova["data_inicio"] = primeiro_dia.isoformat()
            nova["data_fim"] = ultimo_dia.isoformat()
            nova["origem"] = "Replicada da meta de maio/2026"
            nova["atualizado_em"] = datetime.now().isoformat(timespec="seconds")

            lista.append(nova)
            existentes.add(chave)
            criadas += 1

    if criadas:
        if isinstance(conteudo, dict):
            conteudo["metas"] = lista
            conteudo["ultima_replicacao_maio_2026"] = datetime.now().isoformat(
                timespec="seconds"
            )
            saida = conteudo
        else:
            saida = lista

        arquivo.write_text(
            json.dumps(saida, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    return criadas


def preencher_metas_2026_com_maio():
    lojas = _replicar_lista_metas_maio_2026(
        DATA_DIR / "metas_lojas.json", "loja"
    )
    compradores = _replicar_lista_metas_maio_2026(
        DATA_DIR / "metas_compradores.json", "comprador"
    )
    return lojas, compradores


# Preenchimento inicial sem sobrescrever alterações já realizadas.
try:
    preencher_metas_2026_com_maio()
except Exception:
    pass


def salvar_metas_lojas(lista, usuario="Gestor"):
    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    dados_limpos = []
    for item in lista:
        registro = dict(item)
        registro["ultima_atualizacao"] = agora
        registro["usuario_atualizacao"] = usuario or "Gestor"
        dados_limpos.append(registro)
    METAS_LOJAS_FILE.write_text(
        json.dumps(dados_limpos, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    historico = carregar_historico_metas_lojas()
    historico.append({
        "data_hora": agora,
        "usuario": usuario or "Gestor",
        "registros": dados_limpos,
    })
    HISTORICO_METAS_LOJAS_FILE.write_text(
        json.dumps(historico[-100:], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

def dataframe_metas_lojas(periodo=None):
    dados = carregar_metas_lojas()
    df = pd.DataFrame(dados)
    colunas = [
        "periodo_referencia", "regional_loja", "gerente", "meta_mes",
        "meta_margem_bruta_valor", "meta_margem_bruta_pct",
        "representatividade_entrega_pct", "representatividade_entrega_valor", "status"
    ]
    for col in colunas:
        if col not in df.columns:
            df[col] = "" if col in ["periodo_referencia", "regional_loja", "gerente", "status"] else 0.0
    if periodo and "periodo_referencia" in df.columns:
        filtrado = df[df["periodo_referencia"].astype(str) == str(periodo)].copy()
        if not filtrado.empty:
            df = filtrado
    return df[colunas].copy()



@st.cache_data(ttl=600, show_spinner=False, max_entries=24)
def carregar_realizado_filiais_ceo(periodo, token_cache):
    """Agrega faturamento e margem bruta por filial diretamente no SQLite."""
    colunas = ["numero_loja", "loja", "faturamento_atual", "margem_bruta_atual"]
    try:
        with sqlite3.connect(CACHE_DB_FILE, timeout=30) as con:
            tabelas = {linha[0] for linha in con.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()}
            if "base_vendas" not in tabelas:
                return pd.DataFrame(columns=colunas)
            estrutura = {linha[1] for linha in con.execute(
                'PRAGMA table_info("base_vendas")'
            ).fetchall()}
            if "periodo_referencia" not in estrutura:
                return pd.DataFrame(columns=colunas)
            expr_venda = "COALESCE(CAST(valortotal AS REAL), 0)" if "valortotal" in estrutura else "0"
            if "lucro" in estrutura:
                expr_lucro = "COALESCE(CAST(lucro AS REAL), 0)"
            elif "custo" in estrutura and "valortotal" in estrutura:
                expr_lucro = "COALESCE(CAST(valortotal AS REAL), 0) - COALESCE(CAST(custo AS REAL), 0)"
            else:
                expr_lucro = "0"
            numero = "COALESCE(CAST(numero_loja AS TEXT), '')" if "numero_loja" in estrutura else "''"
            nome = "COALESCE(CAST(loja AS TEXT), '')" if "loja" in estrutura else "''"
            consulta = f"""
                SELECT {numero} AS numero_loja, {nome} AS loja,
                       SUM({expr_venda}) AS faturamento_atual,
                       SUM({expr_lucro}) AS margem_bruta_atual
                FROM base_vendas
                WHERE periodo_referencia = ?
                GROUP BY {numero}, {nome}
                ORDER BY {numero}, {nome}
            """
            return pd.read_sql_query(consulta, con, params=(str(periodo),))
    except Exception:
        return pd.DataFrame(columns=colunas)


def montar_quadro_filiais_ceo(periodo):
    """Cruza metas cadastradas por filial com o realizado da base de vendas."""
    metas_filiais = dataframe_metas_lojas(periodo).copy()
    if metas_filiais.empty:
        return pd.DataFrame()

    realizado = carregar_realizado_filiais_ceo(periodo, _arquivo_token(CACHE_DB_FILE)).copy()

    def chave_numero(valor):
        numeros = re.findall(r"\d+", str(valor or "").strip())
        return str(int(numeros[-1])).zfill(2) if numeros else ""

    metas_filiais["chave_filial"] = metas_filiais["regional_loja"].map(chave_numero)
    if realizado.empty:
        realizado = pd.DataFrame(columns=["numero_loja", "loja", "faturamento_atual", "margem_bruta_atual"])
    realizado["chave_filial"] = realizado["numero_loja"].map(chave_numero)

    mapa_romano = {" I":"01", " II":"02", " III":"03", " IV":"04", " V":"05", " VI":"06", " VII":"07", " VIII":"08", " IX":"09", " X":"10"}
    for indice in realizado[realizado["chave_filial"].eq("")].index:
        nome_loja = " " + str(realizado.at[indice, "loja"]).strip().upper()
        for romano, numero in sorted(mapa_romano.items(), key=lambda item: len(item[0]), reverse=True):
            if nome_loja.endswith(romano):
                realizado.at[indice, "chave_filial"] = numero
                break

    realizado = realizado[["chave_filial", "loja", "faturamento_atual", "margem_bruta_atual"]].copy()
    quadro = metas_filiais.merge(realizado, on="chave_filial", how="left")
    for coluna in ["meta_mes", "meta_margem_bruta_valor", "meta_margem_bruta_pct", "faturamento_atual", "margem_bruta_atual"]:
        quadro[coluna] = pd.to_numeric(quadro.get(coluna, 0), errors="coerce").fillna(0.0)

    quadro["Atingimento Faturamento (%)"] = quadro.apply(
        lambda linha: linha["faturamento_atual"] / linha["meta_mes"] * 100 if linha["meta_mes"] else 0.0, axis=1)
    quadro["Atingimento Margem Bruta (%)"] = quadro.apply(
        lambda linha: linha["margem_bruta_atual"] / linha["meta_margem_bruta_valor"] * 100 if linha["meta_margem_bruta_valor"] else 0.0, axis=1)
    quadro["Margem Bruta Atual (%)"] = quadro.apply(
        lambda linha: linha["margem_bruta_atual"] / linha["faturamento_atual"] * 100 if linha["faturamento_atual"] else 0.0, axis=1)

    quadro["Filial"] = quadro["regional_loja"].fillna("")
    quadro["Gerente"] = quadro["gerente"].fillna("")
    quadro["Faturamento Total META"] = quadro["meta_mes"]
    quadro["Faturamento Total Atual"] = quadro["faturamento_atual"]
    quadro["Margem Bruta META"] = quadro["meta_margem_bruta_valor"]
    quadro["Margem Bruta Atual"] = quadro["margem_bruta_atual"]
    quadro["Margem Bruta META (%)"] = quadro["meta_margem_bruta_pct"]
    return quadro[["Filial", "Gerente", "Faturamento Total META", "Faturamento Total Atual", "Atingimento Faturamento (%)", "Margem Bruta META", "Margem Bruta Atual", "Atingimento Margem Bruta (%)", "Margem Bruta META (%)", "Margem Bruta Atual (%)"]].sort_values(["Filial", "Gerente"]).reset_index(drop=True)


# =========================================================
# IMPORTAÇÃO TEMPORÁRIA DA RUPTURA
# =========================================================

RUPTURA_FILE = DATA_DIR / "ruptura_importada.csv"
RUPTURA_META_FILE = DATA_DIR / "ruptura_importacao_meta.json"

COLUNAS_RUPTURA_PADRAO = [
    "Loja",
    "Comprador",
    "Classificação 3º Nível",
    "Código Interno",
    "EAN",
    "Produto",
    "Ruptura Ativa",
    "Data Referência",
]

def normalizar_nome_coluna(nome):
    return (
        str(nome).strip()
        .replace("\n", " ")
        .replace("  ", " ")
    )

def detectar_coluna(df, alternativas):
    mapa = {normalizar_nome_coluna(c).lower(): c for c in df.columns}
    for alternativa in alternativas:
        chave = alternativa.lower()
        if chave in mapa:
            return mapa[chave]
    return None

def preparar_ruptura_importada(df):
    df = df.copy()
    df.columns = [normalizar_nome_coluna(c) for c in df.columns]

    col_loja = detectar_coluna(df, ["loja", "numero_loja", "filial"])
    col_comprador = detectar_coluna(df, ["comprador", "comprador responsável", "comprador responsavel"])
    col_classificacao = detectar_coluna(df, ["classificação 3º nível", "classificacao 3º nivel", "classificacao 3 nivel", "classificação"])
    col_codigo = detectar_coluna(df, ["código interno", "codigo interno", "cod_interno", "produtoid"])
    col_ean = detectar_coluna(df, ["ean", "código de barras", "codigo de barras", "codigobarras"])
    col_produto = detectar_coluna(df, ["produto", "descrição", "descricao"])
    col_ruptura = detectar_coluna(df, ["ruptura ativa", "ruptura", "valor ruptura", "ruptura_r$"])
    col_data = detectar_coluna(df, ["data referência", "data referencia", "data", "competência", "competencia"])

    if col_ruptura is None:
        raise ValueError(
            "Não foi encontrada a coluna obrigatória de ruptura. "
            "Use uma coluna chamada 'Ruptura Ativa' ou 'Ruptura'."
        )

    saida = pd.DataFrame()
    saida["Loja"] = df[col_loja] if col_loja else ""
    saida["Comprador"] = df[col_comprador] if col_comprador else ""
    saida["Classificação 3º Nível"] = df[col_classificacao] if col_classificacao else ""
    saida["Código Interno"] = df[col_codigo] if col_codigo else ""
    saida["EAN"] = df[col_ean] if col_ean else ""
    saida["Produto"] = df[col_produto] if col_produto else ""
    saida["Ruptura Ativa"] = pd.to_numeric(
        df[col_ruptura].astype(str)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False),
        errors="coerce"
    ).fillna(0)

    if col_data:
        saida["Data Referência"] = pd.to_datetime(df[col_data], errors="coerce", dayfirst=True)
    else:
        saida["Data Referência"] = pd.Timestamp.today().normalize()

    saida["Data Referência"] = saida["Data Referência"].dt.strftime("%Y-%m-%d")
    return saida

def carregar_ruptura_importada():
    if RUPTURA_FILE.exists():
        try:
            return pd.read_csv(RUPTURA_FILE, sep=";", encoding="utf-8-sig")
        except Exception:
            return pd.DataFrame(columns=COLUNAS_RUPTURA_PADRAO)
    return pd.DataFrame(columns=COLUNAS_RUPTURA_PADRAO)

def salvar_ruptura_importada(df, nome_arquivo):
    df.to_csv(RUPTURA_FILE, sep=";", index=False, encoding="utf-8-sig")
    meta = {
        "arquivo_origem": nome_arquivo,
        "data_importacao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "linhas": int(len(df)),
        "valor_total_ruptura": float(df["Ruptura Ativa"].sum()),
    }
    RUPTURA_META_FILE.write_text(
        json.dumps(meta, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

def carregar_meta_ruptura():
    if RUPTURA_META_FILE.exists():
        try:
            return json.loads(RUPTURA_META_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

RUPTURA_IMPORTADA = carregar_ruptura_importada()
META_RUPTURA_IMPORTADA = carregar_meta_ruptura()


def carregar_logo_economize():
    caminho = Path("assets/logo_rede_economize.png")
    if caminho.exists():
        return base64.b64encode(caminho.read_bytes()).decode("utf-8")
    return ""

LOGO_ECONOMIZE_B64 = carregar_logo_economize()


# =========================================================
# BANCO DE DADOS E ATUALIZAÇÕES MENSAIS
# =========================================================

CONFIG_DIR = Path("config")
SQL_DIR = Path("sql")
DATA_DIR = Path("data")
CONFIG_DIR.mkdir(exist_ok=True)
SQL_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)

DB_CONFIG_FILE = CONFIG_DIR / "database.json"
CACHE_DB_FILE = DATA_DIR / "kpis_mensal.sqlite"
ANALISE_COMERCIAL_CONFIG_FILE = CONFIG_DIR / "analise_comercial.json"
PLANO_CONTAS_PAGAMENTO_PADRAO = "Resultado > 03.1 - Despesas Operacionais > 2-CUSTOS VARIAVEIS > 1-C.M.V / DUPL. PAGAS"
PLANOS_CONTAS_CATALOGO_FILE = CONFIG_DIR / "planos_contas_catalogo.json"

@st.cache_data(show_spinner=False)
def carregar_config_analise_comercial():
    padrao = {
        "plano_contas_padrao": PLANO_CONTAS_PAGAMENTO_PADRAO,
        "planos_contas_selecionados": [PLANO_CONTAS_PAGAMENTO_PADRAO],
        "planos_adicionais": [],
    }
    if ANALISE_COMERCIAL_CONFIG_FILE.exists():
        try:
            dados = json.loads(ANALISE_COMERCIAL_CONFIG_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, dict):
                padrao.update(dados)
        except Exception:
            pass
    return padrao

def salvar_config_analise_comercial(dados):
    ANALISE_COMERCIAL_CONFIG_FILE.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    _limpar_cache_dados()

@st.cache_data(show_spinner=False)
def carregar_catalogo_planos_contas():
    if PLANOS_CONTAS_CATALOGO_FILE.exists():
        try:
            dados = json.loads(
                PLANOS_CONTAS_CATALOGO_FILE.read_text(encoding="utf-8")
            )
            planos = dados.get("planos", []) if isinstance(dados, dict) else []
            return sorted({
                str(plano).strip()
                for plano in planos
                if str(plano).strip()
            })
        except Exception:
            pass
    return []


def _numero_br(valor):
    """Converte valores brasileiros ou numéricos para float."""
    if valor is None or pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float, np.number)):
        return float(valor)
    texto = str(valor).strip().replace("R$", "").strip()
    if not texto:
        return 0.0
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except Exception:
        return 0.0


def normalizar_contas_pagar_df(df):
    """Adapta a saída oficial do DBeaver ao formato interno do projeto."""
    base = df.copy()
    mapa = {normalizar_nome_coluna(c).casefold(): c for c in base.columns}

    def coluna(*nomes):
        for nome in nomes:
            achada = mapa.get(normalizar_nome_coluna(nome).casefold())
            if achada is not None:
                return achada
        return None

    col_status = coluna("Status", "status")
    col_venc = coluna("Data de Vencimento", "data_vencimento")
    col_pag = coluna("Data Pagamento", "data_pagamento")
    col_valor = coluna("Valor Documento", "valor_documento", "Valor")
    col_fornecedor = coluna("Credor", "fornecedor")
    col_unidade = coluna("Unidade", "unidade")
    col_apelido = coluna("Apelido Un. Neg.", "apelido_unidade")
    col_plano = coluna("Plano de Contas", "plano_contas")
    col_doc = coluna("Número Documento", "numero_documento")

    saida = pd.DataFrame(index=base.index)
    saida["status"] = base[col_status].astype(str) if col_status else ""
    saida["data_vencimento"] = (
        pd.to_datetime(base[col_venc], errors="coerce", dayfirst=True)
        if col_venc else pd.NaT
    )
    saida["data_pagamento"] = (
        pd.to_datetime(base[col_pag], errors="coerce", dayfirst=True)
        if col_pag else pd.NaT
    )
    saida["valor_documento"] = (
        base[col_valor].map(_numero_br) if col_valor else 0.0
    )
    saida["fornecedor"] = base[col_fornecedor].astype(str) if col_fornecedor else ""
    saida["unidade"] = base[col_unidade].astype(str) if col_unidade else ""
    saida["apelido_unidade"] = base[col_apelido].astype(str) if col_apelido else ""
    saida["plano_contas"] = base[col_plano].astype(str).str.strip() if col_plano else ""
    saida["numero_documento"] = base[col_doc].astype(str) if col_doc else ""

    paga = saida["data_pagamento"].notna() | saida["status"].str.casefold().eq("paga")
    pendente = saida["status"].str.casefold().eq("pendente")
    saida["valor_pago"] = saida["valor_documento"].where(paga, 0.0)
    saida["saldo_aberto"] = saida["valor_documento"].where(pendente, 0.0)

    saida["data_vencimento"] = saida["data_vencimento"].dt.strftime("%Y-%m-%d")
    saida["data_pagamento"] = saida["data_pagamento"].dt.strftime("%Y-%m-%d")
    return saida


FONTES_BANCO = {
    "vendas": {
        "titulo": "Vendas",
        "arquivo_sql": SQL_DIR / "vendas.sql",
        "tabela_cache": "base_vendas",
    },
    "estoque": {
        "titulo": "Estoque",
        "arquivo_sql": SQL_DIR / "estoque.sql",
        "tabela_cache": "base_estoque",
    },
    "entradas": {
        "titulo": "Entradas",
        "arquivo_sql": SQL_DIR / "entradas.sql",
        "tabela_cache": "base_entradas",
    },
    "contas_pagar": {
        "titulo": "Contas a Pagar",
        "arquivo_sql": SQL_DIR / "contas_pagar.sql",
        "tabela_cache": "base_contas_pagar",
    },
}

DB_CONFIG_PADRAO = {
    "tipo": "PostgreSQL",
    "host": "",
    "porta": 5432,
    "banco": "",
    "usuario": "",
    "senha": "",
    "sslmode": "prefer",
    "salvar_senha": True,
    "ultima_validacao": "",
}

@st.cache_data(show_spinner=False)
def carregar_config_banco():
    if DB_CONFIG_FILE.exists():
        try:
            dados = json.loads(DB_CONFIG_FILE.read_text(encoding="utf-8"))
            return {**DB_CONFIG_PADRAO, **dados}
        except Exception:
            pass
    return DB_CONFIG_PADRAO.copy()

def salvar_config_banco(dados):
    DB_CONFIG_FILE.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

def montar_url_banco(cfg):
    usuario = quote_plus(str(cfg.get("usuario", "")))
    senha = quote_plus(str(cfg.get("senha", "")))
    host = cfg.get("host", "")
    porta = int(cfg.get("porta", 5432))
    banco = cfg.get("banco", "")
    sslmode = cfg.get("sslmode", "prefer")
    return f"postgresql+psycopg2://{usuario}:{senha}@{host}:{porta}/{banco}?sslmode={sslmode}"

def criar_engine_banco(cfg):
    return create_engine(
        montar_url_banco(cfg),
        pool_pre_ping=True,
        pool_recycle=1800,
        connect_args={"connect_timeout": 15},
    )

def garantir_sqls():
    exemplos = {
        "vendas.sql": """-- Use os parâmetros :data_inicio e :data_fim
-- Cole aqui o SQL validado de vendas.
-- Exemplo de filtro:
-- WHERE datahora_venda_final::date BETWEEN :data_inicio AND :data_fim
SELECT
    NULL::text AS loja,
    NULL::text AS classificacao_3_nivel,
    NULL::numeric AS valor_venda
WHERE 1 = 0;
""",
        "estoque.sql": """-- Use os parâmetros :data_inicio e :data_fim
-- Cole aqui o SQL validado de estoque.
-- Para estoque atual, os parâmetros podem ser ignorados.
SELECT
    NULL::text AS loja,
    NULL::text AS classificacao_3_nivel,
    NULL::numeric AS valor_estoque
WHERE 1 = 0;
""",
        "entradas.sql": """-- Use os parâmetros :data_inicio e :data_fim
-- Cole aqui o SQL validado de entradas.
-- Exemplo de filtro:
-- WHERE datahoraentrada::date BETWEEN :data_inicio AND :data_fim
SELECT
    NULL::text AS loja,
    NULL::text AS classificacao_3_nivel,
    NULL::numeric AS valor_entrada
WHERE 1 = 0;
""",
        "contas_pagar.sql": """-- A consulta oficial de Contas a Pagar
-- é armazenada em sql/contas_pagar.sql.
-- Use os parâmetros :data_inicio e :data_fim.
""",
    }
    for nome, conteudo in exemplos.items():
        caminho = SQL_DIR / nome
        if not caminho.exists() or not caminho.read_text(encoding="utf-8").strip():
            caminho.write_text(conteudo, encoding="utf-8")

def ler_sql(caminho):
    if caminho.exists():
        return caminho.read_text(encoding="utf-8")
    return ""


def diagnosticar_sql_fonte(fonte, sql):
    """Valida se a fonte possui uma consulta real antes da execução."""
    texto_sql = str(sql or "").strip()
    normalizado = re.sub(r"\s+", " ", texto_sql).casefold()

    if not texto_sql:
        return False, "O arquivo SQL está vazio."

    marcadores_modelo = [
        "where 1 = 0",
        "cole aqui o sql",
        "null::text as unidade",
        "null::numeric as valor_documento",
    ]
    if any(marcador in normalizado for marcador in marcadores_modelo):
        if fonte == "contas_pagar":
            return (
                False,
                "A consulta de Contas a Pagar ainda é apenas um modelo. "
                "Cole o SQL oficial no menu Banco de Dados > Editar SQL."
            )
        return False, "A consulta desta fonte ainda é apenas um modelo."

    if fonte == "contas_pagar":
        if "plano" not in normalizado or "valor" not in normalizado:
            return (
                False,
                "O SQL de Contas a Pagar deve retornar o Plano de Contas "
                "e o Valor Documento."
            )

    return True, "SQL configurado."


def salvar_sql(caminho, conteudo):
    caminho.write_text(conteudo.strip() + "\n", encoding="utf-8")

def conexao_cache():
    con = sqlite3.connect(CACHE_DB_FILE, timeout=30)
    try:
        con.execute("PRAGMA journal_mode=WAL")
        con.execute("PRAGMA synchronous=NORMAL")
        con.execute("PRAGMA temp_store=MEMORY")
        con.execute("PRAGMA cache_size=-64000")
        con.execute("PRAGMA busy_timeout=30000")
        con.execute("PRAGMA mmap_size=268435456")
    except Exception:
        pass
    con.execute("""
        CREATE TABLE IF NOT EXISTS atualizacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fonte TEXT NOT NULL,
            periodo_referencia TEXT NOT NULL,
            data_inicio TEXT NOT NULL,
            data_fim TEXT NOT NULL,
            registros INTEGER NOT NULL,
            status TEXT NOT NULL,
            mensagem TEXT,
            atualizado_em TEXT NOT NULL
        )
    """)
    con.commit()
    return con


def reconstruir_posicoes_mensais(con, periodo=None):
    """Recalcula posições mensais preservando todas as bases existentes."""
    garantir_tabelas_analise(con)
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    existe_cp = con.execute(
        "SELECT 1 FROM sqlite_master "
        "WHERE type='table' AND name='base_contas_pagar'"
    ).fetchone()

    if existe_cp:
        filtro_cp = " WHERE periodo_referencia = ?" if periodo else ""
        params_cp = (periodo,) if periodo else ()
        linhas_cp = con.execute(
            """
            SELECT
                periodo_referencia,
                SUM(COALESCE(CAST(valor_documento AS REAL), 0))
            FROM base_contas_pagar
            """
            + filtro_cp
            + """
            GROUP BY periodo_referencia
            """,
            params_cp,
        ).fetchall()

        for competencia, valor_cp in linhas_cp:
            existe = con.execute(
                """
                SELECT 1
                FROM analise_posicao_resumo
                WHERE periodo_referencia = ?
                """,
                (competencia,),
            ).fetchone()

            if existe:
                con.execute(
                    """
                    UPDATE analise_posicao_resumo
                    SET contas_pagar = ?,
                        origem = 'Reconstrução',
                        atualizado_em = ?
                    WHERE periodo_referencia = ?
                    """,
                    (float(valor_cp or 0), agora, competencia),
                )
            else:
                con.execute(
                    """
                    INSERT INTO analise_posicao_resumo
                        (
                            periodo_referencia,
                            contas_pagar,
                            estoque,
                            origem,
                            atualizado_em
                        )
                    VALUES (?, ?, 0, 'Reconstrução', ?)
                    """,
                    (competencia, float(valor_cp or 0), agora),
                )

    existe_estoque = con.execute(
        "SELECT 1 FROM sqlite_master "
        "WHERE type='table' AND name='analise_estoque_resumo'"
    ).fetchone()

    if existe_estoque:
        filtro_est = " WHERE periodo_referencia = ?" if periodo else ""
        params_est = (periodo,) if periodo else ()
        linhas_est = con.execute(
            """
            SELECT
                periodo_referencia,
                COALESCE(estoque, 0)
            FROM analise_estoque_resumo
            """
            + filtro_est,
            params_est,
        ).fetchall()

        for competencia, valor_estoque in linhas_est:
            existe = con.execute(
                """
                SELECT 1
                FROM analise_posicao_resumo
                WHERE periodo_referencia = ?
                """,
                (competencia,),
            ).fetchone()

            if existe:
                con.execute(
                    """
                    UPDATE analise_posicao_resumo
                    SET estoque = ?,
                        origem = 'Reconstrução',
                        atualizado_em = ?
                    WHERE periodo_referencia = ?
                    """,
                    (float(valor_estoque or 0), agora, competencia),
                )
            else:
                con.execute(
                    """
                    INSERT INTO analise_posicao_resumo
                        (
                            periodo_referencia,
                            contas_pagar,
                            estoque,
                            origem,
                            atualizado_em
                        )
                    VALUES (?, 0, ?, 'Reconstrução', ?)
                    """,
                    (competencia, float(valor_estoque or 0), agora),
                )

    con.commit()



def garantir_tabelas_analise(con):
    """Cria todas as tabelas de resumo necessárias antes de qualquer atualização."""
    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_vendas_resumo (
            periodo_referencia TEXT NOT NULL,
            classificacao TEXT,
            curva TEXT,
            venda REAL NOT NULL DEFAULT 0,
            custo REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (periodo_referencia, classificacao, curva)
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_entradas_resumo (
            periodo_referencia TEXT NOT NULL,
            classificacao TEXT,
            entrada REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (periodo_referencia, classificacao)
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_contas_resumo (
            periodo_referencia TEXT NOT NULL,
            plano_contas TEXT,
            pagamento REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (periodo_referencia, plano_contas)
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_estoque_resumo (
            periodo_referencia TEXT NOT NULL PRIMARY KEY,
            estoque REAL NOT NULL DEFAULT 0
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_posicao_resumo (
            periodo_referencia TEXT NOT NULL PRIMARY KEY,
            contas_pagar REAL NOT NULL DEFAULT 0,
            estoque REAL NOT NULL DEFAULT 0,
            origem TEXT,
            atualizado_em TEXT
        )
    """)

    con.commit()


def atualizar_resumos_analise(con, fonte=None, periodo=None):
    garantir_tabelas_analise(con)
    """Mantém tabelas pequenas usadas pela Análise Comercial.

    A tela nunca precisa carregar as bases brutas. Os resumos são refeitos
    somente após atualização da fonte ou em uma reconstrução manual.
    """
    filtros_delete = " WHERE periodo_referencia = ?" if periodo else ""
    params = (periodo,) if periodo else ()

    if fonte in (None, "vendas"):
        con.execute("""
            CREATE TABLE IF NOT EXISTS analise_vendas_resumo (
                periodo_referencia TEXT NOT NULL,
                classificacao TEXT,
                venda REAL NOT NULL DEFAULT 0,
                custo REAL NOT NULL DEFAULT 0,
                PRIMARY KEY (periodo_referencia, classificacao)
            )
        """)
        con.execute("DELETE FROM analise_vendas_resumo" + filtros_delete, params)
        where = " WHERE periodo_referencia = ?" if periodo else ""
        con.execute("""
            INSERT OR REPLACE INTO analise_vendas_resumo
                (periodo_referencia, classificacao, venda, custo)
            SELECT
                SUBSTR(
                    COALESCE(
                        NULLIF(TRIM(datahora_venda_final), ''),
                        NULLIF(TRIM(datahora), ''),
                        periodo_referencia
                    ), 1, 7
                ) AS periodo_referencia,
                COALESCE(NULLIF(TRIM(classificacao_3_nivel), ''),
                         NULLIF(TRIM(classificacao_resumida), ''),
                         'SEM CLASSIFICACAO') AS classificacao,
                SUM(COALESCE(CAST(valortotal AS REAL), 0)) AS venda,
                SUM(
                    ABS(COALESCE(CAST(qtd_mov AS REAL), CAST(quantidade AS REAL), 0))
                    * COALESCE(CAST(custo AS REAL), 0)
                ) AS custo
            FROM base_vendas
        """ + (
            " WHERE SUBSTR(COALESCE(NULLIF(TRIM(datahora_venda_final), ''), "
            "NULLIF(TRIM(datahora), ''), periodo_referencia), 1, 7) = ?"
            if periodo else ""
        ) + """
            GROUP BY
                SUBSTR(
                    COALESCE(
                        NULLIF(TRIM(datahora_venda_final), ''),
                        NULLIF(TRIM(datahora), ''),
                        periodo_referencia
                    ), 1, 7
                ),
                     COALESCE(NULLIF(TRIM(classificacao_3_nivel), ''),
                              NULLIF(TRIM(classificacao_resumida), ''),
                              'SEM CLASSIFICACAO')
        """, params)

    if fonte in (None, "entradas"):
        con.execute("""
            CREATE TABLE IF NOT EXISTS analise_entradas_resumo (
                periodo_referencia TEXT NOT NULL,
                classificacao TEXT,
                compra REAL NOT NULL DEFAULT 0,
                PRIMARY KEY (periodo_referencia, classificacao)
            )
        """)
        con.execute("DELETE FROM analise_entradas_resumo" + filtros_delete, params)
        where = " WHERE periodo_referencia = ?" if periodo else ""
        con.execute("""
            INSERT OR REPLACE INTO analise_entradas_resumo
                (periodo_referencia, classificacao, compra)
            SELECT
                SUBSTR(
                    COALESCE(
                        NULLIF(TRIM(data_entrada), ''),
                        NULLIF(TRIM(data_emissao), ''),
                        periodo_referencia
                    ), 1, 7
                ) AS periodo_referencia,
                COALESCE(NULLIF(TRIM(classificacao_3_nivel), ''), 'SEM CLASSIFICACAO'),
                SUM(COALESCE(CAST(entrada_custo_total AS REAL), 0))
            FROM base_entradas
        """ + (
            " WHERE SUBSTR(COALESCE(NULLIF(TRIM(data_entrada), ''), "
            "NULLIF(TRIM(data_emissao), ''), periodo_referencia), 1, 7) = ?"
            if periodo else ""
        ) + """
            GROUP BY
                SUBSTR(
                    COALESCE(
                        NULLIF(TRIM(data_entrada), ''),
                        NULLIF(TRIM(data_emissao), ''),
                        periodo_referencia
                    ), 1, 7
                ),
                     COALESCE(NULLIF(TRIM(classificacao_3_nivel), ''), 'SEM CLASSIFICACAO')
        """, params)

    if fonte in (None, "contas_pagar"):
        con.execute("""
            CREATE TABLE IF NOT EXISTS analise_contas_resumo (
                periodo_referencia TEXT NOT NULL,
                plano_contas TEXT,
                pagamento REAL NOT NULL DEFAULT 0,
                PRIMARY KEY (periodo_referencia, plano_contas)
            )
        """)
        con.execute("DELETE FROM analise_contas_resumo" + filtros_delete, params)

        con.execute("""
            INSERT OR REPLACE INTO analise_contas_resumo
                (periodo_referencia, plano_contas, pagamento)
            SELECT
                SUBSTR(
                    COALESCE(NULLIF(TRIM(data_pagamento), ''), periodo_referencia),
                    1, 7
                ) AS periodo_referencia,
                COALESCE(NULLIF(TRIM(plano_contas), ''), 'SEM PLANO DE CONTAS'),
                SUM(COALESCE(CAST(valor_pago AS REAL), 0))
            FROM base_contas_pagar
        """ + (
            " WHERE SUBSTR(COALESCE(NULLIF(TRIM(data_pagamento), ''), "
            "periodo_referencia), 1, 7) = ? "
            "AND COALESCE(CAST(valor_pago AS REAL), 0) <> 0"
            if periodo else
            " WHERE COALESCE(CAST(valor_pago AS REAL), 0) <> 0"
        ) + """
            GROUP BY
                SUBSTR(
                    COALESCE(NULLIF(TRIM(data_pagamento), ''), periodo_referencia),
                    1, 7
                ),
                COALESCE(NULLIF(TRIM(plano_contas), ''), 'SEM PLANO DE CONTAS')
        """, params)

        con.execute("""
            CREATE TABLE IF NOT EXISTS analise_posicao_resumo (
                periodo_referencia TEXT NOT NULL PRIMARY KEY,
                contas_pagar REAL NOT NULL DEFAULT 0,
                estoque REAL NOT NULL DEFAULT 0,
                origem TEXT,
                atualizado_em TEXT
            )
        """)

        if periodo:
            saldo_contas = con.execute(
                """
                SELECT SUM(COALESCE(CAST(valor_documento AS REAL), 0))
                FROM base_contas_pagar
                WHERE periodo_referencia = ?
                """,
                (periodo,),
            ).fetchone()[0] or 0

            con.execute(
                """
                INSERT INTO analise_posicao_resumo
                    (periodo_referencia, contas_pagar, estoque, origem, atualizado_em)
                VALUES (?, ?, 0, 'Banco de Dados', ?)
                ON CONFLICT(periodo_referencia) DO UPDATE SET
                    contas_pagar = excluded.contas_pagar,
                    origem = excluded.origem,
                    atualizado_em = excluded.atualizado_em
                """,
                (
                    periodo,
                    float(saldo_contas),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )

    if fonte in (None, "estoque"):
        con.execute("""
            CREATE TABLE IF NOT EXISTS analise_estoque_resumo (
                periodo_referencia TEXT NOT NULL PRIMARY KEY,
                estoque REAL NOT NULL DEFAULT 0
            )
        """)
        con.execute("DELETE FROM analise_estoque_resumo" + filtros_delete, params)

        existe_estoque = con.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='base_estoque'"
        ).fetchone()
        if existe_estoque:
            colunas = {
                str(linha[1]).strip().lower(): str(linha[1])
                for linha in con.execute("PRAGMA table_info(base_estoque)").fetchall()
            }
            col_periodo = colunas.get("periodo_referencia")
            col_qtd = next(
                (colunas[c] for c in [
                    "estoque", "quantidade estoque", "saldo estoque", "qtd estoque"
                ] if c in colunas),
                None
            )
            col_custo_medio = next(
                (colunas[c] for c in [
                    "custo_medio_atual", "custo medio atual",
                    "custo médio atual", "customedio", "custo medio"
                ] if c in colunas),
                None
            )
            col_custo_unit = next(
                (colunas[c] for c in [
                    "custo_unit_atual", "custo unit atual",
                    "custo unitario atual", "custo", "custo_final_r"
                ] if c in colunas),
                None
            )
            col_valor = next(
                (colunas[c] for c in [
                    "valor_estoque", "valor estoque", "estoque total",
                    "estoque a custo", "estoque x custo medio"
                ] if c in colunas),
                None
            )

            if col_periodo and (col_valor or (col_qtd and (col_custo_medio or col_custo_unit))):
                where_est = f' WHERE "{col_periodo}" = ?' if periodo else ""
                if col_valor:
                    expr_valor = f'COALESCE(CAST("{col_valor}" AS REAL), 0)'
                else:
                    if col_custo_medio and col_custo_unit:
                        expr_custo = (
                            f'CASE WHEN COALESCE(CAST("{col_custo_medio}" AS REAL),0) > 0 '
                            f'THEN CAST("{col_custo_medio}" AS REAL) '
                            f'ELSE COALESCE(CAST("{col_custo_unit}" AS REAL),0) END'
                        )
                    elif col_custo_medio:
                        expr_custo = f'COALESCE(CAST("{col_custo_medio}" AS REAL),0)'
                    else:
                        expr_custo = f'COALESCE(CAST("{col_custo_unit}" AS REAL),0)'
                    expr_valor = (
                        f'COALESCE(CAST("{col_qtd}" AS REAL),0) * ({expr_custo})'
                    )

                con.execute(
                    f"""
                    INSERT OR REPLACE INTO analise_estoque_resumo
                        (periodo_referencia, estoque)
                    SELECT
                        "{col_periodo}",
                        SUM({expr_valor})
                    FROM base_estoque
                    {where_est}
                    GROUP BY "{col_periodo}"
                    """,
                    params
                )

                con.execute("""
                    CREATE TABLE IF NOT EXISTS analise_posicao_resumo (
                        periodo_referencia TEXT NOT NULL PRIMARY KEY,
                        contas_pagar REAL NOT NULL DEFAULT 0,
                        estoque REAL NOT NULL DEFAULT 0,
                        origem TEXT,
                        atualizado_em TEXT
                    )
                """)

                if periodo:
                    linha_estoque = con.execute(
                        """
                        SELECT COALESCE(estoque, 0)
                        FROM analise_estoque_resumo
                        WHERE periodo_referencia = ?
                        """,
                        (periodo,),
                    ).fetchone()
                    valor_estoque_posicao = (
                        float(linha_estoque[0])
                        if linha_estoque is not None
                        else 0.0
                    )
                    agora_posicao = datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    )
                    existe_posicao = con.execute(
                        """
                        SELECT 1
                        FROM analise_posicao_resumo
                        WHERE periodo_referencia = ?
                        """,
                        (periodo,),
                    ).fetchone()
                    if existe_posicao:
                        con.execute(
                            """
                            UPDATE analise_posicao_resumo
                            SET estoque = ?,
                                origem = 'Banco de Dados',
                                atualizado_em = ?
                            WHERE periodo_referencia = ?
                            """,
                            (
                                valor_estoque_posicao,
                                agora_posicao,
                                periodo,
                            ),
                        )
                    else:
                        con.execute(
                            """
                            INSERT INTO analise_posicao_resumo
                                (
                                    periodo_referencia,
                                    contas_pagar,
                                    estoque,
                                    origem,
                                    atualizado_em
                                )
                            VALUES (?, 0, ?, 'Banco de Dados', ?)
                            """,
                            (
                                periodo,
                                valor_estoque_posicao,
                                agora_posicao,
                            ),
                        )
    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_posicao_resumo (
            periodo_referencia TEXT NOT NULL PRIMARY KEY,
            contas_pagar REAL NOT NULL DEFAULT 0,
            estoque REAL NOT NULL DEFAULT 0,
            origem TEXT,
            atualizado_em TEXT
        )
    """)

    con.commit()


def salvar_contas_pagar_independente(df, periodo, data_inicio, data_fim):
    """Fluxo exclusivo de consulta já normalizada até o cache e resumos."""
    obrigatorias = {
        "status", "data_vencimento", "data_pagamento",
        "valor_documento", "valor_pago", "saldo_aberto",
        "fornecedor", "unidade", "plano_contas",
    }
    faltantes = sorted(obrigatorias - set(df.columns))
    if faltantes:
        raise ValueError(
            "Campos obrigatórios ausentes: " + ", ".join(faltantes)
        )
    if df.empty:
        raise ValueError("A base normalizada está vazia.")

    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    base = df.copy()

    for coluna in ("valor_documento", "valor_pago", "saldo_aberto"):
        base[coluna] = pd.to_numeric(
            base[coluna], errors="coerce"
        ).fillna(0.0)

    for coluna in (
        "status", "data_vencimento", "data_pagamento",
        "fornecedor", "unidade", "apelido_unidade",
        "plano_contas", "numero_documento",
    ):
        if coluna not in base.columns:
            base[coluna] = ""
        base[coluna] = base[coluna].fillna("").astype(str)

    base["periodo_referencia"] = str(periodo)
    base["data_inicio_meta"] = str(data_inicio)
    base["data_fim_meta"] = str(data_fim)
    base["atualizado_em"] = agora

    destino = "base_contas_pagar"
    temporaria = "_tmp_contas_pagar_atualizacao"
    etapa = "abertura do SQLite"
    con = conexao_cache()

    try:
        etapa = "gravação temporária"
        con.execute(f'DROP TABLE IF EXISTS "{temporaria}"')
        con.commit()
        base.to_sql(temporaria, con, if_exists="replace", index=False)

        gravados_tmp = con.execute(
            f'SELECT COUNT(*) FROM "{temporaria}"'
        ).fetchone()[0]
        if int(gravados_tmp) != len(base):
            raise RuntimeError(
                f"Tabela temporária: esperado {len(base)}, "
                f"gravado {gravados_tmp}."
            )

        etapa = "adequação da tabela definitiva"
        existe = con.execute(
            "SELECT 1 FROM sqlite_master "
            "WHERE type='table' AND name=?",
            (destino,),
        ).fetchone()

        if not existe:
            con.execute(
                f'CREATE TABLE "{destino}" AS '
                f'SELECT * FROM "{temporaria}" WHERE 1=0'
            )
            con.commit()
        else:
            cols_destino = {
                r[1] for r in con.execute(
                    f'PRAGMA table_info("{destino}")'
                ).fetchall()
            }
            cols_tmp = [
                r[1] for r in con.execute(
                    f'PRAGMA table_info("{temporaria}")'
                ).fetchall()
            ]
            for coluna in cols_tmp:
                if coluna not in cols_destino:
                    nome_seguro = coluna.replace('"', '""')
                    con.execute(
                        f'ALTER TABLE "{destino}" '
                        f'ADD COLUMN "{nome_seguro}"'
                    )
            con.commit()

        cols_destino = {
            r[1] for r in con.execute(
                f'PRAGMA table_info("{destino}")'
            ).fetchall()
        }
        cols_tmp = [
            r[1] for r in con.execute(
                f'PRAGMA table_info("{temporaria}")'
            ).fetchall()
        ]
        cols = [c for c in cols_tmp if c in cols_destino]
        if not cols:
            raise RuntimeError("Nenhuma coluna compatível para inserção.")

        campos = ", ".join(
            '"' + c.replace('"', '""') + '"' for c in cols
        )

        etapa = "substituição atômica do mês"
        con.execute("BEGIN IMMEDIATE")
        con.execute(
            f'DELETE FROM "{destino}" '
            'WHERE periodo_referencia=?',
            (str(periodo),),
        )
        con.execute(
            f'INSERT INTO "{destino}" ({campos}) '
            f'SELECT {campos} FROM "{temporaria}"'
        )

        gravados = con.execute(
            f'SELECT COUNT(*) FROM "{destino}" '
            'WHERE periodo_referencia=?',
            (str(periodo),),
        ).fetchone()[0]
        if int(gravados) != len(base):
            raise RuntimeError(
                f"Tabela definitiva: esperado {len(base)}, "
                f"gravado {gravados}."
            )

        etapa = "criação dos resumos"
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS analise_contas_resumo (
                periodo_referencia TEXT NOT NULL,
                plano_contas TEXT NOT NULL,
                pagamento REAL NOT NULL DEFAULT 0,
                PRIMARY KEY (periodo_referencia, plano_contas)
            )
            """
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS analise_posicao_resumo (
                periodo_referencia TEXT NOT NULL PRIMARY KEY,
                contas_pagar REAL NOT NULL DEFAULT 0,
                estoque REAL NOT NULL DEFAULT 0,
                origem TEXT,
                atualizado_em TEXT
            )
            """
        )

        con.execute(
            "DELETE FROM analise_contas_resumo "
            "WHERE periodo_referencia=?",
            (str(periodo),),
        )
        con.execute(
            """
            INSERT OR REPLACE INTO analise_contas_resumo
                (periodo_referencia, plano_contas, pagamento)
            SELECT
                periodo_referencia,
                COALESCE(
                    NULLIF(TRIM(plano_contas), ''),
                    'SEM PLANO DE CONTAS'
                ),
                SUM(COALESCE(CAST(valor_pago AS REAL), 0))
            FROM base_contas_pagar
            WHERE periodo_referencia=?
            GROUP BY
                periodo_referencia,
                COALESCE(
                    NULLIF(TRIM(plano_contas), ''),
                    'SEM PLANO DE CONTAS'
                )
            """,
            (str(periodo),),
        )

        # Total de documentos com vencimento na competência.
        # Pagamentos realizados continuam separados no indicador de caixa.
        saldo = con.execute(
            """
            SELECT COALESCE(
                SUM(COALESCE(CAST(valor_documento AS REAL), 0)), 0
            )
            FROM base_contas_pagar
            WHERE periodo_referencia=?
            """,
            (str(periodo),),
        ).fetchone()[0] or 0

        con.execute(
            """
            INSERT INTO analise_posicao_resumo
                (periodo_referencia, contas_pagar, estoque,
                 origem, atualizado_em)
            VALUES (?, ?, 0, 'Banco de Dados', ?)
            ON CONFLICT(periodo_referencia) DO UPDATE SET
                contas_pagar=excluded.contas_pagar,
                origem=excluded.origem,
                atualizado_em=excluded.atualizado_em
            """,
            (str(periodo), float(saldo), agora),
        )

        etapa = "histórico"
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS atualizacoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fonte TEXT,
                periodo_referencia TEXT,
                data_inicio TEXT,
                data_fim TEXT,
                registros INTEGER,
                status TEXT,
                mensagem TEXT,
                atualizado_em TEXT
            )
            """
        )
        con.execute(
            """
            INSERT INTO atualizacoes
                (fonte, periodo_referencia, data_inicio, data_fim,
                 registros, status, mensagem, atualizado_em)
            VALUES (?, ?, ?, ?, ?, 'Sucesso', ?, ?)
            """,
            (
                "contas_pagar", str(periodo),
                str(data_inicio), str(data_fim),
                int(len(base)),
                "Rotina exclusiva concluída",
                agora,
            ),
        )

        con.execute(f'DROP TABLE IF EXISTS "{temporaria}"')
        con.commit()
        _limpar_cache_dados()
        st.cache_data.clear()
        return int(len(base))

    except Exception as erro:
        try:
            con.rollback()
        except Exception:
            pass
        try:
            con.execute(f'DROP TABLE IF EXISTS "{temporaria}"')
            con.commit()
        except Exception:
            pass

        LOG_DIR.mkdir(parents=True, exist_ok=True)
        log = LOG_DIR / "contas_pagar_gravacao_erro.txt"
        detalhe = (
            f"Etapa: {etapa}\n"
            f"Período: {periodo}\n"
            f"Registros: {len(base)}\n"
            f"Tipo: {type(erro).__name__}\n"
            f"Mensagem: {erro}\n\n"
            f"{traceback.format_exc()}"
        )
        log.write_text(detalhe, encoding="utf-8")
        raise RuntimeError(
            "Falha na atualização exclusiva de Contas a Pagar.\n"
            f"Etapa: {etapa}\nDetalhe: {erro}\nLog: {log}"
        ) from erro
    finally:
        con.close()


def salvar_snapshot_mensal(df, fonte, periodo, data_inicio, data_fim):
    with conexao_cache() as con_estrutura:
        garantir_tabelas_analise(con_estrutura)

    info = FONTES_BANCO[fonte]
    tabela = info["tabela_cache"]
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    base = df.copy()
    base["periodo_referencia"] = periodo
    base["data_inicio_meta"] = data_inicio
    base["data_fim_meta"] = data_fim
    base["atualizado_em"] = agora

    con = conexao_cache()
    try:
        # Remove somente o período que será atualizado.
        try:
            con.execute(
                f'DELETE FROM "{tabela}" WHERE periodo_referencia = ?',
                (periodo,)
            )
            con.commit()
        except sqlite3.OperationalError:
            pass

        base.to_sql(tabela, con, if_exists="append", index=False)
        # Atualiza somente o pequeno resumo da fonte/período alterado.
        atualizar_resumos_analise(con, fonte=fonte, periodo=periodo)
        con.execute(
            """
            INSERT INTO atualizacoes
            (fonte, periodo_referencia, data_inicio, data_fim, registros, status, mensagem, atualizado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                fonte, periodo, data_inicio, data_fim,
                len(base), "Sucesso", "Atualização concluída", agora
            )
        )
        con.commit()
        _limpar_cache_dados()
    finally:
        con.close()

def limpar_dados_operacionais():
    """Remove bases, resumos e histórico técnico sem apagar cadastros e metas."""
    tabelas = [
        "base_vendas",
        "base_estoque",
        "base_entradas",
        "base_contas_pagar",
        "analise_vendas_resumo",
        "analise_entradas_resumo",
        "analise_contas_resumo",
        "analise_estoque_resumo",
        "analise_posicao_resumo",
        "atualizacoes",
    ]
    con = conexao_cache()
    try:
        for tabela in tabelas:
            try:
                con.execute(f'DELETE FROM "{tabela}"')
            except sqlite3.OperationalError:
                pass
        con.commit()
        try:
            con.execute("VACUUM")
        except Exception:
            pass
    finally:
        con.close()

    _limpar_cache_dados()
    st.cache_data.clear()


def _config_psycopg2(cfg):
    return {
        "host": cfg.get("host") or cfg.get("servidor"),
        "port": int(cfg.get("port") or cfg.get("porta") or 5432),
        "dbname": cfg.get("database") or cfg.get("banco"),
        "user": cfg.get("user") or cfg.get("usuario"),
        "password": cfg.get("password") or cfg.get("senha"),
        "connect_timeout": 20,
        "application_name": "Rede Economize KPI Comercial",
    }


def _gravar_log_contas_pagar(texto):
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    caminho = LOG_DIR / "contas_pagar_erros.log"
    with caminho.open("a", encoding="utf-8") as arquivo:
        arquivo.write(
            f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]\n{texto}\n"
        )
    return caminho


def executar_contas_pagar_psycopg2(cfg, sql, data_inicio, data_fim, periodo):
    """Executa exatamente o SQL validado no DBeaver via psycopg2 direto."""
    inicio = datetime.strptime(str(data_inicio)[:10], "%Y-%m-%d").date()
    fim = datetime.strptime(str(data_fim)[:10], "%Y-%m-%d").date()
    sql_final = str(sql)
    sql_final, qtd_inicio = re.subn(
        r"DATE\s+'\d{4}-\d{2}-\d{2}'\s+AS\s+data_inicial",
        f"DATE '{inicio.isoformat()}' AS data_inicial",
        sql_final, count=1, flags=re.IGNORECASE,
    )
    sql_final, qtd_fim = re.subn(
        r"DATE\s+'\d{4}-\d{2}-\d{2}'\s+AS\s+data_final",
        f"DATE '{fim.isoformat()}' AS data_final",
        sql_final, count=1, flags=re.IGNORECASE,
    )
    if qtd_inicio != 1 or qtd_fim != 1:
        raise RuntimeError("As datas do CTE parametros não foram encontradas no SQL oficial.")

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    arquivo_sql = LOG_DIR / "contas_pagar_sql_executado.sql"
    arquivo_erro = LOG_DIR / "contas_pagar_erro.txt"
    arquivo_sql.write_text(sql_final, encoding="utf-8")

    parametros = {
        "host": cfg.get("host"),
        "port": int(cfg.get("porta", 5432)),
        "dbname": cfg.get("banco"),
        "user": cfg.get("usuario"),
        "password": cfg.get("senha"),
        "connect_timeout": 20,
        "application_name": "Rede Economize KPI Comercial",
        "options": "-c statement_timeout=600000",
    }
    sslmode = str(cfg.get("sslmode", "prefer") or "prefer").strip()
    if sslmode:
        parametros["sslmode"] = sslmode

    conexao = None
    cursor = None
    try:
        conexao = psycopg2.connect(**parametros)
        conexao.autocommit = False
        cursor = conexao.cursor()
        cursor.execute("SELECT current_database(), current_user, current_schema(), current_setting('search_path')")
        ambiente = cursor.fetchone()
        cursor.execute(sql_final)
        if cursor.description is None:
            raise RuntimeError("O SQL executou, mas não retornou colunas de resultado.")
        colunas = [d.name for d in cursor.description]
        registros = cursor.fetchall()
        conexao.rollback()
        if not registros:
            raise ValueError(f"O SQL executou corretamente, mas retornou 0 registros entre {inicio} e {fim}.")
        df = pd.DataFrame.from_records(registros, columns=colunas)
        df.attrs["ambiente_postgresql"] = ambiente
        return df
    except Exception as erro:
        if conexao is not None:
            try: conexao.rollback()
            except Exception: pass
        detalhes=[f"Competência: {periodo}",f"Período SQL: {inicio} até {fim}",f"Tipo: {type(erro).__name__}",f"Mensagem: {erro}"]
        if getattr(erro,'pgcode',None): detalhes.append(f"SQLSTATE: {erro.pgcode}")
        if getattr(erro,'pgerror',None): detalhes.append(f"PostgreSQL: {erro.pgerror}")
        diag=getattr(erro,'diag',None)
        if diag is not None:
            for attr,label in [("severity","Severidade"),("message_primary","Mensagem principal"),("message_detail","Detalhe"),("message_hint","Sugestão"),("statement_position","Posição no SQL"),("context","Contexto"),("schema_name","Schema"),("table_name","Tabela"),("column_name","Coluna")]:
                val=getattr(diag,attr,None)
                if val: detalhes.append(f"{label}: {val}")
        detalhes += ["", "TRACEBACK:", traceback.format_exc()]
        texto="\n".join(detalhes)
        arquivo_erro.write_text(texto,encoding='utf-8')
        raise RuntimeError(texto) from erro
    finally:
        if cursor is not None:
            try: cursor.close()
            except Exception: pass
        if conexao is not None:
            try: conexao.close()
            except Exception: pass


def executar_sql_com_fallback_datas(conn, sql, data_inicio, data_fim, periodo):
    inicio_date = datetime.strptime(str(data_inicio)[:10], "%Y-%m-%d").date()
    fim_date = datetime.strptime(str(data_fim)[:10], "%Y-%m-%d").date()

    return pd.read_sql_query(
        text(sql),
        conn,
        params={
            "data_inicio": inicio_date,
            "data_fim": fim_date,
            "periodo_referencia": str(periodo),
        },
    )



def importar_csv_anual_contas_pagar(arquivo):
    """Importa uma base anual e cria snapshots mensais pelo vencimento."""
    nome = str(getattr(arquivo, "name", "contas_pagar.csv"))
    arquivo.seek(0)
    try:
        bruto = pd.read_csv(
            arquivo,
            sep=None,
            engine="python",
            dtype=str,
            encoding="utf-8-sig",
        )
    except UnicodeDecodeError:
        arquivo.seek(0)
        bruto = pd.read_csv(
            arquivo,
            sep=None,
            engine="python",
            dtype=str,
            encoding="latin-1",
        )

    if bruto.empty:
        raise ValueError("O CSV de Contas a Pagar está vazio.")

    normalizado = normalizar_contas_pagar_df(bruto)
    vencimentos = pd.to_datetime(
        normalizado["data_vencimento"], errors="coerce"
    )
    normalizado = normalizado.loc[vencimentos.notna()].copy()
    vencimentos = vencimentos.loc[vencimentos.notna()]
    normalizado["periodo_referencia"] = vencimentos.dt.strftime("%Y-%m")

    if normalizado.empty:
        raise ValueError(
            "Não foi possível reconhecer a coluna Data de Vencimento no CSV."
        )

    resultados = []
    planos = sorted(
        p for p in normalizado["plano_contas"].dropna().astype(str).str.strip().unique()
        if p
    )

    for periodo, grupo in normalizado.groupby("periodo_referencia"):
        ano, mes = map(int, periodo.split("-"))
        import calendar
        inicio = f"{periodo}-01"
        fim = f"{periodo}-{calendar.monthrange(ano, mes)[1]:02d}"
        salvar_snapshot_mensal(
            grupo.drop(columns=["periodo_referencia"]),
            "contas_pagar",
            periodo,
            inicio,
            fim,
        )
        resultados.append({"Período": periodo, "Registros": len(grupo)})

    _gravar_log_contas_pagar(
        f"Importação CSV concluída: {nome}; "
        f"{len(normalizado)} registros; {len(resultados)} competências."
    )
    return pd.DataFrame(resultados), planos


def registrar_erro_atualizacao(fonte, periodo, data_inicio, data_fim, mensagem):
    con = conexao_cache()
    try:
        con.execute(
            """
            INSERT INTO atualizacoes
            (fonte, periodo_referencia, data_inicio, data_fim, registros, status, mensagem, atualizado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                fonte, periodo, data_inicio, data_fim,
                0, "Erro", str(mensagem)[:8000],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
        )
        con.commit()
    finally:
        con.close()

def executar_atualizacao_fonte(fonte, cfg, periodo, data_inicio, data_fim):
    info = FONTES_BANCO[fonte]
    sql = ler_sql(info["arquivo_sql"])

    if not str(sql).strip():
        raise ValueError(f"O SQL de {info['titulo']} está vazio.")

    engine = None

    try:
        if fonte == "contas_pagar":
            df = executar_contas_pagar_psycopg2(
                cfg,
                sql,
                data_inicio,
                data_fim,
                periodo,
            )
        else:
            engine = criar_engine_banco(cfg)
            with engine.connect() as conn:
                df = executar_sql_com_fallback_datas(
                    conn,
                    sql,
                    data_inicio,
                    data_fim,
                    periodo,
                )

        if fonte == "contas_pagar":
            colunas_originais = [str(c) for c in df.columns]
            df = normalizar_contas_pagar_df(df)

            if df.empty:
                raise ValueError(
                    "O SQL retornou registros, mas a normalização resultou em base vazia. "
                    f"Colunas recebidas: {colunas_originais}"
                )

            if "plano_contas" not in df.columns:
                raise ValueError(
                    "A coluna Plano de Contas não foi reconhecida no resultado."
                )

        if fonte == "contas_pagar":
            return salvar_contas_pagar_independente(
                df,
                periodo,
                data_inicio,
                data_fim,
            )

        salvar_snapshot_mensal(
            df,
            fonte,
            periodo,
            data_inicio,
            data_fim,
        )
        return len(df)

    except Exception as erro:
        mensagem = str(erro)
        registrar_erro_atualizacao(
            fonte, periodo, data_inicio, data_fim, mensagem,
        )
        if fonte == "contas_pagar":
            raise
        raise RuntimeError(
            f"Falha na consulta de {info['titulo']}:\n{mensagem}"
        ) from erro

    finally:
        if engine is not None:
            engine.dispose()



def status_configuracao_fontes():
    resultado = {}
    for codigo, info in FONTES_BANCO.items():
        sql = ler_sql(info["arquivo_sql"])
        configurado, mensagem = diagnosticar_sql_fonte(codigo, sql)
        resultado[codigo] = {
            "configurado": configurado,
            "mensagem": mensagem,
        }
    return resultado


def historico_atualizacoes():
    con = conexao_cache()
    try:
        return pd.read_sql_query(
            "SELECT * FROM atualizacoes ORDER BY id DESC",
            con
        )
    finally:
        con.close()

def carregar_snapshot(fonte, periodo):
    tabela = FONTES_BANCO[fonte]["tabela_cache"]
    con = conexao_cache()
    try:
        try:
            return pd.read_sql_query(
                f'SELECT * FROM "{tabela}" WHERE periodo_referencia = ?',
                con,
                params=(periodo,)
            )
        except Exception:
            return pd.DataFrame()
    finally:
        con.close()


def garantir_indices_cache():
    """Índices mínimos e objetivos para acelerar consultas por competência."""
    con = conexao_cache()
    try:
        tabelas = ["base_vendas", "base_entradas", "base_estoque", "base_contas_pagar"]
        candidatas_class = [
            "classificacao_resumida", "classificacao_3_nivel",
            "classificacao_geral", "classificacao_principal", "classificacao"
        ]
        for tabela in tabelas:
            try:
                cols = {r[1] for r in con.execute(f'PRAGMA table_info("{tabela}")').fetchall()}
                if "periodo_referencia" not in cols:
                    continue
                con.execute(
                    f'CREATE INDEX IF NOT EXISTS "idx_{tabela}_periodo" '
                    f'ON "{tabela}" (periodo_referencia)'
                )
                col_class = next((c for c in candidatas_class if c in cols), None)
                if col_class:
                    con.execute(
                        f'CREATE INDEX IF NOT EXISTS "idx_{tabela}_periodo_class" '
                        f'ON "{tabela}" (periodo_referencia, "{col_class}")'
                    )
            except Exception:
                continue
        con.commit()
    finally:
        con.close()


garantir_sqls()
# PERFORMANCE: não criar índices sobre milhões de linhas durante a abertura.
# A manutenção de índices deve ocorrer somente após atualização manual do banco.
CONFIG_BANCO = carregar_config_banco()


# =========================================================
# RUPTURA AUTOMÁTICA POR PERÍODO DA META
# =========================================================

PASTA_RUPTURA_AUTO = Path("IMPORTAR_RUPTURA")
PASTA_RUPTURA_AUTO.mkdir(exist_ok=True)
RUPTURA_AUTO_DB = DATA_DIR / "ruptura_mensal.sqlite"
RUPTURA_AUTO_CONTROLE = DATA_DIR / "controle_ruptura_auto.json"

MAPA_COMPRADORES_RUPTURA = {
    "PRINCIPAL > NAO MED > BRINQUEDOS": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA COMBATE": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA ENCALHADOS": "Francieli",
    "PRINCIPAL > NAO MED > CONVENIÊNCIA": "Francieli",
    "PRINCIPAL > NAO MED > SUPLEMENTOS": "Francieli",
    "PRINCIPAL > MED > ÉTICOS": "Paulo",
    "PRINCIPAL > NAO MED > MASCARA": "Sebastião",
    "PRINCIPAL > NAO MED > FRALDAS": "Sebastião",
    "PRINCIPAL > MED > GEN - SIM": "Sebastião",
    "PRINCIPAL > MED > MEDICAMENTO CURVA D (TOP 15)": "Sebastião",
    "PRINCIPAL > MED > NATURAIS": "Sebastião",
    "PRINCIPAL > MED > PRÓPRIOS": "Sebastião",
    "PRINCIPAL > NAO MED > HOSPITALARES": "Sebastião",
    "PRINCIPAL > NAO MED > LEITES": "Sebastião",
    "PRINCIPAL > NAO MED > VAREJO": "Sebastião",
}

def _hash_arquivo(caminho):
    h = hashlib.sha256()
    with open(caminho, "rb") as f:
        for bloco in iter(lambda: f.read(1024 * 1024), b""):
            h.update(bloco)
    return h.hexdigest()

def _numero(serie):
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce").fillna(0)
    s = serie.astype(str).str.strip()
    tem_virgula = s.str.contains(",", regex=False)
    r = pd.Series(index=s.index, dtype="float64")
    r.loc[tem_virgula] = pd.to_numeric(
        s.loc[tem_virgula].str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
        errors="coerce"
    )
    r.loc[~tem_virgula] = pd.to_numeric(s.loc[~tem_virgula], errors="coerce")
    return r.fillna(0)

def _comprador(classificacao):
    chave = " ".join(str(classificacao or "").strip().upper().split())
    mapa = {" ".join(k.upper().split()): v for k, v in MAPA_COMPRADORES_RUPTURA.items()}
    return mapa.get(chave, "Não mapeado")

def _arquivo_mais_recente():
    arquivos = []
    for padrao in ("*.xlsx", "*.xls", "*.csv"):
        arquivos.extend(PASTA_RUPTURA_AUTO.glob(padrao))
    arquivos = [p for p in arquivos if not p.name.startswith("~$")]
    return max(arquivos, key=lambda p: p.stat().st_mtime) if arquivos else None

def _ler_modelo_ruptura(caminho):
    if caminho.suffix.lower() == ".csv":
        try:
            bruto = pd.read_csv(caminho, sep=";", encoding="utf-8-sig")
        except Exception:
            bruto = pd.read_csv(caminho)
    else:
        bruto = pd.read_excel(caminho)

    bruto.columns = [str(c).strip() for c in bruto.columns]
    obrigatorias = [
        "Un. Neg.", "Apelido Un. Neg.", "Produto", "Fabricante",
        "Ruptura Venda", "Necessidade", "Estoque", "Custo Médio",
        "Curva Valor", "Cód. de Barras", "Classificação Principal"
    ]
    faltantes = [c for c in obrigatorias if c not in bruto.columns]
    if faltantes:
        raise ValueError("Colunas ausentes: " + ", ".join(faltantes))

    out = pd.DataFrame()
    out["Loja"] = bruto["Un. Neg."].astype(str).str.strip()
    out["Apelido Loja"] = bruto["Apelido Un. Neg."].astype(str).str.strip()
    out["Produto"] = bruto["Produto"].astype(str).str.strip()
    out["Fabricante"] = bruto["Fabricante"].astype(str).str.strip()
    out["EAN"] = bruto["Cód. de Barras"].astype(str).str.replace(r"\.0$", "", regex=True)
    out["Classificação Principal"] = bruto["Classificação Principal"].astype(str).str.strip()
    out["Comprador"] = out["Classificação Principal"].map(_comprador)
    out["Curva Valor"] = bruto["Curva Valor"].astype(str).str.strip()
    out["Curva Qtd."] = bruto["Curva Qtd."].astype(str).str.strip() if "Curva Qtd." in bruto else ""
    out["Ruptura Venda"] = _numero(bruto["Ruptura Venda"])
    out["Necessidade"] = _numero(bruto["Necessidade"])
    out["Estoque"] = _numero(bruto["Estoque"])
    out["Custo Médio"] = _numero(bruto["Custo Médio"])
    out["Valor Necessidade Custo"] = out["Necessidade"] * out["Custo Médio"]
    out["Valor Ruptura"] = out["Ruptura Venda"]
    out = out[(out["Valor Ruptura"] != 0) | (out["Necessidade"] != 0) | (out["Estoque"] != 0)]
    return out

def _con_ruptura():
    con = sqlite3.connect(RUPTURA_AUTO_DB)
    con.execute("""
        CREATE TABLE IF NOT EXISTS importacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            periodo_referencia TEXT,
            arquivo TEXT,
            hash_arquivo TEXT,
            registros INTEGER,
            valor_ruptura REAL,
            importado_em TEXT
        )
    """)
    con.commit()
    return con

def processar_ruptura_automatica(periodo=None, forcar=False):
    periodo = periodo or METAS_GESTOR.get("periodo_referencia", "")
    arquivo = _arquivo_mais_recente()
    if arquivo is None:
        return {"status": "sem_arquivo", "mensagem": "Nenhum arquivo na pasta IMPORTAR_RUPTURA."}

    hash_atual = _hash_arquivo(arquivo)
    controle = {}
    if RUPTURA_AUTO_CONTROLE.exists():
        try:
            controle = json.loads(RUPTURA_AUTO_CONTROLE.read_text(encoding="utf-8"))
        except Exception:
            controle = {}

    chave = f"{periodo}|{arquivo.name}"
    if not forcar and controle.get(chave) == hash_atual:
        return {"status": "sem_alteracao", "mensagem": "Arquivo já processado.", "arquivo": arquivo.name}

    df = _ler_modelo_ruptura(arquivo)
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    base = df.copy()
    base["periodo_referencia"] = periodo
    base["arquivo_origem"] = arquivo.name
    base["importado_em"] = agora

    con = _con_ruptura()
    try:
        try:
            con.execute("DELETE FROM ruptura_detalhe WHERE periodo_referencia = ?", (periodo,))
            con.commit()
        except Exception:
            pass
        base.to_sql("ruptura_detalhe", con, if_exists="append", index=False)
        con.execute(
            "INSERT INTO importacoes(periodo_referencia,arquivo,hash_arquivo,registros,valor_ruptura,importado_em) VALUES(?,?,?,?,?,?)",
            (periodo, arquivo.name, hash_atual, len(base), float(base["Valor Ruptura"].sum()), agora)
        )
        con.commit()
    finally:
        con.close()

    controle[chave] = hash_atual
    controle["ultima_importacao"] = {
        "periodo": periodo,
        "arquivo": arquivo.name,
        "registros": len(base),
        "valor_ruptura": float(base["Valor Ruptura"].sum()),
        "data": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    }
    RUPTURA_AUTO_CONTROLE.write_text(json.dumps(controle, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"status": "importado", **controle["ultima_importacao"]}

@st.cache_data(ttl=600, show_spinner=False)
def carregar_ruptura_auto(periodo):
    con = _con_ruptura()
    try:
        try:
            base = pd.read_sql_query(
                "SELECT * FROM ruptura_detalhe WHERE periodo_referencia = ?",
                con, params=(periodo,)
            )
        except Exception:
            return pd.DataFrame()
    finally:
        con.close()

    # Recalcula sempre o comprador pela Classificação Principal.
    # Isso corrige também registros antigos que foram gravados como
    # "Não mapeado", sem exigir uma nova importação da planilha.
    if not base.empty:
        coluna_classificacao = next(
            (
                coluna for coluna in [
                    "Classificação Principal", "classificacao principal",
                    "classificacao_principal", "classificacao_geral",
                    "classificacao"
                ]
                if coluna in base.columns
            ),
            None
        )
        if coluna_classificacao:
            base["Comprador"] = base[coluna_classificacao].map(_mapear_comprador)
        elif "Comprador" not in base.columns:
            base["Comprador"] = "Não mapeado"

    return base

def historico_ruptura_auto():
    con = _con_ruptura()
    try:
        return pd.read_sql_query("SELECT * FROM importacoes ORDER BY id DESC", con)
    finally:
        con.close()

# Executa uma vez ao abrir e não repete se o arquivo não mudou.
RESULTADO_AUTO_RUPTURA = processar_ruptura_automatica()


# =========================================================
# GESTÃO EDITÁVEL DE COMPRADORES POR CLASSIFICAÇÃO
# =========================================================

MAPA_COMPRADORES_FILE = DATA_DIR / "mapa_compradores_editavel.json"

MAPA_COMPRADORES_PADRAO = [
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > BRINQUEDOS", "Comprador": "Francieli"},
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > PERFUMARIA", "Comprador": "Francieli"},
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > PERFUMARIA COMBATE", "Comprador": "Francieli"},
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > PERFUMARIA ENCALHADOS", "Comprador": "Francieli"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > NAO MED > CONVENIÊNCIA", "Comprador": "Francieli"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > NAO MED > SUPLEMENTOS", "Comprador": "Francieli"},
    {"Área": "Propagados", "Classificação Principal": "PRINCIPAL > MED > ÉTICOS", "Comprador": "Paulo"},
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > MASCARA", "Comprador": "Não mapeado"},
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > FRALDAS", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > MED > GEN - SIM", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > MED > MEDICAMENTO CURVA D (TOP 15)", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > MED > NATURAIS", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > MED > PRÓPRIOS", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > NAO MED > HOSPITALARES", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > NAO MED > LEITES", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > NAO MED > VAREJO", "Comprador": "Não mapeado"},
]

def carregar_mapa_compradores_editavel():
    if MAPA_COMPRADORES_FILE.exists():
        try:
            dados = json.loads(MAPA_COMPRADORES_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, list):
                return dados
        except Exception:
            pass
    MAPA_COMPRADORES_FILE.write_text(
        json.dumps(MAPA_COMPRADORES_PADRAO, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    return MAPA_COMPRADORES_PADRAO.copy()

def salvar_mapa_compradores_editavel(dados):
    MAPA_COMPRADORES_FILE.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

def mapa_compradores_dict():
    dados = carregar_mapa_compradores_editavel()
    return {
        _normalizar_classificacao(item.get("Classificação Principal", "")): item.get("Comprador", "Não mapeado")
        for item in dados
        if item.get("Classificação Principal")
    }

MAPA_COMPRADORES_EDITAVEL = carregar_mapa_compradores_editavel()


# =========================================================
# CADASTRO EDITÁVEL DE COMPRADORES
# =========================================================

COMPRADORES_FILE = DATA_DIR / "cadastro_compradores.json"

COMPRADORES_PADRAO = [
    {"Comprador": "Geane", "Status": "Ativo"},
    {"Comprador": "Renato", "Status": "Ativo"},
]

def carregar_cadastro_compradores():
    if COMPRADORES_FILE.exists():
        try:
            dados = json.loads(COMPRADORES_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, list) and dados:
                return dados
        except Exception:
            pass
    COMPRADORES_FILE.write_text(
        json.dumps(COMPRADORES_PADRAO, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    return COMPRADORES_PADRAO.copy()

def salvar_cadastro_compradores(dados):
    COMPRADORES_FILE.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

def lista_compradores_ativos():
    return [
        str(item.get("Comprador", "")).strip()
        for item in carregar_cadastro_compradores()
        if str(item.get("Status", "Ativo")).strip() == "Ativo"
        and str(item.get("Comprador", "")).strip()
    ]

COMPRADORES = lista_compradores_ativos()


def _conjunto_compradores_ativos():
    return {nome.casefold(): nome for nome in lista_compradores_ativos()}


def filtrar_dataframe_compradores_ativos(df):
    if not isinstance(df, pd.DataFrame) or df.empty or "Comprador" not in df.columns:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    ativos = set(_conjunto_compradores_ativos())
    nomes = df["Comprador"].astype(str).str.strip().str.casefold()
    return df.loc[nomes.isin(ativos)].copy()


# =========================================================
# METAS EDITÁVEIS POR COMPRADOR E PERÍODO
# =========================================================

METAS_COMPRADORES_FILE = DATA_DIR / "metas_por_comprador.json"

def estrutura_meta_comprador_padrao(comprador, periodo, participacao_inicial=0.0):
    # Compradores e participações não são fixos no código. Para compradores
    # recém-identificados na base, a participação inicial pode ser calculada
    # pelo faturamento real do período e depois alterada na Gestão de Metas.
    participacao = float(participacao_inicial or 0.0)
    meta_venda = float(METAS_GESTOR.get("meta_venda_total_mes", 0)) * participacao / 100.0
    meta_cmv_pct = float(METAS_GESTOR.get("meta_cmv_mes", 0))
    meta_cmv = meta_venda * meta_cmv_pct / 100.0
    meta_estoque = meta_cmv * float(METAS_GESTOR.get("fator_cobertura", 0))

    return {
        "periodo_referencia": periodo,
        "comprador": comprador,
        "meta_venda": meta_venda,
        "participacao_venda_pct": participacao,
        "meta_cmv_pct": meta_cmv_pct,
        "meta_cmv_valor": meta_cmv,
        "fator_cobertura": float(METAS_GESTOR.get("fator_cobertura", 0)),
        "meta_estoque_total": meta_estoque,
        "meta_curva_a_pct": float(METAS_GESTOR.get("curva_a", 0)),
        "meta_curva_b_pct": float(METAS_GESTOR.get("curva_b", 0)),
        "meta_curva_c_pct": float(METAS_GESTOR.get("curva_c", 0)),
        "meta_curva_d_pct": float(METAS_GESTOR.get("curva_d", 0)),
        "meta_ruptura_pct": float(METAS_GESTOR.get("meta_ruptura", 0)),
        "meta_reposicao_pct": float(METAS_GESTOR.get("meta_reposicao", 0)),
        "valor_premio": float(METAS_GESTOR.get("valor_premio_total", 0)),
        "status": "Ativa",
        "ultima_atualizacao": "",
    }

def carregar_metas_por_comprador():
    if METAS_COMPRADORES_FILE.exists():
        try:
            dados = json.loads(METAS_COMPRADORES_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, list):
                return dados
        except Exception:
            pass
    return []

def salvar_metas_por_comprador(dados):
    METAS_COMPRADORES_FILE.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def _nome_comprador_valido(valor):
    texto = str(valor or "").strip()
    return texto and texto.upper() not in {
        "NÃO MAPEADO", "NAO MAPEADO", "NAN", "NONE", "NULL"
    }


def sincronizar_compradores_reconhecidos(nomes):
    """Inclui no cadastro compradores encontrados nas bases, sem apagar os existentes."""
    atuais = carregar_cadastro_compradores()
    conhecidos = {
        str(item.get("Comprador", "")).strip().casefold()
        for item in atuais if str(item.get("Comprador", "")).strip()
    }
    alterado = False
    for nome in sorted({str(x).strip() for x in nomes if _nome_comprador_valido(x)}):
        if nome.casefold() not in conhecidos:
            atuais.append({"Comprador": nome, "Status": "Ativo", "Origem": "Base de dados"})
            conhecidos.add(nome.casefold())
            alterado = True
    if alterado:
        salvar_cadastro_compradores(atuais)
    return atuais


def garantir_meta_comprador_periodo(comprador, periodo, participacao_inicial=0.0):
    dados = carregar_metas_por_comprador()
    for item in dados:
        if (
            str(item.get("periodo_referencia", "")) == str(periodo)
            and str(item.get("comprador", "")).strip().casefold() == str(comprador).strip().casefold()
        ):
            return item
    novo = estrutura_meta_comprador_padrao(comprador, periodo, participacao_inicial)
    dados.append(novo)
    salvar_metas_por_comprador(dados)
    return novo

def garantir_metas_compradores_periodo(periodo):
    dados = carregar_metas_por_comprador()
    ativos = lista_compradores_ativos()
    existentes = {
        (str(x.get("periodo_referencia", "")), str(x.get("comprador", "")))
        for x in dados
    }
    alterado = False

    for comprador in ativos:
        chave = (periodo, comprador)
        if chave not in existentes:
            dados.append(estrutura_meta_comprador_padrao(comprador, periodo))
            alterado = True

    if alterado:
        salvar_metas_por_comprador(dados)
    return dados

def obter_meta_comprador(comprador, periodo):
    dados = garantir_metas_compradores_periodo(periodo)
    for item in dados:
        if (
            str(item.get("periodo_referencia", "")) == str(periodo)
            and str(item.get("comprador", "")) == str(comprador)
        ):
            return item
    return estrutura_meta_comprador_padrao(comprador, periodo)

def atualizar_nome_comprador_metas(nome_antigo, nome_novo):
    dados = carregar_metas_por_comprador()
    alterado = False
    for item in dados:
        if str(item.get("comprador", "")).strip() == nome_antigo:
            item["comprador"] = nome_novo
            alterado = True
    if alterado:
        salvar_metas_por_comprador(dados)

garantir_metas_compradores_periodo(METAS_GESTOR.get("periodo_referencia", ""))

# Compradores inativos permanecem apenas no histórico.
_dados_metas_ativos = carregar_metas_por_comprador()
_alterou_status = False
_ativos_cf = set(_conjunto_compradores_ativos())
for _item_meta in _dados_metas_ativos:
    _nome_cf = str(_item_meta.get("comprador", "")).strip().casefold()
    if _nome_cf and _nome_cf not in _ativos_cf and str(_item_meta.get("status", "")).casefold() != "inativa":
        _item_meta["status"] = "Inativa"
        _item_meta["motivo_inativacao"] = "Comprador fora do cadastro ativo"
        _alterou_status = True
if _alterou_status:
    salvar_metas_por_comprador(_dados_metas_ativos)

REALIZADOS = filtrar_dataframe_compradores_ativos(REALIZADOS)
METAS = filtrar_dataframe_compradores_ativos(METAS)
RESULTADO = filtrar_dataframe_compradores_ativos(RESULTADO)
PREMIO = filtrar_dataframe_compradores_ativos(PREMIO)


# =========================================================
# MOTOR DINÂMICO DAS VISÕES
# =========================================================

MAPA_CLASSIFICACAO_COMPRADOR = {
    "PRINCIPAL > NAO MED > BRINQUEDOS": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA COMBATE": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA ENCALHADOS": "Francieli",
    "PRINCIPAL > NAO MED > CONVENIÊNCIA": "Francieli",
    "PRINCIPAL > NAO MED > SUPLEMENTOS": "Francieli",
    "PRINCIPAL > MED > ÉTICOS": "Paulo",
    "PRINCIPAL > NAO MED > MASCARA": "Não mapeado",
    "PRINCIPAL > NAO MED > FRALDAS": "Não mapeado",
    "PRINCIPAL > MED > GEN - SIM": "Não mapeado",
    "PRINCIPAL > MED > MEDICAMENTO CURVA D (TOP 15)": "Não mapeado",
    "PRINCIPAL > MED > NATURAIS": "Não mapeado",
    "PRINCIPAL > MED > PRÓPRIOS": "Não mapeado",
    "PRINCIPAL > NAO MED > HOSPITALARES": "Não mapeado",
    "PRINCIPAL > NAO MED > LEITES": "Não mapeado",
    "PRINCIPAL > NAO MED > VAREJO": "Não mapeado",
}

def _norm_coluna(valor):
    import unicodedata
    txt = unicodedata.normalize("NFKD", str(valor))
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return " ".join(txt.lower().replace("_", " ").strip().split())

def _achar_coluna(df, candidatos):
    if df is None or df.empty:
        return None
    mapa = {_norm_coluna(c): c for c in df.columns}
    for candidato in candidatos:
        chave = _norm_coluna(candidato)
        if chave in mapa:
            return mapa[chave]
    for candidato in candidatos:
        chave = _norm_coluna(candidato)
        for norm, original in mapa.items():
            if chave in norm or norm in chave:
                return original
    return None

def _numero_df(df, candidatos):
    coluna = _achar_coluna(df, candidatos)
    if coluna is None:
        return pd.Series(0.0, index=df.index)
    serie = df[coluna]
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce").fillna(0.0)
    texto = serie.astype(str).str.strip()
    tem_virgula = texto.str.contains(",", regex=False)
    resultado = pd.Series(index=texto.index, dtype="float64")
    resultado.loc[tem_virgula] = pd.to_numeric(
        texto.loc[tem_virgula]
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False),
        errors="coerce"
    )
    resultado.loc[~tem_virgula] = pd.to_numeric(texto.loc[~tem_virgula], errors="coerce")
    return resultado.fillna(0.0)

def _texto_df(df, candidatos, padrao=""):
    coluna = _achar_coluna(df, candidatos)
    if coluna is None:
        return pd.Series(padrao, index=df.index, dtype="object")
    return df[coluna].fillna(padrao).astype(str).str.strip()

def _normalizar_classificacao(valor):
    """Normaliza o caminho sem perder os níveis da classificação.

    Trata acentos, quebras de linha, espaços duplicados e diferenças de
    digitação entre PostgreSQL, Excel e o cadastro do comprador.
    """
    texto = str(valor or "").replace("\n", " ").replace("\r", " ")
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.upper().strip()
    texto = re.sub(r"\s*>\s*", " > ", texto)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip(" >")


def _mapa_compradores_ordenado():
    """Retorna somente mapeamentos de compradores ativos.

    O Cadastro de Compradores é a fonte oficial. Metas antigas, cache e
    mapeamentos históricos nunca reativam um comprador removido/inativo.
    """
    ativos = {nome.casefold(): nome for nome in lista_compradores_ativos()}
    itens = []
    for item in carregar_mapa_compradores_editavel():
        caminho = _normalizar_classificacao(
            item.get("Classificação Principal", "")
        )
        comprador_original = str(item.get("Comprador", "Não mapeado")).strip()
        comprador = ativos.get(comprador_original.casefold(), "")
        if caminho and comprador:
            itens.append((caminho, comprador))

    # Mais níveis e caminhos maiores primeiro. Isso impede que
    # PRINCIPAL > ETICOS capture antes de PRINCIPAL > ETICOS > ETICO LINEAR.
    itens.sort(
        key=lambda par: (par[0].count(" > "), len(par[0])),
        reverse=True,
    )
    return itens


def _mapear_comprador(valor):
    """Reconhece o comprador pelo caminho completo da classificação.

    Ordem de decisão:
    1. caminho completo exatamente igual ao cadastro;
    2. caminho cadastrado como ancestral mais específico;
    3. subclassificação recebida como ancestral de um único comprador;
    4. não mapeado quando houver ambiguidade.

    A regra nunca escolhe uma classificação genérica antes de uma
    subclassificação mais detalhada.
    """
    chave = _normalizar_classificacao(valor)
    if not chave:
        return "Não mapeado"

    itens = _mapa_compradores_ordenado()
    mapa_exato = {caminho: comprador for caminho, comprador in itens}

    # 1. A classificação completa é a fonte prioritária.
    if chave in mapa_exato:
        return mapa_exato[chave]

    # 2. Procura o ancestral cadastrado mais específico.
    ancestrais = [
        (caminho, comprador)
        for caminho, comprador in itens
        if chave.startswith(caminho + " > ")
    ]
    if ancestrais:
        maior_nivel = max(caminho.count(" > ") for caminho, _ in ancestrais)
        compradores = {
            comprador
            for caminho, comprador in ancestrais
            if caminho.count(" > ") == maior_nivel
        }
        if len(compradores) == 1:
            return next(iter(compradores))

    # 3. Quando a base vier resumida, aceita apenas se todas as
    # subclassificações abaixo dela pertencem ao mesmo comprador.
    descendentes = {
        comprador
        for caminho, comprador in itens
        if caminho.startswith(chave + " > ")
    }
    if len(descendentes) == 1:
        return next(iter(descendentes))

    return "Não mapeado"


def _atribuir_comprador(df):
    if df is None or df.empty:
        return pd.Series(dtype="object")

    # Prioriza sempre a classificação completa/subclassificada.
    classificacao = _texto_df(
        df,
        [
            "classificacao_resumida", "classificação resumida",
            "classificacao 3 nivel", "classificação 3º nível",
            "classificacao_3_nivel",
            "classificacao principal", "classificação principal",
            "classificacao_principal",
            "classificacao geral", "classificacao_geral",
            "classificacao",
        ],
        "",
    )

    # Recalcula pela classificação em todas as bases. Um comprador antigo
    # salvo no cache não pode prevalecer sobre um caminho atualizado.
    comprador_calculado = classificacao.map(_mapear_comprador)

    # Usa a coluna existente somente quando não existe classificação para
    # recalcular e o valor existente é realmente válido.
    comprador_existente = _texto_df(
        df, ["comprador", "comprador responsavel"], ""
    ).astype(str).str.strip()
    valido_existente = ~comprador_existente.str.upper().isin({
        "", "NÃO MAPEADO", "NAO MAPEADO", "NAN", "NONE", "NULL"
    })
    sem_classificacao = classificacao.astype(str).str.strip().eq("")
    resultado = comprador_calculado.where(
        ~(sem_classificacao & valido_existente),
        comprador_existente,
    )
    return resultado.replace("", "Não mapeado").fillna("Não mapeado")

def _snapshot_seguro(fonte, periodo):
    try:
        return carregar_snapshot(fonte, periodo)
    except Exception:
        return pd.DataFrame()

def _meta_comprador(comprador):
    periodo = METAS_GESTOR.get("periodo_referencia", "")
    individual = obter_meta_comprador(comprador, periodo)

    meta_venda = float(individual.get("meta_venda", 0))
    participacao = float(individual.get("participacao_venda_pct", 0)) / 100.0
    meta_cmv_pct = float(individual.get("meta_cmv_pct", 0))
    meta_cmv = float(individual.get("meta_cmv_valor", 0))
    if meta_cmv <= 0:
        meta_cmv = meta_venda * meta_cmv_pct / 100.0

    fator_cobertura = float(individual.get("fator_cobertura", 0))
    meta_estoque_total = float(individual.get("meta_estoque_total", 0))
    if meta_estoque_total <= 0:
        meta_estoque_total = meta_cmv * fator_cobertura

    return {
        "participacao": participacao,
        "meta_venda": meta_venda,
        "meta_cmv": meta_cmv,
        "meta_estoque_total": meta_estoque_total,
        "meta_curva_a": meta_estoque_total * float(individual.get("meta_curva_a_pct", 0)) / 100.0,
        "meta_curva_b": meta_estoque_total * float(individual.get("meta_curva_b_pct", 0)) / 100.0,
        "meta_curva_c": meta_estoque_total * float(individual.get("meta_curva_c_pct", 0)) / 100.0,
        "meta_curva_d": meta_estoque_total * float(individual.get("meta_curva_d_pct", 0)) / 100.0,
        "meta_ruptura": meta_venda * float(individual.get("meta_ruptura_pct", 0)) / 100.0,
        "meta_entradas": meta_cmv * float(individual.get("meta_reposicao_pct", 0)) / 100.0,
        "meta_cmv_pct": meta_cmv_pct,
        "fator_cobertura": fator_cobertura,
        "meta_curva_a_pct": float(individual.get("meta_curva_a_pct", 0)),
        "meta_curva_b_pct": float(individual.get("meta_curva_b_pct", 0)),
        "meta_curva_c_pct": float(individual.get("meta_curva_c_pct", 0)),
        "meta_curva_d_pct": float(individual.get("meta_curva_d_pct", 0)),
        "meta_ruptura_pct": float(individual.get("meta_ruptura_pct", 0)),
        "meta_reposicao_pct": float(individual.get("meta_reposicao_pct", 0)),
        "valor_premio": float(individual.get("valor_premio", METAS_GESTOR.get("valor_premio_total", 0))),
    }

def _atingimento_maior(real, meta):
    if meta <= 0:
        return 0.0
    return max(0.0, min(real / meta, 1.0)) * 100.0

def _atingimento_menor(real, meta):
    if real <= 0:
        return 100.0
    if meta <= 0:
        return 0.0
    return max(0.0, min(meta / real, 1.0)) * 100.0

def _agregar_vendas(df):
    saida = {}
    if df.empty:
        return saida
    base = df.copy()
    base["Comprador"] = _atribuir_comprador(base)
    valor = _numero_df(base, ["valortotal", "valor total", "faturamento", "valor venda", "receita"])
    custo_unit = _numero_df(base, ["custo", "custo unitario", "custo_unit_r", "cmv"])
    quantidade = _numero_df(base, ["quantidade", "qtd", "quantidade vendida"])
    custo_total_existente = _numero_df(base, ["custo total", "cmv total", "valor custo"])
    custo_total = custo_total_existente.where(custo_total_existente != 0, custo_unit * quantidade)
    base["_venda"] = valor
    base["_cmv"] = custo_total
    for comprador, grupo in base.groupby("Comprador"):
        saida[comprador] = {
            "faturamento": float(grupo["_venda"].sum()),
            "cmv": float(grupo["_cmv"].sum()),
        }
    return saida

def _agregar_entradas(df):
    saida = {}
    if df.empty:
        return saida
    base = df.copy()
    base["Comprador"] = _atribuir_comprador(base)
    valor = _numero_df(
        base,
        [
            "entrada custo total", "entradas custo", "valor entrada",
            "valor nf total", "custo total", "entrada_custo_total"
        ]
    )
    if float(valor.abs().sum()) == 0:
        custo = _numero_df(base, ["custo_final_r", "custo final", "custo"])
        qtd = _numero_df(base, ["quantidade_por_produto", "quantidade", "qtd"])
        valor = custo * qtd
    base["_entrada"] = valor
    for comprador, grupo in base.groupby("Comprador"):
        saida[comprador] = float(grupo["_entrada"].sum())
    return saida

def _agregar_estoque(df):
    saida = {}
    if df.empty:
        return saida

    base = df.copy()
    base["Comprador"] = _atribuir_comprador(base)

    curva = _texto_df(
        base,
        ["curva valor", "curva", "curva abc", "curva qtd", "cabc nome"],
        ""
    ).str.upper().str.strip()

    # Caso a consulta já traga o valor financeiro do estoque, utiliza diretamente.
    valor = _numero_df(
        base,
        [
            "estoque x custo medio", "valor estoque", "estoque a custo",
            "estoque total", "valor_estoque", "valor estoque atual"
        ]
    )

    # No novo script oficial, o valor é:
    # estoque × custo_medio_atual.
    # Se custo médio estiver zerado, usa custo_unit_atual.
    if float(valor.abs().sum()) == 0:
        qtd = _numero_df(
            base,
            ["estoque", "quantidade estoque", "saldo estoque", "qtd estoque"]
        )
        custo_medio = _numero_df(
            base,
            [
                "custo_medio_atual", "custo medio atual",
                "custo médio atual", "customedio", "custo medio"
            ]
        )
        custo_unitario = _numero_df(
            base,
            [
                "custo_unit_atual", "custo unit atual",
                "custo unitario atual", "custo", "custo_final_r"
            ]
        )
        custo_usado = custo_medio.where(custo_medio > 0, custo_unitario)
        valor = qtd * custo_usado

    base["_valor_estoque"] = valor
    base["_curva"] = curva

    for comprador, grupo in base.groupby("Comprador"):
        curvas = {}
        for letra in ["A", "B", "C", "D"]:
            mascara = grupo["_curva"].str.contains(
                rf"(^|[^A-Z]){letra}([^A-Z]|$)",
                regex=True,
                na=False
            )
            curvas[letra] = float(
                grupo.loc[mascara, "_valor_estoque"].sum()
            )

        saida[comprador] = {
            "total": float(grupo["_valor_estoque"].sum()),
            "curvas": curvas,
        }

    return saida

@st.cache_data(ttl=600, show_spinner=False)
def _agregar_ruptura(periodo):
    try:
        base = carregar_ruptura_auto(periodo)
    except Exception:
        base = pd.DataFrame()
    saida = {}
    if base.empty:
        return saida
    # Recalcula inclusive quando a coluna já existe, pois versões anteriores
    # gravavam o texto "Não mapeado" e impediam uma nova classificação.
    base["Comprador"] = _atribuir_comprador(base)
    valor_col = "Valor Ruptura" if "Valor Ruptura" in base.columns else _achar_coluna(base, ["ruptura ativa", "ruptura venda"])
    if isinstance(valor_col, str):
        base["_ruptura"] = pd.to_numeric(base[valor_col], errors="coerce").fillna(0)
    else:
        base["_ruptura"] = 0.0
    for comprador, grupo in base.groupby("Comprador"):
        saida[comprador] = float(grupo["_ruptura"].sum())
    return saida

def _colunas_tabela_cache(con, tabela):
    try:
        return {linha[1] for linha in con.execute(f'PRAGMA table_info("{tabela}")').fetchall()}
    except Exception:
        return set()


@st.cache_data(ttl=3600, show_spinner=False, max_entries=16)
def _resolver_periodo_realizado(periodo_solicitado):
    """Retorna a competência realizada disponível no cache.

    Se o mês solicitado ainda não foi atualizado, utiliza o último mês
    disponível até a competência solicitada.
    """
    periodo_solicitado = str(periodo_solicitado or "")[:7]
    periodos = set()
    con = conexao_cache()
    try:
        for tabela in ["base_vendas", "base_entradas", "base_estoque"]:
            existe = con.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                (tabela,),
            ).fetchone()
            if not existe:
                continue
            colunas = _colunas_tabela_cache(con, tabela)
            if "periodo_referencia" not in colunas:
                continue
            sql = (
                f'SELECT DISTINCT periodo_referencia FROM "{tabela}" '
                "WHERE periodo_referencia IS NOT NULL "
                "AND TRIM(periodo_referencia) <> ''"
            )
            for (valor,) in con.execute(sql).fetchall():
                texto = str(valor)[:7]
                if re.match(r"^\d{4}-\d{2}$", texto):
                    periodos.add(texto)
    finally:
        con.close()

    if periodo_solicitado in periodos:
        return periodo_solicitado
    anteriores = sorted(p for p in periodos if p <= periodo_solicitado)
    if anteriores:
        return anteriores[-1]
    return max(periodos) if periodos else periodo_solicitado


def _agregados_cache_rapidos(periodo):
    """Calcula os KPIs no SQLite sem carregar milhões de linhas na memória."""
    vendas, estoque, entradas, contas = {}, {}, {}, {}
    con = conexao_cache()
    try:
        cols = _colunas_tabela_cache(con, "base_vendas")
        if cols and "periodo_referencia" in cols:
            col_class = next((c for c in ["classificacao_resumida", "classificacao_3_nivel", "classificacao_geral", "classificacao_principal", "classificacao"] if c in cols), None)
            if col_class:
                valor_expr = 'COALESCE(SUM(CAST("valortotal" AS REAL)),0)' if "valortotal" in cols else '0'
                cmv_expr = 'COALESCE(SUM(CAST("custo" AS REAL) * CAST("quantidade" AS REAL)),0)' if "custo" in cols and "quantidade" in cols else '0'
                sql = f"SELECT COALESCE(\"{col_class}\", ''), {valor_expr}, {cmv_expr} FROM base_vendas WHERE periodo_referencia = ? GROUP BY COALESCE(\"{col_class}\", '')"
                for classificacao, faturamento, cmv in con.execute(sql, (periodo,)):
                    comprador = _mapear_comprador(classificacao)
                    item = vendas.setdefault(comprador, {"faturamento": 0.0, "cmv": 0.0})
                    item["faturamento"] += float(faturamento or 0)
                    item["cmv"] += float(cmv or 0)

        cols = _colunas_tabela_cache(con, "base_entradas")
        if cols and "periodo_referencia" in cols:
            col_class = next((c for c in ["classificacao_resumida", "classificacao_3_nivel", "classificacao_geral", "classificacao_principal", "classificacao"] if c in cols), None)
            if col_class:
                if "entrada_custo_total" in cols:
                    valor_expr = 'COALESCE(SUM(CAST("entrada_custo_total" AS REAL)),0)'
                elif "custo_final_r" in cols and "quantidade_por_produto" in cols:
                    valor_expr = 'COALESCE(SUM(CAST("custo_final_r" AS REAL) * CAST("quantidade_por_produto" AS REAL)),0)'
                else:
                    valor_expr = '0'
                sql = f"SELECT COALESCE(\"{col_class}\", ''), {valor_expr} FROM base_entradas WHERE periodo_referencia = ? GROUP BY COALESCE(\"{col_class}\", '')"
                for classificacao, valor in con.execute(sql, (periodo,)):
                    comprador = _mapear_comprador(classificacao)
                    entradas[comprador] = entradas.get(comprador, 0.0) + float(valor or 0)

        cols = _colunas_tabela_cache(con, "base_estoque")
        if cols and "periodo_referencia" in cols:
            col_class = next((c for c in ["classificacao_geral", "classificacao_3_nivel", "classificacao_principal", "classificacao"] if c in cols), None)
            col_curva = next((c for c in ["curva", "curva_abc", "curva_valor"] if c in cols), None)
            col_qtd = next((c for c in ["estoque", "saldo_estoque", "quantidade_estoque"] if c in cols), None)
            col_cmed = next((c for c in ["custo_medio_atual", "custo_medio", "customedio"] if c in cols), None)
            col_cunit = next((c for c in ["custo_unit_atual", "custo", "custo_unitario"] if c in cols), None)
            if col_class and col_qtd:
                curva_expr = f"COALESCE(\"{col_curva}\", '')" if col_curva else "''"
                if col_cmed and col_cunit:
                    custo_expr = f'CASE WHEN COALESCE(CAST(\"{col_cmed}\" AS REAL),0)>0 THEN CAST(\"{col_cmed}\" AS REAL) ELSE COALESCE(CAST(\"{col_cunit}\" AS REAL),0) END'
                elif col_cmed:
                    custo_expr = f'COALESCE(CAST(\"{col_cmed}\" AS REAL),0)'
                elif col_cunit:
                    custo_expr = f'COALESCE(CAST(\"{col_cunit}\" AS REAL),0)'
                else:
                    custo_expr = '0'
                sql = f"SELECT COALESCE(\"{col_class}\", ''), {curva_expr}, COALESCE(SUM(CAST(\"{col_qtd}\" AS REAL) * ({custo_expr})),0) FROM base_estoque WHERE periodo_referencia = ? GROUP BY COALESCE(\"{col_class}\", ''), {curva_expr}"
                for classificacao, curva, valor in con.execute(sql, (periodo,)):
                    comprador = _mapear_comprador(classificacao)
                    item = estoque.setdefault(comprador, {"total": 0.0, "curvas": {"A": 0.0, "B": 0.0, "C": 0.0, "D": 0.0}})
                    valor = float(valor or 0)
                    item["total"] += valor
                    curva_txt = str(curva or '').upper().strip()
                    for letra in ["A", "B", "C", "D"]:
                        if curva_txt == letra or curva_txt.startswith(letra + " ") or curva_txt.startswith(letra + "-"):
                            item["curvas"][letra] += valor
                            break
    except Exception as erro_cache:
        # Uma tabela ausente, coluna incompatível ou SQL inválido não pode
        # derrubar todo o painel. Mantém os agregados já calculados e segue.
        try:
            print(f"Aviso ao consolidar cache SQLite: {erro_cache}")
        except Exception:
            pass
    finally:
        con.close()
    return vendas, estoque, entradas, contas


@st.cache_data(ttl=3600, show_spinner=False, max_entries=16)
def construir_visoes_dinamicas(periodo):
    periodo_realizado = _resolver_periodo_realizado(periodo)
    vendas, estoque, entradas, contas = _agregados_cache_rapidos(periodo_realizado)
    ruptura = _agregar_ruptura(periodo_realizado)

    linhas_real = []
    linhas_meta = []

    total_fat = sum(v.get("faturamento", 0) for v in vendas.values())

    # O Cadastro de Compradores ativos é a fonte oficial da interface.
    # Bases, metas antigas e mapeamentos históricos não podem reintroduzir
    # um comprador removido ou inativado.
    compradores_periodo = sorted(
        {nome for nome in lista_compradores_ativos() if _nome_comprador_valido(nome)},
        key=lambda x: x.casefold(),
    )

    # Novos compradores recebem inicialmente a participação real encontrada
    # na base. O percentual continua editável na Gestão de Metas.
    for nome in compradores_periodo:
        fat_nome = float(vendas.get(nome, {}).get("faturamento", 0.0))
        participacao_inicial = (fat_nome / total_fat * 100.0) if total_fat else 0.0
        garantir_meta_comprador_periodo(nome, periodo, participacao_inicial)

    for comprador in compradores_periodo:
        met = _meta_comprador(comprador)
        fat = vendas.get(comprador, {}).get("faturamento", 0.0)
        cmv = vendas.get(comprador, {}).get("cmv", 0.0)
        est = estoque.get(comprador, {})
        est_total = est.get("total", 0.0)
        curvas = est.get("curvas", {})
        ent = entradas.get(comprador, 0.0)
        rup = ruptura.get(comprador, 0.0)

        rep_fat = (fat / total_fat * 100.0) if total_fat else 0.0
        rep_cmv = (cmv / fat * 100.0) if fat else 0.0
        fator_cob = (est_total / cmv) if cmv else 0.0
        rep_a = (curvas.get("A", 0) / est_total * 100.0) if est_total else 0.0
        rep_b = (curvas.get("B", 0) / est_total * 100.0) if est_total else 0.0
        rep_c = (curvas.get("C", 0) / est_total * 100.0) if est_total else 0.0
        rep_d = (curvas.get("D", 0) / est_total * 100.0) if est_total else 0.0
        rup_pct = (rup / fat * 100.0) if fat else 0.0
        reposicao_pct = (ent / cmv * 100.0) if cmv else 0.0

        linhas_real.append([
            comprador, fat, rep_fat, cmv, rep_cmv, est_total, fator_cob,
            curvas.get("A", 0.0), rep_a,
            curvas.get("B", 0.0), rep_b,
            curvas.get("C", 0.0), rep_c,
            curvas.get("D", 0.0), rep_d,
            rup, rup_pct, ent, reposicao_pct
        ])

        linhas_meta.append([
            comprador,
            met["meta_venda"], met["participacao"] * 100,
            met["meta_cmv"], met["meta_cmv_pct"],
            met["meta_estoque_total"], met["fator_cobertura"],
            met["meta_curva_a"], met["meta_curva_a_pct"],
            met["meta_curva_b"], met["meta_curva_b_pct"],
            met["meta_curva_c"], met["meta_curva_c_pct"],
            met["meta_curva_d"], met["meta_curva_d_pct"],
            met["meta_ruptura"], met["meta_ruptura_pct"],
            met["meta_entradas"], met["meta_reposicao_pct"],
        ])

    colunas_real = [
        "Comprador", "Faturamento Total Atual", "Rep. Faturamento",
        "CMV mês Atual", "Rep. CMV", "Estoque Total", "Fator Cobertura",
        "Estoque Curva A", "Rep. Curva A", "Estoque Curva B", "Rep. Curva B",
        "Estoque Curva C", "Rep. Curva C", "Estoque Curva D", "Rep. Curva D",
        "Ruptura Ativa", "Ruptura %", "Entradas CUSTO", "Reposição CMV %"
    ]
    colunas_meta = [
        "Comprador", "Faturamento Total META", "Rep. Faturamento",
        "CMV mês META", "Rep. CMV", "Estoque Total META", "Fator Cobertura",
        "Estoque Curva A", "Rep. Curva A", "Estoque Curva B", "Rep. Curva B",
        "Estoque Curva C", "Rep. Curva C", "Estoque Curva D", "Rep. Curva D",
        "Ruptura Ativa", "Ruptura %", "Entradas CUSTO", "Reposição CMV %"
    ]

    real = pd.DataFrame(linhas_real, columns=colunas_real)
    metas = pd.DataFrame(linhas_meta, columns=colunas_meta)

    resultado = real.copy()
    pares = [
        ("Faturamento Total Atual", "Faturamento Total META"),
        ("CMV mês Atual", "CMV mês META"),
        ("Estoque Total", "Estoque Total META"),
        ("Estoque Curva A", "Estoque Curva A"),
        ("Estoque Curva B", "Estoque Curva B"),
        ("Estoque Curva C", "Estoque Curva C"),
        ("Estoque Curva D", "Estoque Curva D"),
        ("Ruptura Ativa", "Ruptura Ativa"),
        ("Entradas CUSTO", "Entradas CUSTO"),
    ]
    for atual, meta_col in pares:
        resultado[atual] = metas[meta_col] - real[atual]
    resultado["Rep. Faturamento"] = metas["Rep. Faturamento"] - real["Rep. Faturamento"]
    resultado["Rep. CMV"] = metas["Rep. CMV"] - real["Rep. CMV"]
    resultado["Fator Cobertura"] = metas["Fator Cobertura"] - real["Fator Cobertura"]
    for curva in ["A", "B", "C", "D"]:
        resultado[f"Rep. Curva {curva}"] = metas[f"Rep. Curva {curva}"] - real[f"Rep. Curva {curva}"]
    resultado["Ruptura %"] = metas["Ruptura %"] - real["Ruptura %"]
    resultado["Reposição CMV %"] = metas["Reposição CMV %"] - real["Reposição CMV %"]

    pesos = {
        "Faturamento": float(METAS_GESTOR.get("peso_faturamento", 0)),
        "CMV": float(METAS_GESTOR.get("peso_cmv", 0)),
        "Fator Cobertura": float(METAS_GESTOR.get("peso_fator_cobertura", 0)),
        "Estoque Curva A": float(METAS_GESTOR.get("peso_curva_a", 0)),
        "Estoque Curva B": float(METAS_GESTOR.get("peso_curva_b", 0)),
        "Estoque Curva C": float(METAS_GESTOR.get("peso_curva_c", 0)),
        "Estoque Curva D": float(METAS_GESTOR.get("peso_curva_d", 0)),
        "Ruptura Ativa": float(METAS_GESTOR.get("peso_ruptura", 0)),
        "Reposição CMV": float(METAS_GESTOR.get("peso_reposicao", 0)),
    }
    premio_total = float(METAS_GESTOR.get("valor_premio_total", 0))

    premio_linhas = []
    for i, comprador in enumerate(compradores_periodo):
        r = real.iloc[i]
        m = metas.iloc[i]
        ating = {
            "Faturamento": _atingimento_maior(r["Faturamento Total Atual"], m["Faturamento Total META"]),
            "CMV": _atingimento_menor(r["CMV mês Atual"], m["CMV mês META"]),
            "Fator Cobertura": _atingimento_menor(r["Fator Cobertura"], m["Fator Cobertura"]),
            "Estoque Curva A": _atingimento_maior(r["Estoque Curva A"], m["Estoque Curva A"]),
            "Estoque Curva B": _atingimento_maior(r["Estoque Curva B"], m["Estoque Curva B"]),
            "Estoque Curva C": _atingimento_maior(r["Estoque Curva C"], m["Estoque Curva C"]),
            "Estoque Curva D": _atingimento_maior(r["Estoque Curva D"], m["Estoque Curva D"]),
            "Ruptura Ativa": _atingimento_menor(r["Ruptura Ativa"], m["Ruptura Ativa"]),
            "Reposição CMV": _atingimento_maior(r["Reposição CMV %"], m["Reposição CMV %"]),
        }
        premio_linhas.append([
            comprador,
            premio_total * pesos["Faturamento"]/100 * ating["Faturamento"]/100, ating["Faturamento"],
            premio_total * pesos["CMV"]/100 * ating["CMV"]/100, ating["CMV"],
            premio_total * pesos["Fator Cobertura"]/100 * ating["Fator Cobertura"]/100, ating["Fator Cobertura"],
            premio_total * pesos["Estoque Curva A"]/100 * ating["Estoque Curva A"]/100, ating["Estoque Curva A"],
            premio_total * pesos["Estoque Curva B"]/100 * ating["Estoque Curva B"]/100, ating["Estoque Curva B"],
            premio_total * pesos["Estoque Curva C"]/100 * ating["Estoque Curva C"]/100, ating["Estoque Curva C"],
            premio_total * pesos["Estoque Curva D"]/100 * ating["Estoque Curva D"]/100, ating["Estoque Curva D"],
            premio_total * pesos["Ruptura Ativa"]/100 * ating["Ruptura Ativa"]/100, ating["Ruptura Ativa"],
            premio_total * pesos["Reposição CMV"]/100 * ating["Reposição CMV"]/100, ating["Reposição CMV"],
        ])

    premio = pd.DataFrame(premio_linhas, columns=[
        "Comprador", "Faturamento Prêmio", "Faturamento Realizado",
        "CMV Prêmio", "CMV Realizado", "Estoque Total Prêmio", "Estoque Total Realizado",
        "Curva A Prêmio", "Curva A Realizado", "Curva B Prêmio", "Curva B Realizado",
        "Curva C Prêmio", "Curva C Realizado", "Curva D Prêmio", "Curva D Realizado",
        "Ruptura Prêmio", "Ruptura Realizado", "Entradas Prêmio", "Entradas Realizado"
    ])

    # Referência segura para o quadro de prêmio por KPI.
    # Usa Francieli quando existir; caso contrário usa o primeiro comprador
    # disponível. Se não houver resultados, mantém os indicadores zerados.
    referencia = None
    if not premio.empty and "Comprador" in premio.columns:
        referencia_francieli = premio.loc[
            premio["Comprador"].astype(str).str.strip().str.casefold() == "francieli"
        ]
        if not referencia_francieli.empty:
            referencia = referencia_francieli.iloc[0]
        else:
            referencia = premio.iloc[0]
    premio_kpi = pd.DataFrame([
        ["Faturamento", pesos["Faturamento"], premio_total*pesos["Faturamento"]/100, referencia["Faturamento Realizado"] if referencia is not None else 0, referencia["Faturamento Prêmio"] if referencia is not None else 0],
        ["CMV", pesos["CMV"], premio_total*pesos["CMV"]/100, referencia["CMV Realizado"] if referencia is not None else 0, referencia["CMV Prêmio"] if referencia is not None else 0],
        ["Fator Cobertura", pesos["Fator Cobertura"], premio_total*pesos["Fator Cobertura"]/100, referencia["Estoque Total Realizado"] if referencia is not None else 0, referencia["Estoque Total Prêmio"] if referencia is not None else 0],
        ["Estoque Curva A", pesos["Estoque Curva A"], premio_total*pesos["Estoque Curva A"]/100, referencia["Curva A Realizado"] if referencia is not None else 0, referencia["Curva A Prêmio"] if referencia is not None else 0],
        ["Estoque Curva B", pesos["Estoque Curva B"], premio_total*pesos["Estoque Curva B"]/100, referencia["Curva B Realizado"] if referencia is not None else 0, referencia["Curva B Prêmio"] if referencia is not None else 0],
        ["Estoque Curva C", pesos["Estoque Curva C"], premio_total*pesos["Estoque Curva C"]/100, referencia["Curva C Realizado"] if referencia is not None else 0, referencia["Curva C Prêmio"] if referencia is not None else 0],
        ["Estoque Curva D", pesos["Estoque Curva D"], premio_total*pesos["Estoque Curva D"]/100, referencia["Curva D Realizado"] if referencia is not None else 0, referencia["Curva D Prêmio"] if referencia is not None else 0],
        ["Ruptura Ativa", pesos["Ruptura Ativa"], premio_total*pesos["Ruptura Ativa"]/100, referencia["Ruptura Realizado"] if referencia is not None else 0, referencia["Ruptura Prêmio"] if referencia is not None else 0],
        ["Reposição CMV", pesos["Reposição CMV"], premio_total*pesos["Reposição CMV"]/100, referencia["Entradas Realizado"] if referencia is not None else 0, referencia["Entradas Prêmio"] if referencia is not None else 0],
    ], columns=["KPI", "Peso sobre a meta", "Prêmio por KPI atingível", "Atingimento %", "Prêmio Atingido"])

    # As fontes acima já foram consolidadas diretamente no SQLite.
    # Por isso, o quadro de status deve usar as estruturas agregadas e não
    # DataFrames completos, evitando carregar milhões de linhas na memória.
    status_fontes = {
        "Vendas": contar_registros_cache(
            "base_vendas", periodo_realizado
        ),
        "Estoque": contar_registros_cache(
            "base_estoque", periodo_realizado
        ),
        "Entradas": contar_registros_cache(
            "base_entradas", periodo_realizado
        ),
        "Contas a Pagar": contar_registros_cache(
            "base_contas_pagar", periodo_realizado
        ),
        "Ruptura": len(carregar_ruptura_auto(periodo_realizado))
        if "carregar_ruptura_auto" in globals()
        else 0,
    }
    return real, metas, resultado, premio, premio_kpi, status_fontes, periodo_realizado

def _listar_competencias_globais():
    """Lista competências para visualização sem depender de uma única fonte."""
    periodos = set()

    periodo_meta = str(METAS_GESTOR.get("periodo_referencia", "")).strip()[:7]
    if re.match(r"^\d{4}-\d{2}$", periodo_meta):
        periodos.add(periodo_meta)

    try:
        for item in carregar_historico():
            periodo = str(item.get("periodo_referencia", "")).strip()[:7]
            if re.match(r"^\d{4}-\d{2}$", periodo):
                periodos.add(periodo)
    except Exception:
        pass

    for arquivo in [
        DATA_DIR / "metas_lojas.json",
        DATA_DIR / "metas_por_comprador.json",
    ]:
        try:
            dados = json.loads(arquivo.read_text(encoding="utf-8"))
            lista = dados.get("metas", []) if isinstance(dados, dict) else dados
            for item in lista if isinstance(lista, list) else []:
                periodo = str(
                    item.get("periodo_referencia")
                    or item.get("competencia")
                    or ""
                ).strip()[:7]
                if re.match(r"^\d{4}-\d{2}$", periodo):
                    periodos.add(periodo)
        except Exception:
            pass

    # Disponibiliza os doze meses do ano vigente, mesmo antes da carga da base.
    ano_vigente = int(str(periodo_meta or date.today().strftime("%Y-%m"))[:4])
    for mes_numero in range(1, 13):
        periodos.add(f"{ano_vigente:04d}-{mes_numero:02d}")

    return sorted(periodos, reverse=True)


def _meta_para_competencia(periodo):
    """Retorna a meta correspondente ao mês selecionado."""
    periodo = str(periodo)[:7]
    candidatos = []
    try:
        candidatos.extend(carregar_historico())
    except Exception:
        pass
    candidatos.append(METAS_GESTOR)

    for item in candidatos:
        if str(item.get("periodo_referencia", ""))[:7] == periodo:
            return dict(item)

    # Se ainda não houver uma meta específica, usa os valores atuais como
    # modelo, alterando somente a vigência. O usuário poderá salvá-la na
    # Gestão de Metas.
    meta = dict(METAS_GESTOR)
    ano, mes = map(int, periodo.split("-"))
    import calendar as _calendar
    meta["periodo_referencia"] = periodo
    meta["data_inicio"] = f"{periodo}-01"
    meta["data_fim"] = f"{periodo}-{_calendar.monthrange(ano, mes)[1]:02d}"
    return meta


_COMPETENCIAS_GLOBAIS = _listar_competencias_globais()
_PERIODO_PADRAO_GLOBAL = str(
    METAS_GESTOR.get("periodo_referencia", date.today().strftime("%Y-%m"))
)[:7]

with st.sidebar:
    st.markdown("#### Mês de visualização")
    PERIODO_GLOBAL_SELECIONADO = st.selectbox(
        "Competência global",
        _COMPETENCIAS_GLOBAIS,
        index=(
            _COMPETENCIAS_GLOBAIS.index(_PERIODO_PADRAO_GLOBAL)
            if _PERIODO_PADRAO_GLOBAL in _COMPETENCIAS_GLOBAIS
            else 0
        ),
        label_visibility="collapsed",
        key="periodo_global_dashboard",
        help=(
            "Aplica-se às telas operacionais, metas e premiações. "
            "A Análise Comercial utiliza o próprio seletor anual."
        ),
    )
    st.caption(
        "Filtro global do projeto. Não altera a Análise Comercial."
    )

METAS_GESTOR = _meta_para_competencia(PERIODO_GLOBAL_SELECIONADO)
st.session_state["periodo_gestao_unificado_global"] = PERIODO_GLOBAL_SELECIONADO

PERIODO_DASHBOARD = PERIODO_GLOBAL_SELECIONADO
_TOKEN_VISOES = _arquivo_token(CACHE_DB_FILE, RUPTURA_AUTO_DB, METAS_FILE, COMPRADORES_FILE, MAPA_COMPRADORES_FILE)
_CHAVE_VISOES = f"{PERIODO_DASHBOARD}|{_TOKEN_VISOES}"
if st.session_state.get("_chave_visoes") != _CHAVE_VISOES:
    st.session_state["_dados_visoes"] = construir_visoes_dinamicas(PERIODO_DASHBOARD)
    st.session_state["_chave_visoes"] = _CHAVE_VISOES
REALIZADOS, METAS, RESULTADO, PREMIO, PREMIO_KPI, STATUS_FONTES_DINAMICAS, PERIODO_REALIZADO_USADO = st.session_state["_dados_visoes"]

if PERIODO_REALIZADO_USADO != PERIODO_DASHBOARD:
    st.warning(
        f"A competência {PERIODO_DASHBOARD} ainda não possui Vendas, Entradas e Estoque "
        f"atualizados no cache. Os indicadores realizados abaixo usam temporariamente "
        f"{PERIODO_REALIZADO_USADO}. As metas continuam sendo as de {PERIODO_DASHBOARD}. "
        "Atualize o mês no módulo Banco de Dados para substituir o realizado provisório."
    )
# Segurança final: todas as visões operacionais exibem somente compradores ativos.
REALIZADOS = filtrar_dataframe_compradores_ativos(REALIZADOS)
METAS = filtrar_dataframe_compradores_ativos(METAS)
RESULTADO = filtrar_dataframe_compradores_ativos(RESULTADO)
PREMIO = filtrar_dataframe_compradores_ativos(PREMIO)

# Filtros e seletores seguem os compradores reconhecidos nas bases do período.
COMPRADORES = sorted(
    {
        str(x).strip() for x in REALIZADOS.get("Comprador", pd.Series(dtype=str)).tolist()
        if _nome_comprador_valido(x)
    },
    key=lambda x: x.casefold(),
)


@st.cache_data(ttl=600, show_spinner=False, max_entries=8)
def _ler_cache_analise_cached(tabela, token):
    try:
        with sqlite3.connect(CACHE_DB_FILE, timeout=30) as con:
            existe = con.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                (tabela,),
            ).fetchone()
            if not existe:
                return pd.DataFrame()
            return pd.read_sql_query(f'SELECT * FROM "{tabela}"', con)
    except Exception:
        return pd.DataFrame()

# =========================================================
# EIROX DESIGN SYSTEM
# =========================================================

st.markdown("""
<style>
:root{
    --bg:#050b13;
    --panel:#0c1724;
    --panel2:#101f30;
    --line:#1f3850;
    --text:#f3f7fb;
    --muted:#8da2b8;
    --cyan:#22d3ee;
    --blue:#2f80ed;
    --green:#31d07f;
    --gold:#f8c24e;
    --red:#ff6b74;
}
.stApp{
    background:
      radial-gradient(circle at 15% 0%,rgba(47,128,237,.18),transparent 28%),
      radial-gradient(circle at 85% 0%,rgba(34,211,238,.12),transparent 22%),
      linear-gradient(180deg,#07111c 0%,#050b13 100%);
    color:var(--text);
}
[data-testid="stSidebar"]{
    background:linear-gradient(180deg,#07121f,#0c1a2b 58%,#07111d);
    border-right:1px solid #18324a;
}
[data-testid="stSidebar"] *{color:#eef6ff!important;}
.block-container{max-width:1820px;padding:1rem 1.35rem 2rem;}
.eirox-shell{
    background:linear-gradient(135deg,rgba(15,38,60,.96),rgba(7,19,31,.98));
    border:1px solid #204663;
    border-radius:24px;
    padding:26px 28px;
    margin-bottom:16px;
    box-shadow:0 22px 55px rgba(0,0,0,.36);
    position:relative;
    overflow:hidden;
}
.eirox-shell:after{
    content:"";
    position:absolute;
    width:260px;height:260px;
    right:-80px;top:-90px;
    background:radial-gradient(circle,rgba(34,211,238,.25),transparent 66%);
}
.brand-row{display:flex;align-items:center;justify-content:space-between;gap:18px;}
.eirox-brand{display:flex;align-items:center;gap:14px;}
.eirox-mark{
    width:54px;height:54px;border-radius:16px;
    display:flex;align-items:center;justify-content:center;
    background:linear-gradient(135deg,#22d3ee,#2f80ed);
    box-shadow:0 10px 25px rgba(47,128,237,.35);
    color:white;font-weight:950;font-size:22px;letter-spacing:-1px;
}
.eirox-title small{display:block;color:#6fdff1;text-transform:uppercase;font-weight:800;letter-spacing:1.5px;font-size:11px}
.eirox-title h1{margin:2px 0 0;color:white!important;font-size:34px}
.eirox-badge{
    border:1px solid #2a5676;border-radius:999px;padding:8px 12px;
    color:#b9e9f4;background:rgba(14,44,66,.65);font-weight:800;font-size:12px;
}
.kpi-grid{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:12px;margin-bottom:16px}
.kpi-card{
    background:linear-gradient(180deg,rgba(16,31,48,.98),rgba(10,23,37,.98));
    border:1px solid #1d405d;border-radius:18px;padding:16px;
    box-shadow:0 12px 28px rgba(0,0,0,.22);
}
.kpi-card .label{color:#91a8bd;font-size:11px;text-transform:uppercase;letter-spacing:.7px;font-weight:800}
.kpi-card .value{font-size:25px;font-weight:950;color:white;margin-top:6px}
.kpi-card .sub{font-size:12px;color:#56d7ec;margin-top:4px}
.meta-card{
    background:linear-gradient(180deg,#0e1d2d,#0a1724);
    border:1px solid #1d3f5b;border-radius:18px;overflow:hidden;
    box-shadow:0 12px 28px rgba(0,0,0,.20);min-height:100%;
}
.meta-card.premium{border-color:#80661c;box-shadow:0 12px 28px rgba(248,194,78,.08)}
.meta-card-title{
    padding:12px 10px;text-align:center;font-weight:900;color:white;
    background:linear-gradient(180deg,#16324a,#102338);border-bottom:1px solid #244c69;
}
.meta-card.premium .meta-card-title{background:linear-gradient(180deg,#4e3d12,#30270f);border-color:#80661c}
.meta-line{display:grid;grid-template-columns:minmax(0,1fr) 112px;min-height:38px;align-items:center;border-bottom:1px solid #142d42}
.meta-line span{padding:8px 10px;color:#d8e4ef;font-weight:700;font-size:13px}
.meta-line strong{
    height:100%;display:flex;align-items:center;justify-content:flex-end;
    padding:8px 10px;color:#07111c;font-size:13px;
    background:linear-gradient(180deg,#8df4ff,#40d9ef);border-left:1px solid #2a7890;
}
.meta-card.premium .meta-line strong{background:linear-gradient(180deg,#ffe58f,#f8c24e);border-color:#9b7413}
.section-title{
    border-radius:16px 16px 0 0;padding:11px 16px;text-align:center;
    font-weight:950;letter-spacing:.3px;color:white;border:1px solid #284a65;border-bottom:none;
}
.sec-gray{background:linear-gradient(180deg,#26384a,#1b2a38)}
.sec-green{background:linear-gradient(180deg,#17613e,#0f412a)}
.sec-blue{background:linear-gradient(180deg,#185078,#123a58)}
.sec-gold{background:linear-gradient(180deg,#806117,#523e0f)}
[data-testid="stDataFrame"]{
    background:#0a1724;border:1px solid #25445f;border-radius:0 0 16px 16px;overflow:hidden;
    box-shadow:0 12px 28px rgba(0,0,0,.18);
}
[data-testid="stMetric"]{
    background:linear-gradient(180deg,#0f1f30,#0a1724);
    border:1px solid #22425f;padding:14px;border-radius:16px;
}
.premium-box{
    display:flex;align-items:center;justify-content:space-between;
    background:linear-gradient(135deg,#12263a,#0b1927);
    border:1px solid #2d526f;border-radius:18px;padding:16px 18px;margin-bottom:12px;
    box-shadow:0 12px 28px rgba(0,0,0,.20)
}
.premium-box .value{
    background:linear-gradient(180deg,#ffe58f,#f8c24e);
    color:#1d1605;border-radius:12px;padding:10px 28px;font-size:22px;font-weight:950
}
.eirox-footer{margin-top:18px;color:#7890a5;font-size:12px;text-align:center;letter-spacing:.4px}
@media(max-width:1200px){.kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))}}
@media(max-width:700px){.kpi-grid{grid-template-columns:1fr}.brand-row{align-items:flex-start;flex-direction:column}.eirox-title h1{font-size:28px}}

.sidebar-logo img { display:block; }
[data-testid="stSidebar"] [data-testid="stRadio"] label {
    padding: 7px 9px;
    border-radius: 10px;
    margin-bottom: 3px;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
    background: rgba(255,0,0,.08);
}

</style>
""", unsafe_allow_html=True)

with st.sidebar:
    if LOGO_ECONOMIZE_B64:
        st.markdown(
            f"""
            <div style="display:flex;justify-content:center;padding:4px 0 12px;">
                <img src="data:image/png;base64,{LOGO_ECONOMIZE_B64}"
                     style="width:235px;max-width:100%;object-fit:contain;">
            </div>
            """,
            unsafe_allow_html=True
        )

    st.caption("Gestão de Performance Comercial")
    st.caption("⚡ Cache otimizado ativo")
    st.divider()

    visao_label = st.radio(
        "Navegação",
        [
            "📌 Metas e Parâmetros",
            "👔 Resumo CEO",
            "📚 Análise Comercial",
            "📊 Realizados",
            "🎯 Métricas Destaque",
            "📈 Resultado Métricas",
            "🏆 Prêmio Comprador",
            "💰 Prêmio por KPI",
            "🏪 Metas de Loja",
            "🧭 Gestão de Metas",
            "📥 Importar Ruptura",
            "🗄️ Banco de Dados",
            "👥 Compradores por Classificação",
            "🧑‍💼 Cadastro de Compradores",
        ],
        label_visibility="collapsed"
    )

    mapa_visoes = {
        "📌 Metas e Parâmetros": "Metas e Parâmetros",
        "👔 Resumo CEO": "Resumo CEO",
        "📚 Análise Comercial": "Análise Comercial",
        "📊 Realizados": "Realizados",
        "🎯 Métricas Destaque": "Métricas Destaque",
        "📈 Resultado Métricas": "Resultado Métricas",
        "🏆 Prêmio Comprador": "Prêmio Comprador",
        "💰 Prêmio por KPI": "Prêmio por KPI",
        "🏪 Metas de Loja": "Metas de Loja",
        "🧭 Gestão de Metas": "Gestão de Metas",
        "📥 Importar Ruptura": "Importar Ruptura",
        "🗄️ Banco de Dados": "Banco de Dados",
        "👥 Compradores por Classificação": "Compradores por Classificação",
        "🧑‍💼 Cadastro de Compradores": "Cadastro de Compradores",
    }
    visao = mapa_visoes[visao_label]

    st.markdown("### Filtros")
    comprador = st.selectbox("Comprador", ["Todos"] + COMPRADORES)

    st.markdown("#### Período ativo")
    st.info(
        f"{METAS_GESTOR.get('periodo_referencia','-')}\\n\\n"
        f"{data_br(METAS_GESTOR.get('data_inicio',''))} a "
        f"{data_br(METAS_GESTOR.get('data_fim',''))}"
    )

    with st.expander("Status das fontes", expanded=False):
        st.write("**Vendas:** banco de dados")
        st.write("**Estoque:** banco de dados")
        st.write("**Entradas:** banco de dados")
        st.write("**Ruptura:** automática por pasta")
        st.write("**Snapshots mensais:** SQLite local")

    st.divider()
    st.caption("Rede Economize • Enterprise")

iniciar_contexto_exportacao(visao, METAS_GESTOR.get("periodo_referencia", "-"))

st.markdown(f"""
<div class="eirox-shell">
  <div class="brand-row">
    <div style="display:flex;align-items:center;gap:24px;min-width:0;">
      <img src="data:image/png;base64,{LOGO_ECONOMIZE_B64}"
           style="width:290px;max-width:38vw;object-fit:contain;filter:drop-shadow(0 10px 16px rgba(0,0,0,.28));">
      <div class="eirox-title">
        <small>Rede Economize</small>
        <h1>Performance Comercial</h1>
      </div>
    </div>
    <div style="display:flex;gap:8px;flex-wrap:wrap">
      <div class="eirox-badge">Enterprise Edition</div>
      <div class="eirox-badge">Período: {METAS_GESTOR.get("periodo_referencia","-")}</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

fat = REALIZADOS["Faturamento Total Atual"].sum()
cmv = REALIZADOS["CMV mês Atual"].sum()
estoque = REALIZADOS["Estoque Total"].sum()
ruptura = RUPTURA_IMPORTADA["Ruptura Ativa"].sum() if not RUPTURA_IMPORTADA.empty else REALIZADOS["Ruptura Ativa"].sum()
reposicao = REALIZADOS["Reposição CMV %"].mean()

st.markdown(f"""
<div class="kpi-grid">
  <div class="kpi-card"><div class="label">Faturamento Total</div><div class="value">{moeda_real(fat)}</div><div class="sub">Meta: {moeda_real(METAS_GESTOR["meta_venda_total_mes"])}</div></div>
  <div class="kpi-card"><div class="label">CMV Atual</div><div class="value">R$ {moeda(cmv)}</div><div class="sub">68,9% do faturamento</div></div>
  <div class="kpi-card"><div class="label">Estoque Total</div><div class="value">R$ {moeda(estoque)}</div><div class="sub">Cobertura consolidada</div></div>
  <div class="kpi-card"><div class="label">Ruptura Ativa</div><div class="value">{moeda_real(ruptura)}</div><div class="sub">Meta operacional: {percentual(METAS_GESTOR["meta_ruptura"])}</div></div>
  <div class="kpi-card"><div class="label">Reposição CMV</div><div class="value">{percentual(reposicao)}</div><div class="sub">Meta: {percentual(METAS_GESTOR["meta_reposicao"])}</div></div>
</div>
""", unsafe_allow_html=True)

# =========================================================
# GRÁFICOS EXECUTIVOS
# =========================================================

dados_grafico = REALIZADOS.copy()
metas_grafico = METAS.copy()

if comprador != "Todos":
    dados_grafico = dados_grafico[dados_grafico["Comprador"] == comprador]
    metas_grafico = metas_grafico[metas_grafico["Comprador"] == comprador]

col_g1, col_g2 = st.columns([1.1, 0.9], gap="large")

with col_g1:
    st.markdown("### Faturamento realizado x meta")
    df_fat = dados_grafico[["Comprador", "Faturamento Total Atual"]].merge(
        metas_grafico[["Comprador", "Faturamento Total META"]],
        on="Comprador",
        how="left"
    )
    df_fat = df_fat.melt(
        id_vars="Comprador",
        var_name="Indicador",
        value_name="Valor"
    )
    fig_fat = px.bar(
        df_fat,
        x="Comprador",
        y="Valor",
        color="Indicador",
        barmode="group",
        text_auto=False
    )
    fig_fat.update_layout(
        height=350,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font_color="#f3f7fb",
        margin=dict(l=10, r=10, t=15, b=10),
        legend_title_text="",
        yaxis_title="R$",
        xaxis_title=""
    )
    fig_fat.update_yaxes(gridcolor="rgba(255,255,255,.08)")
    plotly_chart_br(fig_fat, use_container_width=True, config={"displayModeBar": False})

with col_g2:
    st.markdown("### Composição do estoque")
    df_curvas = pd.DataFrame({
        "Curva": ["Curva A", "Curva B", "Curva C", "Curva D"],
        "Valor": [
            dados_grafico["Estoque Curva A"].sum(),
            dados_grafico["Estoque Curva B"].sum(),
            dados_grafico["Estoque Curva C"].sum(),
            dados_grafico["Estoque Curva D"].sum(),
        ]
    })
    fig_curvas = px.pie(
        df_curvas,
        names="Curva",
        values="Valor",
        hole=0.58
    )
    fig_curvas.update_traces(textposition="inside", textinfo="percent+label")
    fig_curvas.update_layout(
        height=350,
        paper_bgcolor="rgba(0,0,0,0)",
        font_color="#f3f7fb",
        margin=dict(l=10, r=10, t=15, b=10),
        legend_title_text=""
    )
    plotly_chart_br(fig_curvas, use_container_width=True, config={"displayModeBar": False})

col_g3, col_g4 = st.columns(2, gap="large")

with col_g3:
    st.markdown("### Reposição CMV por comprador")
    fig_rep = px.bar(
        dados_grafico.sort_values("Reposição CMV %"),
        x="Reposição CMV %",
        y="Comprador",
        orientation="h",
        text="Reposição CMV %"
    )
    fig_rep.add_vline(
        x=METAS_GESTOR["meta_reposicao"],
        line_dash="dash",
        annotation_text="Meta"
    )
    fig_rep.update_traces(
        text=[
            percentual(valor)
            for valor in dados_grafico.sort_values(
                "Reposição CMV %"
            )["Reposição CMV %"]
        ],
        texttemplate="%{text}",
        textposition="outside",
        hovertemplate=(
            "Comprador=%{y}<br>"
            "Reposição CMV=%{text}<extra></extra>"
        ),
    )
    fig_rep.update_layout(
        height=300,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font_color="#f3f7fb",
        margin=dict(l=10, r=25, t=15, b=10),
        xaxis_title="Percentual",
        yaxis_title=""
    )
    fig_rep.update_xaxes(gridcolor="rgba(255,255,255,.08)")
    plotly_chart_br(
        fig_rep,
        use_container_width=True,
        config={"displayModeBar": False},
        tipo="percentual",
    )

with col_g4:
    st.markdown("### Ruptura ativa por comprador")
    if (
        not RUPTURA_IMPORTADA.empty
        and "Comprador" in RUPTURA_IMPORTADA.columns
        and RUPTURA_IMPORTADA["Comprador"].astype(str).str.strip().ne("").any()
    ):
        df_rup = RUPTURA_IMPORTADA.copy()
        df_rup["Comprador"] = df_rup["Comprador"].astype(str).str.strip()
        df_rup = (
            df_rup[df_rup["Comprador"] != ""]
            .groupby("Comprador", as_index=False)["Ruptura Ativa"]
            .sum()
        )
    else:
        df_rup = dados_grafico[["Comprador", "Ruptura Ativa"]].copy()

    fig_rup = px.bar(
        df_rup.sort_values("Ruptura Ativa"),
        x="Ruptura Ativa",
        y="Comprador",
        orientation="h",
        text_auto=False
    )
    fig_rup.update_layout(
        height=300,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font_color="#f3f7fb",
        margin=dict(l=10, r=25, t=15, b=10),
        xaxis_title="R$",
        yaxis_title=""
    )
    fig_rup.update_xaxes(gridcolor="rgba(255,255,255,.08)")
    plotly_chart_br(fig_rup, use_container_width=True, config={"displayModeBar": False})


st.markdown("### Status das bases do período")
status_cols = st.columns(5)
for coluna, (fonte, registros) in zip(
    status_cols,
    STATUS_FONTES_DINAMICAS.items(),
):
    with coluna:
        card_status_base(fonte, registros)

def listar_periodos_gestao_metas():
    periodos = set()
    atual = str(METAS_GESTOR.get("periodo_referencia", "")).strip()
    if atual:
        periodos.add(atual)
    for item in carregar_historico():
        valor = str(item.get("periodo_referencia", "")).strip()
        if valor:
            periodos.add(valor)
    for item in carregar_metas_por_comprador():
        valor = str(item.get("periodo_referencia", "")).strip()
        if valor:
            periodos.add(valor)
    df_lojas = dataframe_metas_lojas()
    if not df_lojas.empty and "periodo_referencia" in df_lojas.columns:
        periodos.update(df_lojas["periodo_referencia"].dropna().astype(str).str.strip())
    return sorted([p for p in periodos if p], reverse=True)


if visao == "Metas e Parâmetros":
    c1, c2, c3, c4, c5 = st.columns([1.35,.9,1.15,1.15,1.65], gap="small")
    with c1:
        card_meta("Metas por Tipo", [
            ("Meta Venda Total Mês", moeda_real(METAS_GESTOR["meta_venda_total_mes"])),
            ("Meta CMV Mês", percentual(METAS_GESTOR["meta_cmv_mes"])),
            ("Fator Redução CMV", br_num(METAS_GESTOR["fator_reducao_cmv"], 2)),
            ("Fator Cobertura", br_num(METAS_GESTOR["fator_cobertura"], 2)),
        ])
    with c2:
        card_meta("Metas Operacionais", [
            ("Meta Ruptura", percentual(METAS_GESTOR["meta_ruptura"])),
            ("Meta Reposição", percentual(METAS_GESTOR["meta_reposicao"])),
        ])
    with c3:
        card_meta("Curvas de Estoque", [
            ("Curva A", percentual(METAS_GESTOR["curva_a"])),
            ("Curva B", percentual(METAS_GESTOR["curva_b"])),
            ("Curva C", percentual(METAS_GESTOR["curva_c"])),
            ("Curva D", percentual(METAS_GESTOR["curva_d"])),
        ])
    with c4:
        metas_part_periodo = [
            item for item in carregar_metas_por_comprador()
            if str(item.get("periodo_referencia", "")) == str(PERIODO_DASHBOARD)
            and str(item.get("comprador", "")).strip().casefold() in _conjunto_compradores_ativos()
            and _nome_comprador_valido(item.get("comprador", ""))
        ]
        linhas_rep = [
            (str(item.get("comprador", "")), percentual(float(item.get("participacao_venda_pct", 0))))
            for item in sorted(metas_part_periodo, key=lambda x: str(x.get("comprador", "")).casefold())
        ]
        linhas_rep.append(("Total", percentual(sum(float(x.get("participacao_venda_pct", 0)) for x in metas_part_periodo))))
        card_meta("Rep. Venda Comprador", linhas_rep)
    with c5:
        card_meta("Pesos sobre Prêmio", [
            ("Faturamento", percentual(METAS_GESTOR["peso_faturamento"])),
            ("CMV", percentual(METAS_GESTOR["peso_cmv"])),
            ("Fator Cobertura", percentual(METAS_GESTOR["peso_fator_cobertura"])),
            ("Estoque Curva A", percentual(METAS_GESTOR["peso_curva_a"])),
            ("Estoque Curva B", percentual(METAS_GESTOR["peso_curva_b"])),
            ("Estoque Curva C", percentual(METAS_GESTOR["peso_curva_c"])),
            ("Estoque Curva D", percentual(METAS_GESTOR["peso_curva_d"])),
            ("Ruptura Ativa", percentual(METAS_GESTOR["peso_ruptura"])),
            ("Reposição CMV", percentual(METAS_GESTOR["peso_reposicao"])),
            ("Peso Total", percentual(
                METAS_GESTOR["peso_faturamento"] + METAS_GESTOR["peso_cmv"] +
                METAS_GESTOR["peso_fator_cobertura"] + METAS_GESTOR["peso_curva_a"] +
                METAS_GESTOR["peso_curva_b"] + METAS_GESTOR["peso_curva_c"] +
                METAS_GESTOR["peso_curva_d"] + METAS_GESTOR["peso_ruptura"] +
                METAS_GESTOR["peso_reposicao"]
            )),
        ], destaque=True)

elif visao == "Resumo CEO":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:22px">Resumo Executivo CEO — Meta x Realizado</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">Visão consolidada de todas as metas do período ativo.</div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Executive Overview</div>
    </div>
    """, unsafe_allow_html=True)

    real_ceo = filtrar_dataframe_compradores_ativos(REALIZADOS)
    meta_ceo = filtrar_dataframe_compradores_ativos(METAS)
    def _soma_ceo(df, coluna):
        return float(pd.to_numeric(df.get(coluna, pd.Series(dtype=float)), errors="coerce").fillna(0).sum())

    indicadores_ceo = [
        ("Faturamento", _soma_ceo(meta_ceo, "Faturamento Total META"), _soma_ceo(real_ceo, "Faturamento Total Atual"), "maior"),
        ("CMV", _soma_ceo(meta_ceo, "CMV mês META"), _soma_ceo(real_ceo, "CMV mês Atual"), "menor"),
        ("Estoque Total", _soma_ceo(meta_ceo, "Estoque Total META"), _soma_ceo(real_ceo, "Estoque Total"), "menor"),
        ("Ruptura Ativa", _soma_ceo(meta_ceo, "Ruptura Ativa"), _soma_ceo(real_ceo, "Ruptura Ativa"), "menor"),
        ("Reposição / Entradas", _soma_ceo(meta_ceo, "Entradas CUSTO"), _soma_ceo(real_ceo, "Entradas CUSTO"), "maior"),
    ]
    linhas_ceo = []
    for nome, meta_v, real_v, regra in indicadores_ceo:
        ating = ((real_v / meta_v) if regra == "maior" else (meta_v / real_v if real_v else 0)) * 100 if meta_v else 0
        linhas_ceo.append({"Indicador": nome, "Meta": meta_v, "Realizado": real_v, "Atingimento (%)": ating, "Status": "✅ Atingida" if ating >= 100 else ("🟡 Atenção" if ating >= 90 else "🔴 Abaixo")})
    df_ceo = pd.DataFrame(linhas_ceo)

    cards = st.columns(5)
    for col, linha in zip(cards, linhas_ceo):
        col.metric(linha["Indicador"], moeda_real(linha['Realizado']), f"{percentual(linha['Atingimento (%)'])} da meta")

    c1, c2 = st.columns([1.35, 1])
    with c1:
        fig = go.Figure()
        fig.add_bar(name="Meta", x=df_ceo["Indicador"], y=df_ceo["Meta"])
        fig.add_bar(name="Realizado", x=df_ceo["Indicador"], y=df_ceo["Realizado"])
        fig.update_layout(barmode="group", height=390, title="Meta x Realizado — Consolidado", margin=dict(l=10, r=10, t=55, b=10))
        plotly_chart_br(fig, use_container_width=True)
    with c2:
        st.markdown("### Semáforo das metas")
        dataframe_br(df_ceo[["Indicador", "Atingimento (%)", "Status"]], use_container_width=True, hide_index=True, height=390, column_config={"Atingimento (%)": st.column_config.ProgressColumn(min_value=0, max_value=120, format="%.1f%%")})

    st.markdown("### Meta x realizado por comprador ativo")
    comp_ceo = real_ceo.merge(meta_ceo, on="Comprador", how="outer").fillna(0)
    if not comp_ceo.empty:
        comp_ceo["Atingimento Faturamento (%)"] = comp_ceo.apply(lambda r: (r.get("Faturamento Total Atual", 0) / r.get("Faturamento Total META", 0) * 100) if r.get("Faturamento Total META", 0) else 0, axis=1)
        cols_ceo = [c for c in ["Comprador", "Faturamento Total META", "Faturamento Total Atual", "Atingimento Faturamento (%)", "CMV mês META", "CMV mês Atual", "Estoque Total META", "Estoque Total"] if c in comp_ceo.columns]
        dataframe_br(comp_ceo[cols_ceo], use_container_width=True, hide_index=True)

    st.markdown("### Meta x realizado por filial")
    st.caption(
        "Faturamento e margem bruta por filial, considerando as metas "
        "cadastradas no período ativo e o realizado da base de vendas."
    )
    quadro_filiais_ceo = montar_quadro_filiais_ceo(PERIODO_DASHBOARD)
    if quadro_filiais_ceo.empty:
        st.info("Não há metas de filial cadastradas para o período ativo.")
    else:
        dataframe_br(
            quadro_filiais_ceo,
            use_container_width=True,
            hide_index=True,
            height=min(430, 82 + 36 * max(len(quadro_filiais_ceo), 1)),
        )

elif visao == "Análise Comercial":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:22px">Análise de Desempenho Comercial</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">Venda, custo, entrada, CMV, fornecedores, estoque e resultado por área e competência.</div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Análise Agrupada</div>
    </div>
    """, unsafe_allow_html=True)

    cfg_analise = carregar_config_analise_comercial()
    st.caption("⚡ Modo rápido: a tela utiliza resumos persistentes, sem carregar os movimentos brutos.")

    token_analise = _arquivo_token(CACHE_DB_FILE)

    @st.cache_data(ttl=3600, show_spinner=False, max_entries=12)
    def _carregar_resumos_analise(token):
        with conexao_cache() as con_preparo:
            garantir_tabelas_analise(con_preparo)
        with sqlite3.connect(CACHE_DB_FILE, timeout=30) as con:
            reconstruir_posicoes_mensais(con)
            tabelas = {r[0] for r in con.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()}
            # Compatibilidade: se o projeto ainda não tiver os resumos,
            # cria uma única vez. Nas próximas atualizações eles são mantidos
            # automaticamente por período.
            obrigatorias = {
                "analise_vendas_resumo",
                "analise_entradas_resumo",
                "analise_contas_resumo",
                "analise_estoque_resumo",
                "analise_posicao_resumo",
            }
            if not obrigatorias.issubset(tabelas):
                atualizar_resumos_analise(con)

            vendas = pd.read_sql_query(
                "SELECT periodo_referencia, classificacao, venda, custo FROM analise_vendas_resumo",
                con,
            )
            entradas = pd.read_sql_query(
                "SELECT periodo_referencia, classificacao, compra FROM analise_entradas_resumo",
                con,
            )
            contas = pd.read_sql_query(
                "SELECT periodo_referencia, plano_contas, pagamento FROM analise_contas_resumo",
                con,
            )
            estoque = pd.read_sql_query(
                "SELECT periodo_referencia, estoque FROM analise_estoque_resumo",
                con,
            )
            posicao = pd.read_sql_query(
                "SELECT periodo_referencia, contas_pagar, estoque "
                "FROM analise_posicao_resumo",
                con,
            )
        return vendas, entradas, contas, estoque, posicao

    # Normalmente retorna poucas dezenas de linhas, mesmo quando a base bruta
    # possui milhões de movimentos.
    vendas_an, entradas_an, contas_an, estoque_an, posicao_an = _carregar_resumos_analise(token_analise)

    def _normalizar_texto_analise(valor):
        texto = "" if pd.isna(valor) else str(valor)
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        return re.sub(r"\s+", " ", texto).strip().upper()

    def _area_analise(valor):
        normal = _normalizar_texto_analise(valor)
        mapa = [
            ("1 - ETICOS", "01 - Med Propagados"),
            ("ETICOS", "01 - Med Propagados"),
            ("2 - GENERICOS", "02 - Genericos/Similares"),
            ("GENERICOS", "02 - Genericos/Similares"),
            ("SIMILARES", "02 - Genericos/Similares"),
            ("5 - PERFUMARIA", "04 - Perfumaria / Dermocosméticos"),
            ("PERFUMARIA", "04 - Perfumaria / Dermocosméticos"),
            ("DERMOCOSMET", "04 - Perfumaria / Dermocosméticos"),
            ("FRALDAS", "05 - Fraldas"),
            ("LEITES", "06 - Leite"),
            ("LEITE", "06 - Leite"),
            ("HOSPITALARES", "07 - Varejos"),
            ("VAREJO", "07 - Varejos"),
            ("6 - CONVENIENCIA", "08 - Conveniencias"),
            ("CONVENIENCIA", "08 - Conveniencias"),
            ("SUPLEMENT", "09 - Nutracêuticos, Suplementos, Vitaminas"),
            ("NUTRACE", "09 - Nutracêuticos, Suplementos, Vitaminas"),
            ("VITAMIN", "09 - Nutracêuticos, Suplementos, Vitaminas"),
        ]
        for termo, area in mapa:
            if termo in normal:
                return area
        return "11 - Outros"

    if not vendas_an.empty:
        vendas_an["Área"] = vendas_an["classificacao"].map(_area_analise)
        vendas_an["Competência"] = vendas_an["periodo_referencia"].astype(str).str[:7]
        vendas_an["Venda"] = pd.to_numeric(vendas_an["venda"], errors="coerce").fillna(0)
        vendas_an["Custo"] = pd.to_numeric(vendas_an["custo"], errors="coerce").fillna(0)

    if not entradas_an.empty:
        entradas_an["Área"] = entradas_an["classificacao"].map(_area_analise)
        entradas_an["Competência"] = entradas_an["periodo_referencia"].astype(str).str[:7]
        entradas_an["Compra"] = pd.to_numeric(entradas_an["compra"], errors="coerce").fillna(0)

    if not contas_an.empty:
        contas_an["Competência"] = contas_an["periodo_referencia"].astype(str).str[:7]
        contas_an["Valor"] = pd.to_numeric(contas_an["pagamento"], errors="coerce").fillna(0)

    if not estoque_an.empty:
        estoque_an["Competência"] = estoque_an["periodo_referencia"].astype(str).str[:7]
        estoque_an["Estoque"] = pd.to_numeric(
            estoque_an["estoque"], errors="coerce"
        ).fillna(0)

    if not posicao_an.empty:
        posicao_an["Competência"] = posicao_an["periodo_referencia"].astype(str).str[:7]
        posicao_an["Contas a Pagar"] = pd.to_numeric(
            posicao_an["contas_pagar"], errors="coerce"
        ).fillna(0)
        posicao_an["Estoque"] = pd.to_numeric(
            posicao_an["estoque"], errors="coerce"
        ).fillna(0)

    periodos_disponiveis = sorted(set(
        ([x for x in vendas_an.get("Competência", pd.Series(dtype=str)).dropna().astype(str).tolist() if re.match(r"^\\d{4}-\\d{2}$", x)] +
         [x for x in entradas_an.get("Competência", pd.Series(dtype=str)).dropna().astype(str).tolist() if re.match(r"^\\d{4}-\\d{2}$", x)] +
         [x for x in contas_an.get("Competência", pd.Series(dtype=str)).dropna().astype(str).tolist() if re.match(r"^\\d{4}-\\d{2}$", x)] +
         [x for x in estoque_an.get("Competência", pd.Series(dtype=str)).dropna().astype(str).tolist() if re.match(r"^\\d{4}-\\d{2}$", x)] +
         [x for x in posicao_an.get("Competência", pd.Series(dtype=str)).dropna().astype(str).tolist() if re.match(r"^\\d{4}-\\d{2}$", x)])
    ))
    # A Evolução Comercial possui contexto anual próprio e não obedece
    # ao período ativo utilizado nas demais telas.
    ano_atual = date.today().year
    anos_encontrados = sorted(
        {
            int(str(periodo)[:4])
            for periodo in periodos_disponiveis
            if re.match(r"^\d{4}-\d{2}$", str(periodo))
        },
        reverse=True,
    )
    anos_disponiveis = sorted(
        set(anos_encontrados + [ano_atual]),
        reverse=True,
    )

    with st.expander("Filtros e plano de contas", expanded=True):
        f1, f2 = st.columns([1.2, 2])
        with f1:
            ano_selecionado = st.selectbox(
                "Ano da Evolução Comercial",
                anos_disponiveis,
                index=(
                    anos_disponiveis.index(ano_atual)
                    if ano_atual in anos_disponiveis
                    else 0
                ),
                help=(
                    "Este filtro é independente do período ativo das demais telas. "
                    "A análise sempre mostra os 12 meses do ano escolhido."
                ),
            )

            meses_selecionados = [
                f"{int(ano_selecionado):04d}-{mes:02d}"
                for mes in range(1, 13)
            ]

            st.caption(
                f"Exibindo janeiro a dezembro de {ano_selecionado}. "
                "Meses sem movimento permanecem visíveis com valor zero."
            )
        planos_base = []
        if not contas_an.empty and "plano_contas" in contas_an.columns:
            planos_base = sorted(
                contas_an["plano_contas"].dropna().astype(str).str.strip()
                .loc[lambda x: x.ne("")].unique().tolist()
            )

        planos_catalogo = carregar_catalogo_planos_contas()
        planos_salvos = (
            list(cfg_analise.get("planos_contas_selecionados", []))
            + [cfg_analise.get("plano_contas_padrao", PLANO_CONTAS_PAGAMENTO_PADRAO)]
            + list(cfg_analise.get("planos_adicionais", []))
        )
        opcoes_planos = list(dict.fromkeys(
            [
                plano for plano in
                planos_salvos + planos_base + planos_catalogo
                if str(plano).strip()
            ]
        ))

        selecionados_salvos = [
            plano for plano in cfg_analise.get(
                "planos_contas_selecionados",
                [cfg_analise.get(
                    "plano_contas_padrao",
                    PLANO_CONTAS_PAGAMENTO_PADRAO,
                )],
            )
            if plano in opcoes_planos
        ]
        if not selecionados_salvos:
            selecionados_salvos = [PLANO_CONTAS_PAGAMENTO_PADRAO]

        with f2:
            planos_selecionados = st.multiselect(
                "Planos de contas usados em Pagamento de Fornecedor",
                opcoes_planos,
                default=selecionados_salvos,
                help=(
                    "É possível selecionar um ou vários planos. A configuração "
                    "fica salva para as próximas análises."
                ),
            )

        st.markdown("**Planos selecionados nesta análise:**")
        if planos_selecionados:
            for plano_ativo in planos_selecionados:
                st.caption(f"✓ {plano_ativo}")
        else:
            st.warning("Selecione pelo menos um plano de contas.")

        cadd1, cadd2, cadd3 = st.columns([2, 1, 1])
        with cadd1:
            novo_plano = st.text_input(
                "Adicionar outro plano de contas",
                placeholder="Cole o caminho completo do plano",
            )
        with cadd2:
            if st.button("Adicionar plano", use_container_width=True):
                if novo_plano.strip():
                    adicionais = list(cfg_analise.get("planos_adicionais", []))
                    if novo_plano.strip() not in adicionais:
                        adicionais.append(novo_plano.strip())
                    cfg_analise["planos_adicionais"] = adicionais
                    salvar_config_analise_comercial(cfg_analise)
                    st.success("Plano adicionado à lista.")
                    st.rerun()
        with cadd3:
            if st.button("Salvar configuração", use_container_width=True):
                if not planos_selecionados:
                    st.error("Selecione pelo menos um plano.")
                else:
                    cfg_analise["planos_contas_selecionados"] = planos_selecionados
                    cfg_analise["plano_contas_padrao"] = planos_selecionados[0]
                    salvar_config_analise_comercial(cfg_analise)
                    st.success("Configuração de planos salva.")
                    st.rerun()

    if meses_selecionados:
        if not vendas_an.empty:
            vendas_an = vendas_an[vendas_an["Competência"].isin(meses_selecionados)]
        if not entradas_an.empty:
            entradas_an = entradas_an[entradas_an["Competência"].isin(meses_selecionados)]
        if not contas_an.empty:
            contas_an = contas_an[contas_an["Competência"].isin(meses_selecionados)]
        if not estoque_an.empty:
            estoque_an = estoque_an[estoque_an["Competência"].isin(meses_selecionados)]
        if not posicao_an.empty:
            posicao_an = posicao_an[posicao_an["Competência"].isin(meses_selecionados)]

    contas_filtradas = contas_an.copy()
    if not contas_filtradas.empty and "plano_contas" in contas_filtradas.columns:
        alvos = {
            _normalizar_texto_analise(plano)
            for plano in planos_selecionados
        }
        contas_filtradas = contas_filtradas[
            contas_filtradas["plano_contas"]
            .map(_normalizar_texto_analise)
            .isin(alvos)
        ]

    venda_total = float(vendas_an.get("Venda", pd.Series(dtype=float)).sum()) if not vendas_an.empty else 0.0
    custo_total = float(vendas_an.get("Custo", pd.Series(dtype=float)).sum()) if not vendas_an.empty else 0.0
    entrada_total = float(entradas_an.get("Compra", pd.Series(dtype=float)).sum()) if not entradas_an.empty else 0.0
    pagamento_total = float(contas_filtradas.get("Valor", pd.Series(dtype=float)).sum()) if not contas_filtradas.empty else 0.0
    cmv_total = (custo_total / venda_total * 100) if venda_total else 0.0
    lucro_total = venda_total - custo_total
    margem_total = (lucro_total / venda_total * 100) if venda_total else 0.0

    cards_an = st.columns(6)
    valores_cards = [
        ("Venda Geral", moeda_real(venda_total)),
        ("Custo Médio", moeda_real(custo_total)),
        ("Entrada Geral", moeda_real(entrada_total)),
        ("CMV Geral", percentual(cmv_total)),
        ("Lucro Bruto", moeda_real(lucro_total)),
        ("Pag. Fornecedor", moeda_real(pagamento_total)),
    ]
    for col, (rotulo, valor) in zip(cards_an, valores_cards):
        col.metric(rotulo, valor)

    areas = sorted(set(vendas_an.get("Área", pd.Series(dtype=str)).dropna().tolist()) | set(entradas_an.get("Área", pd.Series(dtype=str)).dropna().tolist()))
    meses = list(meses_selecionados)
    registros = []
    for area in areas:
        for mes in meses:
            vv = vendas_an[(vendas_an.get("Área") == area) & (vendas_an.get("Competência") == mes)] if not vendas_an.empty else pd.DataFrame()
            ee = entradas_an[(entradas_an.get("Área") == area) & (entradas_an.get("Competência") == mes)] if not entradas_an.empty else pd.DataFrame()
            venda = float(vv.get("Venda", pd.Series(dtype=float)).sum()) if not vv.empty else 0.0
            custo = float(vv.get("Custo", pd.Series(dtype=float)).sum()) if not vv.empty else 0.0
            compra = float(ee.get("Compra", pd.Series(dtype=float)).sum()) if not ee.empty else 0.0
            registros.append({
                "Área": area,
                "Competência": mes,
                "Venda": venda,
                "Custo": custo,
                "Compra": compra,
                "CMV (%)": (custo / venda * 100) if venda else 0,
                "Entrada - CMV": compra - custo,
                "Lucro Bruto": venda - custo,
                "Margem (%)": ((venda - custo) / venda * 100) if venda else 0,
            })
    analise_df = pd.DataFrame(registros)

    st.markdown("### Análise agrupada por área")
    indicador = st.selectbox("Indicador da matriz", ["Venda", "Custo", "Compra", "CMV (%)", "Entrada - CMV", "Lucro Bruto", "Margem (%)"])
    if not analise_df.empty:
        matriz = analise_df.pivot_table(
            index="Área",
            columns="Competência",
            values=indicador,
            aggfunc="sum",
            fill_value=0,
        )
        matriz["Total Geral"] = (
            matriz.sum(axis=1)
            if indicador not in ["CMV (%)", "Margem (%)"]
            else analise_df.groupby("Área")[indicador].mean()
        )
        matriz_exibicao = matriz.copy().astype(object)
        for coluna_matriz in matriz_exibicao.columns:
            if indicador in ["CMV (%)", "Margem (%)"]:
                matriz_exibicao[coluna_matriz] = matriz[coluna_matriz].map(percentual)
            else:
                matriz_exibicao[coluna_matriz] = matriz[coluna_matriz].map(moeda_real)
        dataframe_br(matriz_exibicao, use_container_width=True, height=390)
    else:
        st.info("Ainda não existem dados suficientes no cache para montar a matriz.")

    ano_numero_status = int(ano_selecionado)
    ultimo_mes_status = (
        12 if ano_numero_status < date.today().year
        else date.today().month if ano_numero_status == date.today().year
        else 0
    )
    competencias_realizadas_status = [
        f"{ano_numero_status:04d}-{mes_numero:02d}"
        for mes_numero in range(1, ultimo_mes_status + 1)
    ]
    competencias_com_dados_status = set(
        vendas_an.get("Competência", pd.Series(dtype=str))
        .dropna().astype(str).tolist()
    ) | set(
        entradas_an.get("Competência", pd.Series(dtype=str))
        .dropna().astype(str).tolist()
    )
    faltantes_status = [
        competencia
        for competencia in competencias_realizadas_status
        if competencia not in competencias_com_dados_status
    ]
    if faltantes_status:
        st.warning(
            "Competências realizadas ainda não atualizadas: "
            + ", ".join(faltantes_status)
            + ". Use o botão Atualizar análise."
        )

    ac_btn, ac_msg = st.columns([1.1, 3])
    with ac_btn:
        if st.button(
            "🔄 Atualizar análise",
            key=f"atualizar_evolucao_{ano_selecionado}",
            use_container_width=True,
            help="Atualiza cada competência diretamente no PostgreSQL.",
        ):
            cfg_atualizacao = carregar_config_banco()
            ano_numero = int(ano_selecionado)
            ultimo_mes = (
                12 if ano_numero < date.today().year
                else date.today().month if ano_numero == date.today().year
                else 0
            )

            if ultimo_mes == 0:
                st.info("O ano selecionado ainda não possui competências realizadas.")
            else:
                competencias_existentes = set(
                    vendas_an.get("Competência", pd.Series(dtype=str))
                    .dropna().astype(str).tolist()
                ) | set(
                    entradas_an.get("Competência", pd.Series(dtype=str))
                    .dropna().astype(str).tolist()
                )

                competencias_realizadas = [
                    f"{ano_numero:04d}-{numero_mes:02d}"
                    for numero_mes in range(1, ultimo_mes + 1)
                ]
                competencias_faltantes = [
                    competencia
                    for competencia in competencias_realizadas
                    if competencia not in competencias_existentes
                ]

                # Atualização manual prioriza somente competências faltantes.
                # Caso todas existam, permite atualizar novamente o ano realizado.
                competencias_atualizar = (
                    competencias_faltantes
                    if competencias_faltantes
                    else competencias_realizadas
                )

                total_etapas = len(competencias_atualizar) * 3 + 1
                etapa = 0
                mensagens_ano = []
                progresso_ano = st.progress(
                    0,
                    text=(
                        "Atualizando competências faltantes..."
                        if competencias_faltantes
                        else "Atualizando competências realizadas..."
                    ),
                )

                for competencia_atualizacao in competencias_atualizar:
                    numero_mes = int(competencia_atualizacao[-2:])
                    ultimo_dia = calendar.monthrange(ano_numero, numero_mes)[1]
                    data_inicio_atualizacao = f"{competencia_atualizacao}-01"
                    data_fim_atualizacao = f"{competencia_atualizacao}-{ultimo_dia:02d} 23:59:59"

                    for fonte_atualizacao in ["vendas", "entradas", "contas_pagar"]:
                        etapa += 1
                        titulo = FONTES_BANCO[fonte_atualizacao]["titulo"]
                        progresso_ano.progress(
                            etapa / total_etapas,
                            text=f"{competencia_atualizacao} • {titulo}",
                        )
                        try:
                            qtd = executar_atualizacao_fonte(
                                fonte_atualizacao,
                                cfg_atualizacao,
                                competencia_atualizacao,
                                data_inicio_atualizacao,
                                data_fim_atualizacao,
                            )
                            mensagens_ano.append(
                                f"{competencia_atualizacao} • {titulo}: {qtd:,} registros"
                            )
                        except Exception as erro:
                            mensagens_ano.append(
                                f"{competencia_atualizacao} • {titulo}: ERRO — {erro}"
                            )

                competencia_estoque = f"{ano_numero:04d}-{ultimo_mes:02d}"
                ultimo_dia = calendar.monthrange(ano_numero, ultimo_mes)[1]
                etapa += 1
                progresso_ano.progress(
                    etapa / total_etapas,
                    text=f"{competencia_estoque} • Estoque",
                )
                try:
                    executar_atualizacao_fonte(
                        "estoque",
                        cfg_atualizacao,
                        competencia_estoque,
                        f"{competencia_estoque}-01",
                        f"{competencia_estoque}-{ultimo_dia:02d} 23:59:59",
                    )
                except Exception as erro:
                    mensagens_ano.append(
                        f"{competencia_estoque} • Estoque: ERRO — {erro}"
                    )

                st.cache_data.clear()
                progresso_ano.progress(1.0, text="Atualização concluída.")
                st.success("Evolução Comercial atualizada mês a mês.")
                with st.expander("Resultado da atualização"):
                    st.code("\n".join(mensagens_ano[-60:]), language=None)
                st.rerun()
    with ac_msg:
        st.caption(
            "A atualização é executada somente quando solicitada. "
            "Na navegação normal, a tela continua lendo os resumos rápidos."
        )

    st.markdown("### Evolução Comercial — mês a mês")
    st.caption(
        "Mesma dinâmica do relatório: indicadores nas linhas, competências nas colunas "
        "e Total Geral ao final."
    )

    resumo_mensal = []
    for mes in meses:
        vm = vendas_an[vendas_an.get("Competência") == mes] if not vendas_an.empty else pd.DataFrame()
        em = entradas_an[entradas_an.get("Competência") == mes] if not entradas_an.empty else pd.DataFrame()
        pm = contas_filtradas[contas_filtradas.get("Competência") == mes] if not contas_filtradas.empty else pd.DataFrame()
        psm = posicao_an[posicao_an.get("Competência") == mes] if not posicao_an.empty else pd.DataFrame()
        stm = estoque_an[estoque_an.get("Competência") == mes] if not estoque_an.empty else pd.DataFrame()

        venda = float(vm.get("Venda", pd.Series(dtype=float)).sum()) if not vm.empty else 0.0
        custo = float(vm.get("Custo", pd.Series(dtype=float)).sum()) if not vm.empty else 0.0
        entrada = float(em.get("Compra", pd.Series(dtype=float)).sum()) if not em.empty else 0.0
        pagamento = float(pm.get("Valor", pd.Series(dtype=float)).sum()) if not pm.empty else 0.0
        contas_total = float(psm.get("Contas a Pagar", pd.Series(dtype=float)).sum()) if not psm.empty else 0.0
        estoque_posicao = (
            float(psm.get("Estoque", pd.Series(dtype=float)).sum())
            if not psm.empty
            else 0.0
        )
        estoque_resumo = (
            float(stm.get("Estoque", pd.Series(dtype=float)).sum())
            if not stm.empty
            else 0.0
        )
        estoque_mes = (
            estoque_posicao
            if abs(estoque_posicao) > 0.000001
            else estoque_resumo
        )

        resumo_mensal.append({
            "Competência": mes,
            "Venda geral": venda,
            "Custo Médio Geral": custo,
            "Entrada geral": entrada,
            "CMV Geral": (custo / venda * 100) if venda else 0.0,
            "Custo Médio Geral - Entrada Geral (competência)": custo - entrada,
            "Lucro Bruto Geral": venda - custo,
            "Margem Contribuição": ((venda - custo) / venda * 100) if venda else 0.0,
            "Pagamento de Fornecedor": pagamento,
            "Custo médio - Pagamento de Fornecedor (caixa)": custo - pagamento,
            "Contas a Pagar Fornecedor Total": contas_total,
            "Estoque Mês": estoque_mes,
            "Estoque - Contas a Pagar": estoque_mes - contas_total,
        })

    if resumo_mensal:
        evolucao_df = pd.DataFrame(resumo_mensal).set_index("Competência").T.fillna(0)

        nomes_meses = {
            f"{int(ano_selecionado):04d}-01": "Jan",
            f"{int(ano_selecionado):04d}-02": "Fev",
            f"{int(ano_selecionado):04d}-03": "Mar",
            f"{int(ano_selecionado):04d}-04": "Abr",
            f"{int(ano_selecionado):04d}-05": "Mai",
            f"{int(ano_selecionado):04d}-06": "Jun",
            f"{int(ano_selecionado):04d}-07": "Jul",
            f"{int(ano_selecionado):04d}-08": "Ago",
            f"{int(ano_selecionado):04d}-09": "Set",
            f"{int(ano_selecionado):04d}-10": "Out",
            f"{int(ano_selecionado):04d}-11": "Nov",
            f"{int(ano_selecionado):04d}-12": "Dez",
        }

        indicadores_percentuais = {"CMV Geral", "Margem Contribuição"}
        colunas_meses = list(evolucao_df.columns)

        # Total Geral conforme a natureza de cada indicador.
        total_geral = {}
        for indicador in evolucao_df.index:
            if indicador == "CMV Geral":
                total_venda = sum(item["Venda geral"] for item in resumo_mensal)
                total_custo = sum(item["Custo Médio Geral"] for item in resumo_mensal)
                total_geral[indicador] = (
                    total_custo / total_venda * 100 if total_venda else 0.0
                )
            elif indicador == "Margem Contribuição":
                total_venda = sum(item["Venda geral"] for item in resumo_mensal)
                total_custo = sum(item["Custo Médio Geral"] for item in resumo_mensal)
                total_geral[indicador] = (
                    (total_venda - total_custo) / total_venda * 100
                    if total_venda else 0.0
                )
            elif indicador in {
                "Contas a Pagar Fornecedor Total",
                "Estoque Mês",
                "Estoque - Contas a Pagar",
            }:
                # Indicadores de posição: usa o último mês que realmente possui
                # posição, sem considerar meses futuros zerados.
                serie_posicao = pd.to_numeric(
                    evolucao_df.loc[indicador, colunas_meses],
                    errors="coerce",
                ).fillna(0)
                meses_com_posicao = serie_posicao[serie_posicao.ne(0)]
                total_geral[indicador] = (
                    float(meses_com_posicao.iloc[-1])
                    if not meses_com_posicao.empty
                    else 0.0
                )
            else:
                total_geral[indicador] = float(
                    pd.to_numeric(evolucao_df.loc[indicador], errors="coerce")
                    .fillna(0).sum()
                )

        evolucao_df["Total Geral"] = pd.Series(total_geral)

        def _formatar_evolucao(valor, indicador):
            numero = float(valor or 0)
            if indicador in indicadores_percentuais:
                return percentual(numero)
            return moeda_real(numero)

        evolucao_exibicao = evolucao_df.copy().astype(object)
        evolucao_exibicao = evolucao_exibicao.rename(columns=nomes_meses)
        for indicador in evolucao_exibicao.index:
            for coluna in evolucao_exibicao.columns:
                coluna_origem = next(
                    (
                        competencia
                        for competencia, nome_mes in nomes_meses.items()
                        if nome_mes == coluna
                    ),
                    coluna,
                )
                evolucao_exibicao.loc[indicador, coluna] = _formatar_evolucao(
                    evolucao_df.loc[indicador, coluna_origem], indicador
                )

        dataframe_br(
            evolucao_exibicao,
            use_container_width=True,
            height=500,
        )

        # A tabela mantém janeiro a dezembro, mas meses totalmente zerados
        # não participam dos gráficos nem das médias visuais.
        grafico_mensal = pd.DataFrame(resumo_mensal)
        grafico_mensal["Mês"] = grafico_mensal["Competência"].map(nomes_meses)

        colunas_movimento = [
            "Venda geral",
            "Custo Médio Geral",
            "Entrada geral",
            "Pagamento de Fornecedor",
            "Contas a Pagar Fornecedor Total",
            "Estoque Mês",
        ]
        mascara_movimento = (
            grafico_mensal[colunas_movimento]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0)
            .abs()
            .sum(axis=1)
            > 0
        )
        grafico_realizado = grafico_mensal.loc[mascara_movimento].copy()

        if not grafico_realizado.empty:
            fig_evol = go.Figure()
            for coluna in [
                "Venda geral",
                "Custo Médio Geral",
                "Entrada geral",
                "Pagamento de Fornecedor",
            ]:
                fig_evol.add_scatter(
                    x=grafico_realizado["Mês"],
                    y=grafico_realizado[coluna],
                    mode="lines+markers",
                    name=coluna,
                    connectgaps=False,
                )
            fig_evol.update_layout(
                height=380,
                title="Evolução dos principais indicadores",
                margin=dict(l=10, r=10, t=50, b=10),
                yaxis_title="R$",
                xaxis_title=f"Meses realizados de {ano_selecionado}",
            )
            plotly_chart_br(fig_evol, use_container_width=True)
        else:
            st.info("Ainda não existem meses realizados para exibir no gráfico.")

        st.markdown("### CMV por área")
        if not analise_df.empty:
            cmv_area = analise_df.groupby("Área", as_index=False).agg(
                Venda=("Venda", "sum"),
                Custo=("Custo", "sum"),
            )
            cmv_area["CMV (%)"] = cmv_area.apply(
                lambda r: r["Custo"] / r["Venda"] * 100 if r["Venda"] else 0,
                axis=1,
            )
            fig_cmv = px.bar(
                cmv_area.sort_values("CMV (%)"),
                x="CMV (%)",
                y="Área",
                orientation="h",
                text_auto=False,
            )
            fig_cmv.update_layout(
                height=400,
                margin=dict(l=10, r=10, t=30, b=10),
            )
            plotly_chart_br(fig_cmv, use_container_width=True)
    else:
        st.info("Ainda não existem competências suficientes para montar a evolução comercial.")

    st.caption("Planos de contas aplicados: " + (" | ".join(planos_selecionados) if planos_selecionados else "Nenhum"))

elif visao == "Realizados":
    section("Realizados", "sec-gray")
    df = REALIZADOS if comprador == "Todos" else REALIZADOS[REALIZADOS["Comprador"] == comprador]
    dataframe_br(preparar_tabela(df), use_container_width=True, hide_index=True, height=270)

elif visao == "Métricas Destaque":
    section("Métricas Destaque", "sec-green")
    df = METAS if comprador == "Todos" else METAS[METAS["Comprador"] == comprador]
    dataframe_br(preparar_tabela(df), use_container_width=True, hide_index=True, height=270)

elif visao == "Resultado Métricas":
    section("Resultado Métricas - Realizado", "sec-blue")
    df = RESULTADO if comprador == "Todos" else RESULTADO[RESULTADO["Comprador"] == comprador]
    dataframe_br(preparar_tabela(df), use_container_width=True, hide_index=True, height=270)

elif visao == "Prêmio Comprador":
    section("Prêmio Comprador x Métrica", "sec-gold")
    df = PREMIO if comprador == "Todos" else PREMIO[PREMIO["Comprador"] == comprador]
    dataframe_br(preparar_tabela(df), use_container_width=True, hide_index=True, height=280)

    st.markdown("### Gerente Comercial")
    gerente = pd.DataFrame([[
        "Gerente Comercial", 14.95, 99.7, 15.00, 100.0, 48.96, 97.9,
        44.10, 58.8, 73.56, 98.1, 44.60, 89.2, 39.00, 78.0,
        57.26, 57.3, 67.07, 95.8
    ]], columns=PREMIO.columns)
    dataframe_br(preparar_tabela(gerente), use_container_width=True, hide_index=True, height=110)

elif visao == "Prêmio por KPI":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff">Valor total de prêmio atingível</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:3px">Composição ponderada pelos nove KPIs</div>
      </div>
      <div class="value">R$ 3.000,00</div>
    </div>
    """, unsafe_allow_html=True)

    dataframe_br(preparar_tabela(PREMIO_KPI), use_container_width=True, hide_index=True, height=390)

    c1,c2,c3 = st.columns(3)
    with c1: st.metric("Prêmio atingido","R$ 2.638,34")
    with c2: st.metric("Atingimento geral","87,9%")
    with c3: st.metric("Saldo não atingido","R$ 361,66")

    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=87.9,
        number={"suffix":"%","font":{"color":"#f3f7fb"}},
        title={"text":"Atingimento total do prêmio","font":{"color":"#f3f7fb"}},
        gauge={
            "axis":{"range":[0,100],"tickcolor":"#8da2b8"},
            "bar":{"color":"#22d3ee"},
            "bgcolor":"#0c1724",
            "bordercolor":"#1f3850",
            "steps":[
                {"range":[0,70],"color":"#4a2024"},
                {"range":[70,90],"color":"#5a491a"},
                {"range":[90,100],"color":"#16452d"},
            ]
        }
    ))
    fig.update_layout(height=330,margin=dict(l=20,r=20,t=60,b=20),paper_bgcolor="rgba(0,0,0,0)")
    plotly_chart_br(fig,use_container_width=True,config={"displayModeBar":False})


elif visao == "Metas de Loja":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:21px">Metas de Loja: Faturamento e Margem Bruta</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Acompanhamento executivo das metas financeiras por filial e gerente.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Margem Bruta mínima: 34%</div>
    </div>
    """, unsafe_allow_html=True)

    df_todas_lojas = dataframe_metas_lojas()
    periodos_lojas = sorted(
        df_todas_lojas["periodo_referencia"].astype(str).dropna().unique().tolist(),
        reverse=True,
    ) if not df_todas_lojas.empty else []
    periodo_ativo_lojas = str(METAS_GESTOR.get("periodo_referencia", datetime.now().strftime("%Y-%m")))
    if periodo_ativo_lojas not in periodos_lojas:
        periodos_lojas = [periodo_ativo_lojas] + periodos_lojas

    periodo_painel_lojas = st.selectbox(
        "Período das metas de loja",
        periodos_lojas or [periodo_ativo_lojas],
        index=(periodos_lojas.index(periodo_ativo_lojas) if periodo_ativo_lojas in periodos_lojas else 0),
        key="periodo_painel_metas_lojas",
    )
    df_lojas = dataframe_metas_lojas(periodo_painel_lojas)
    if df_lojas.empty:
        st.warning("Nenhuma meta de loja cadastrada para o período selecionado.")
    else:
        total_meta = pd.to_numeric(df_lojas["meta_mes"], errors="coerce").fillna(0).sum()
        total_mb = pd.to_numeric(df_lojas["meta_margem_bruta_valor"], errors="coerce").fillna(0).sum()
        total_entrega = pd.to_numeric(df_lojas["representatividade_entrega_valor"], errors="coerce").fillna(0).sum()
        mb_pct_global = (total_mb / total_meta * 100) if total_meta else 0
        entrega_pct_global = (total_entrega / total_meta * 100) if total_meta else 0

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Meta global do mês", moeda_real(total_meta))
        c2.metric("Meta de margem bruta", moeda_real(total_mb))
        c3.metric("Margem bruta global", percentual(mb_pct_global))
        c4.metric("Representatividade da entrega", moeda_real(total_entrega), percentual(entrega_pct_global))

        tabela = df_lojas.rename(columns={
            "regional_loja": "Regional / Loja",
            "gerente": "Gerente",
            "meta_mes": "META Mês (R$)",
            "meta_margem_bruta_valor": "META Margem Bruta (R$)",
            "meta_margem_bruta_pct": "META MB (%)",
            "representatividade_entrega_pct": "Representatividade Entrega (%)",
            "representatividade_entrega_valor": "Representatividade Entrega (R$)",
            "status": "Status",
        })
        tabela = tabela[[
            "Regional / Loja", "Gerente", "META Mês (R$)",
            "META Margem Bruta (R$)", "META MB (%)",
            "Representatividade Entrega (%)", "Representatividade Entrega (R$)", "Status"
        ]]
        dataframe_br(
            tabela.style.format({
                "META Mês (R$)": lambda x: moeda_real(float(x)),
                "META Margem Bruta (R$)": lambda x: moeda_real(float(x)),
                "META MB (%)": lambda x: percentual(float(x)),
                "Representatividade Entrega (%)": lambda x: percentual(float(x)),
                "Representatividade Entrega (R$)": lambda x: moeda_real(float(x)),
            }),
            use_container_width=True, hide_index=True, height=360
        )

        graf = df_lojas.copy()
        graf["Meta do mês"] = pd.to_numeric(graf["meta_mes"], errors="coerce").fillna(0)
        graf["Margem bruta"] = pd.to_numeric(graf["meta_margem_bruta_valor"], errors="coerce").fillna(0)
        fig_loja = go.Figure()
        fig_loja.add_bar(name="Meta do mês", x=graf["regional_loja"], y=graf["Meta do mês"])
        fig_loja.add_bar(name="Meta margem bruta", x=graf["regional_loja"], y=graf["Margem bruta"])
        fig_loja.update_layout(
            barmode="group", height=390, paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)", font_color="#f3f7fb",
            margin=dict(l=20, r=20, t=45, b=20), yaxis_title="R$", xaxis_title=""
        )
        fig_loja.update_yaxes(gridcolor="rgba(255,255,255,.08)")
        plotly_chart_br(fig_loja, use_container_width=True, config={"displayModeBar": False})

elif visao == "Gestão de Metas de Loja":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:21px">Gestão de Metas de Loja</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Cadastre, altere e acompanhe metas por filial, gerente e período.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Dados persistentes</div>
    </div>
    """, unsafe_allow_html=True)

    df_edicao = dataframe_metas_lojas()
    periodos_existentes = sorted(df_edicao["periodo_referencia"].astype(str).unique().tolist(), reverse=True) if not df_edicao.empty else []
    periodo_padrao = METAS_GESTOR.get("periodo_referencia", datetime.now().strftime("%Y-%m"))
    if periodo_padrao not in periodos_existentes:
        periodos_existentes = [periodo_padrao] + periodos_existentes

    f1, f2 = st.columns([1, 2])
    with f1:
        periodo_gestao = st.selectbox("Período para gestão", periodos_existentes or [periodo_padrao])
    with f2:
        usuario_meta_loja = st.text_input("Responsável pela alteração", value="Gestor")

    with st.expander("📅 Alterar o período de uma meta já cadastrada", expanded=False):
        st.caption(
            "Transfira todas as metas do período selecionado para uma nova competência. "
            "O histórico anterior será preservado."
        )
        try:
            ano_atual, mes_atual = [int(x) for x in str(periodo_gestao).split("-")[:2]]
            data_periodo_atual = date(ano_atual, mes_atual, 1)
        except Exception:
            data_periodo_atual = date.today().replace(day=1)

        ap1, ap2 = st.columns([1, 2])
        with ap1:
            nova_competencia_data = st.date_input(
                "Nova competência",
                value=data_periodo_atual,
                key=f"nova_competencia_meta_loja_{periodo_gestao}",
            )
        with ap2:
            manter_periodo_original = st.checkbox(
                "Manter uma cópia no período original",
                value=False,
                help="Marcado: copia as metas. Desmarcado: move as metas para a nova competência.",
            )

        nova_competencia = nova_competencia_data.strftime("%Y-%m")
        alterar_periodo = st.button(
            "📅 Aplicar novo período",
            type="primary",
            use_container_width=True,
            key=f"alterar_periodo_meta_loja_{periodo_gestao}",
        )

        if alterar_periodo:
            if nova_competencia == str(periodo_gestao):
                st.warning("A nova competência deve ser diferente do período atual.")
            else:
                origem = df_edicao[
                    df_edicao["periodo_referencia"].astype(str) == str(periodo_gestao)
                ].copy()
                if origem.empty:
                    st.warning("Não existem metas cadastradas no período selecionado.")
                else:
                    transferidas = origem.copy()
                    transferidas["periodo_referencia"] = nova_competencia

                    if manter_periodo_original:
                        base_final = pd.concat([df_edicao, transferidas], ignore_index=True)
                    else:
                        base_sem_origem = df_edicao[
                            df_edicao["periodo_referencia"].astype(str) != str(periodo_gestao)
                        ].copy()
                        base_final = pd.concat([base_sem_origem, transferidas], ignore_index=True)

                    # Quando já houver a mesma filial no destino, prevalece a meta transferida.
                    base_final["periodo_referencia"] = base_final["periodo_referencia"].astype(str)
                    base_final["regional_loja"] = base_final["regional_loja"].astype(str)
                    base_final = base_final.drop_duplicates(
                        subset=["periodo_referencia", "regional_loja"],
                        keep="last",
                    )
                    salvar_metas_lojas(base_final.to_dict("records"), usuario_meta_loja)
                    acao = "copiadas" if manter_periodo_original else "transferidas"
                    st.success(
                        f"Metas {acao} de {periodo_gestao} para {nova_competencia}. "
                        "O histórico foi registrado."
                    )
                    st.session_state["periodo_painel_metas_lojas"] = nova_competencia
                    st.rerun()

    df_periodo = df_edicao[df_edicao["periodo_referencia"].astype(str) == str(periodo_gestao)].copy()
    if df_periodo.empty:
        df_periodo = pd.DataFrame([{
            "periodo_referencia": periodo_gestao, "regional_loja": "Nova filial", "gerente": "",
            "meta_mes": 0.0, "meta_margem_bruta_valor": 0.0, "meta_margem_bruta_pct": 34.0,
            "representatividade_entrega_pct": 0.0, "representatividade_entrega_valor": 0.0, "status": "Planejada"
        }])

    st.info("Você pode adicionar ou excluir linhas diretamente na tabela. A Margem Bruta em R$ pode ser recalculada automaticamente a partir da meta do mês e do percentual.")
    editado = st.data_editor(
        df_periodo, num_rows="dynamic", use_container_width=True, hide_index=True,
        column_config={
            "periodo_referencia": st.column_config.TextColumn("Período", disabled=True),
            "regional_loja": st.column_config.TextColumn("Regional / Loja", required=True),
            "gerente": st.column_config.TextColumn("Gerente", required=True),
            "meta_mes": st.column_config.NumberColumn("META Mês (R$)", min_value=0.0, format="R$ %.2f"),
            "meta_margem_bruta_valor": st.column_config.NumberColumn("META Margem Bruta (R$)", min_value=0.0, format="R$ %.2f"),
            "meta_margem_bruta_pct": st.column_config.NumberColumn("META MB (%)", min_value=0.0, max_value=100.0, format="%.2f%%"),
            "representatividade_entrega_pct": st.column_config.NumberColumn("Representatividade Entrega (%)", min_value=0.0, max_value=100.0, format="%.2f%%"),
            "representatividade_entrega_valor": st.column_config.NumberColumn("Representatividade Entrega (R$)", min_value=0.0, format="R$ %.2f"),
            "status": st.column_config.SelectboxColumn("Status", options=["Planejada", "Ativa", "Encerrada", "Cancelada"]),
        },
        key=f"editor_metas_lojas_{periodo_gestao}"
    )

    b1, b2, b3 = st.columns(3)
    recalcular = b1.button("🧮 Recalcular Margem em R$", use_container_width=True)
    salvar = b2.button("💾 Salvar metas de loja", type="primary", use_container_width=True)
    restaurar = b3.button("↩️ Restaurar exemplo inicial", use_container_width=True)

    if recalcular:
        temp = editado.copy()
        temp["periodo_referencia"] = periodo_gestao
        temp["meta_margem_bruta_valor"] = (
            pd.to_numeric(temp["meta_mes"], errors="coerce").fillna(0) *
            pd.to_numeric(temp["meta_margem_bruta_pct"], errors="coerce").fillna(0) / 100
        ).round(2)
        for c in ["meta_mes", "meta_margem_bruta_valor", "meta_margem_bruta_pct", "representatividade_entrega_pct", "representatividade_entrega_valor"]:
            temp[c] = pd.to_numeric(temp[c], errors="coerce").fillna(0.0)
        temp = temp[temp["regional_loja"].astype(str).str.strip() != ""].copy()
        restantes = df_edicao[df_edicao["periodo_referencia"].astype(str) != str(periodo_gestao)].copy()
        salvar_metas_lojas(pd.concat([restantes, temp], ignore_index=True).to_dict("records"), usuario_meta_loja)
        st.success("Margem bruta recalculada, salva e registrada no histórico.")
        st.rerun()

    if salvar:
        temp = editado.copy()
        temp["periodo_referencia"] = periodo_gestao
        for c in ["meta_mes", "meta_margem_bruta_valor", "meta_margem_bruta_pct", "representatividade_entrega_pct", "representatividade_entrega_valor"]:
            temp[c] = pd.to_numeric(temp[c], errors="coerce").fillna(0.0)
        temp = temp[temp["regional_loja"].astype(str).str.strip() != ""].copy()
        restantes = df_edicao[df_edicao["periodo_referencia"].astype(str) != str(periodo_gestao)].copy()
        consolidado = pd.concat([restantes, temp], ignore_index=True)
        salvar_metas_lojas(consolidado.to_dict("records"), usuario_meta_loja)
        st.success("Metas de loja salvas e histórico registrado.")
        st.rerun()

    if restaurar:
        outros = df_edicao[df_edicao["periodo_referencia"].astype(str) != str(periodo_gestao)].copy()
        padrao = pd.DataFrame([dict(x, periodo_referencia=periodo_gestao) for x in METAS_LOJAS_PADRAO])
        salvar_metas_lojas(pd.concat([outros, padrao], ignore_index=True).to_dict("records"), usuario_meta_loja)
        st.success("Exemplo inicial restaurado para o período selecionado.")
        st.rerun()

    st.markdown("### Resumo do período")
    resumo = editado.copy()
    total_meta = pd.to_numeric(resumo["meta_mes"], errors="coerce").fillna(0).sum()
    total_mb = pd.to_numeric(resumo["meta_margem_bruta_valor"], errors="coerce").fillna(0).sum()
    total_entrega = pd.to_numeric(resumo["representatividade_entrega_valor"], errors="coerce").fillna(0).sum()
    r1, r2, r3 = st.columns(3)
    r1.metric("Meta global", moeda_real(total_meta))
    r2.metric("Margem bruta global", moeda_real(total_mb), percentual(total_mb / total_meta * 100 if total_meta else 0))
    r3.metric("Entrega representativa", moeda_real(total_entrega), percentual(total_entrega / total_meta * 100 if total_meta else 0))

    with st.expander("📚 Histórico de alterações", expanded=False):
        hist_lojas = carregar_historico_metas_lojas()
        if hist_lojas:
            hist_resumo = pd.DataFrame([{
                "Data/Hora": h.get("data_hora", ""),
                "Usuário": h.get("usuario", ""),
                "Quantidade de lojas": len(h.get("registros", [])),
            } for h in reversed(hist_lojas)])
            dataframe_br(hist_resumo, use_container_width=True, hide_index=True)
        else:
            st.caption("Ainda não há alterações registradas.")

elif visao == "Gestão de Metas":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Centro de Gestão de Metas</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Cadastre metas por período, preserve histórico e controle a vigência.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Rede Economize Control Center</div>
    </div>
    """, unsafe_allow_html=True)

    periodos_unificados = listar_periodos_gestao_metas()
    periodo_atual_gestao = str(METAS_GESTOR.get("periodo_referencia", datetime.now().strftime("%Y-%m")))
    periodo_estado = str(st.session_state.get("periodo_gestao_unificado_global", "")).strip()
    for _periodo_extra in [periodo_atual_gestao, periodo_estado]:
        if _periodo_extra and _periodo_extra not in periodos_unificados:
            periodos_unificados.insert(0, _periodo_extra)

    def _abrir_competencia_global_callback():
        nova_data = st.session_state.get("nova_competencia_gestao_global")
        if nova_data:
            st.session_state["periodo_gestao_unificado_global"] = nova_data.strftime("%Y-%m")

    pg1, pg2, pg3 = st.columns([1.3, 1, 1.7])
    with pg1:
        periodo_gestao_unificado = st.selectbox("Período de trabalho para todas as metas", periodos_unificados or [periodo_atual_gestao], key="periodo_gestao_unificado_global")
    with pg2:
        nova_competencia_global = st.date_input("Nova competência", value=datetime.strptime(periodo_gestao_unificado + "-01", "%Y-%m-%d").date(), key="nova_competencia_gestao_global")
    with pg3:
        st.caption("Este período controla Meta Geral, Metas de Loja e Metas por Comprador.")
        st.button("➕ Abrir competência selecionada", key="abrir_competencia_global", use_container_width=True, on_click=_abrir_competencia_global_callback)

    aba1, aba_lojas, aba_compradores, aba2 = st.tabs([
        "📝 Meta Geral",
        "🏪 Metas de Loja",
        "🎯 Metas por Comprador",
        "📚 Histórico de Metas",
    ])

    with aba1:
        with st.form("form_metas_gestor"):
            st.markdown("### Identificação e período")
            p1, p2, p3, p4 = st.columns([1.2, 1, 1, 1])
            with p1:
                periodo_referencia = st.text_input(
                    "Período de referência",
                    value=periodo_gestao_unificado,
                    placeholder="Ex.: 2026-05",
                    key="periodo_referencia_meta_geral_unificado"
                )
            with p2:
                data_inicio = st.date_input(
                    "Data inicial",
                    value=datetime.strptime(
                        METAS_GESTOR.get("data_inicio", "2026-07-01"), "%Y-%m-%d"
                    ).date()
                )
            with p3:
                data_fim = st.date_input(
                    "Data final",
                    value=datetime.strptime(
                        METAS_GESTOR.get("data_fim", "2026-07-31"), "%Y-%m-%d"
                    ).date()
                )
            with p4:
                status_meta = st.selectbox(
                    "Status",
                    ["Planejada", "Ativa", "Encerrada", "Cancelada"],
                    index=["Planejada", "Ativa", "Encerrada", "Cancelada"].index(
                        METAS_GESTOR.get("status", "Ativa")
                    )
                )

            descricao = st.text_input(
                "Descrição da meta",
                value=METAS_GESTOR.get("descricao", "Metas comerciais do período")
            )
            usuario_cadastro = st.text_input(
                "Gestor responsável",
                value=METAS_GESTOR.get("usuario_cadastro", "Gestor")
            )

            st.markdown("### Metas gerais")
            g1, g2, g3, g4 = st.columns(4)
            with g1:
                meta_venda = st.number_input("Meta Venda Total Mês (R$)", min_value=0.0, value=float(METAS_GESTOR["meta_venda_total_mes"]), step=10000.0, format="%.2f")
            with g2:
                meta_cmv = st.number_input("Meta CMV Mês (%)", min_value=0.0, max_value=100.0, value=float(METAS_GESTOR["meta_cmv_mes"]), step=0.1, format="%.2f")
            with g3:
                fator_reducao = st.number_input("Fator Redução CMV", min_value=0.0, value=float(METAS_GESTOR["fator_reducao_cmv"]), step=0.01, format="%.2f")
            with g4:
                fator_cobertura = st.number_input("Fator Cobertura", min_value=0.0, value=float(METAS_GESTOR["fator_cobertura"]), step=0.05, format="%.2f")

            st.markdown("### Metas operacionais")
            o1, o2 = st.columns(2)
            with o1:
                meta_ruptura = st.number_input("Meta Ruptura (%)", min_value=0.0, max_value=100.0, value=float(METAS_GESTOR["meta_ruptura"]), step=0.1, format="%.2f")
            with o2:
                meta_reposicao = st.number_input("Meta Reposição (%)", min_value=0.0, max_value=200.0, value=float(METAS_GESTOR["meta_reposicao"]), step=0.1, format="%.2f")

            st.markdown("### Distribuição por curva de estoque")
            c1, c2, c3, c4 = st.columns(4)
            with c1: curva_a = st.number_input("Curva A (%)", 0.0, 100.0, float(METAS_GESTOR["curva_a"]), 1.0)
            with c2: curva_b = st.number_input("Curva B (%)", 0.0, 100.0, float(METAS_GESTOR["curva_b"]), 1.0)
            with c3: curva_c = st.number_input("Curva C (%)", 0.0, 100.0, float(METAS_GESTOR["curva_c"]), 1.0)
            with c4: curva_d = st.number_input("Curva D (%)", 0.0, 100.0, float(METAS_GESTOR["curva_d"]), 1.0)

            total_curvas = curva_a + curva_b + curva_c + curva_d
            st.caption(f"Total das curvas: {percentual(total_curvas)}")

            st.markdown("### Participação de venda por comprador")
            st.caption(
                "Os compradores são reconhecidos automaticamente pelas bases e pelo mapa "
                "de classificações. Os percentuais abaixo podem ser alterados."
            )
            compradores_reconhecidos = sorted(lista_compradores_ativos(), key=lambda x: x.casefold())
            metas_participacao = carregar_metas_por_comprador()
            linhas_participacao = []
            for nome in compradores_reconhecidos:
                atual = next((
                    item for item in metas_participacao
                    if str(item.get("periodo_referencia", "")) == str(periodo_referencia)
                    and str(item.get("comprador", "")).strip().casefold() == nome.casefold()
                ), {})
                realizado_nome = REALIZADOS.loc[
                    REALIZADOS["Comprador"].astype(str).str.strip().str.casefold() == nome.casefold(),
                    "Rep. Faturamento",
                ]
                participacao_real = float(realizado_nome.iloc[0]) if not realizado_nome.empty else 0.0
                linhas_participacao.append({
                    "Comprador": nome,
                    "Participação Meta (%)": float(atual.get("participacao_venda_pct", participacao_real)),
                    "Participação Real (%)": participacao_real,
                })
            editor_participacao = st.data_editor(
                pd.DataFrame(linhas_participacao),
                use_container_width=True,
                hide_index=True,
                disabled=["Comprador", "Participação Real (%)"],
                column_config={
                    "Comprador": st.column_config.TextColumn("Comprador reconhecido"),
                    "Participação Meta (%)": st.column_config.NumberColumn(
                        "Participação Meta (%)", min_value=0.0, max_value=100.0,
                        step=0.1, format="%.2f%%"
                    ),
                    "Participação Real (%)": st.column_config.NumberColumn(
                        "Participação Real (%)", format="%.2f%%"
                    ),
                },
                key="editor_participacao_compradores_dinamicos",
            )
            total_rep = pd.to_numeric(
                editor_participacao.get("Participação Meta (%)", pd.Series(dtype=float)),
                errors="coerce",
            ).fillna(0).sum()
            st.caption(f"Total da participação configurada: {percentual(total_rep)}")

            st.markdown("### Pesos da premiação")
            p1, p2, p3 = st.columns(3)
            with p1:
                peso_fat = st.number_input("Peso Faturamento (%)", 0.0, 100.0, float(METAS_GESTOR["peso_faturamento"]), 1.0)
                peso_cmv = st.number_input("Peso CMV (%)", 0.0, 100.0, float(METAS_GESTOR["peso_cmv"]), 1.0)
                peso_cob = st.number_input("Peso Fator Cobertura (%)", 0.0, 100.0, float(METAS_GESTOR["peso_fator_cobertura"]), 1.0)
            with p2:
                peso_a = st.number_input("Peso Curva A (%)", 0.0, 100.0, float(METAS_GESTOR["peso_curva_a"]), 1.0)
                peso_b = st.number_input("Peso Curva B (%)", 0.0, 100.0, float(METAS_GESTOR["peso_curva_b"]), 1.0)
                peso_c = st.number_input("Peso Curva C (%)", 0.0, 100.0, float(METAS_GESTOR["peso_curva_c"]), 1.0)
            with p3:
                peso_d = st.number_input("Peso Curva D (%)", 0.0, 100.0, float(METAS_GESTOR["peso_curva_d"]), 1.0)
                peso_ruptura = st.number_input("Peso Ruptura (%)", 0.0, 100.0, float(METAS_GESTOR["peso_ruptura"]), 1.0)
                peso_reposicao = st.number_input("Peso Reposição CMV (%)", 0.0, 100.0, float(METAS_GESTOR["peso_reposicao"]), 1.0)

            total_pesos = peso_fat + peso_cmv + peso_cob + peso_a + peso_b + peso_c + peso_d + peso_ruptura + peso_reposicao
            st.caption(f"Peso total configurado: {percentual(total_pesos)}")

            valor_premio = st.number_input("Valor total de prêmio atingível (R$)", min_value=0.0, value=float(METAS_GESTOR["valor_premio_total"]), step=100.0, format="%.2f")

            salvar = st.form_submit_button("💾 Salvar meta do período", use_container_width=True)

            if salvar:
                erros = []
                if not periodo_referencia.strip():
                    erros.append("Informe o período de referência.")
                if data_fim < data_inicio:
                    erros.append("A data final não pode ser anterior à data inicial.")
                if abs(total_curvas - 100.0) > 0.01:
                    erros.append("A soma das curvas deve ser igual a 100%.")
                if abs(total_rep - 100.0) > 0.01:
                    erros.append("A soma da participação dos compradores deve ser igual a 100%.")
                if abs(total_pesos - 100.0) > 0.01:
                    erros.append("A soma dos pesos da premiação deve ser igual a 100%.")

                if erros:
                    for erro in erros:
                        st.error(erro)
                else:
                    id_meta = f"META-{periodo_referencia.strip()}"
                    novas_metas = {
                        "id_meta": id_meta,
                        "periodo_referencia": periodo_referencia.strip(),
                        "data_inicio": data_inicio.strftime("%Y-%m-%d"),
                        "data_fim": data_fim.strftime("%Y-%m-%d"),
                        "descricao": descricao,
                        "status": status_meta,
                        "usuario_cadastro": usuario_cadastro,
                        "data_cadastro": METAS_GESTOR.get("data_cadastro", ""),
                        "meta_venda_total_mes": meta_venda,
                        "meta_cmv_mes": meta_cmv,
                        "fator_reducao_cmv": fator_reducao,
                        "fator_cobertura": fator_cobertura,
                        "meta_ruptura": meta_ruptura,
                        "meta_reposicao": meta_reposicao,
                        "curva_a": curva_a,
                        "curva_b": curva_b,
                        "curva_c": curva_c,
                        "curva_d": curva_d,
                        # Campos legados preservados apenas por compatibilidade.
                        # A participação atual é armazenada por comprador logo abaixo.
                        "rep_paulo": float(METAS_GESTOR.get("rep_paulo", 0)),
                        "rep_francieli": float(METAS_GESTOR.get("rep_francieli", 0)),
                        "rep_sebastiao": float(METAS_GESTOR.get("rep_sebastiao", 0)),
                        "peso_faturamento": peso_fat,
                        "peso_cmv": peso_cmv,
                        "peso_fator_cobertura": peso_cob,
                        "peso_curva_a": peso_a,
                        "peso_curva_b": peso_b,
                        "peso_curva_c": peso_c,
                        "peso_curva_d": peso_d,
                        "peso_ruptura": peso_ruptura,
                        "peso_reposicao": peso_reposicao,
                        "valor_premio_total": valor_premio,
                    }
                    # Salva a participação para cada comprador reconhecido.
                    metas_individuais = carregar_metas_por_comprador()
                    for _, linha_part in editor_participacao.iterrows():
                        nome_part = str(linha_part.get("Comprador", "")).strip()
                        pct_part = float(linha_part.get("Participação Meta (%)", 0) or 0)
                        encontrado = False
                        for item in metas_individuais:
                            if (
                                str(item.get("periodo_referencia", "")) == str(periodo_referencia.strip())
                                and str(item.get("comprador", "")).strip().casefold() == nome_part.casefold()
                            ):
                                item["participacao_venda_pct"] = pct_part
                                item["meta_venda"] = float(meta_venda) * pct_part / 100.0
                                item["ultima_atualizacao"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                encontrado = True
                                break
                        if not encontrado and nome_part:
                            novo_item = estrutura_meta_comprador_padrao(
                                nome_part, periodo_referencia.strip(), pct_part
                            )
                            novo_item["meta_venda"] = float(meta_venda) * pct_part / 100.0
                            novo_item["ultima_atualizacao"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            metas_individuais.append(novo_item)
                    salvar_metas_por_comprador(metas_individuais)
                    salvar_metas(novas_metas)
                    st.success("Meta do período salva com sucesso.")
                    st.rerun()

        if METAS_GESTOR.get("ultima_atualizacao"):
            st.caption(f"Última atualização: {METAS_GESTOR['ultima_atualizacao']}")

    with aba_lojas:
        st.markdown("### Gestão de Metas de Loja")
        st.caption(
            "Cadastre e altere as metas de faturamento e margem bruta por filial, "
            "usando o mesmo padrão de período, vigência, status e histórico da meta geral."
        )

        df_lojas_todas = dataframe_metas_lojas()
        periodos_lojas = sorted(
            df_lojas_todas["periodo_referencia"].astype(str).dropna().unique().tolist(),
            reverse=True,
        ) if not df_lojas_todas.empty else []
        periodo_padrao_lojas = str(METAS_GESTOR.get("periodo_referencia", datetime.now().strftime("%Y-%m")))
        if periodo_padrao_lojas not in periodos_lojas:
            periodos_lojas = [periodo_padrao_lojas] + periodos_lojas

        c_periodo_loja, c_inicio_loja, c_fim_loja, c_usuario_loja = st.columns([1, 1, 1, 1.3])
        with c_periodo_loja:
            periodo_gestao_lojas = periodo_gestao_unificado
            st.text_input("Período de referência", value=periodo_gestao_lojas, disabled=True, key="periodo_lojas_unificado_exibicao")
        try:
            ano_l, mes_l = [int(x) for x in str(periodo_gestao_lojas).split("-")[:2]]
            inicio_padrao_loja = date(ano_l, mes_l, 1)
            if mes_l == 12:
                fim_padrao_loja = date(ano_l, 12, 31)
            else:
                fim_padrao_loja = date(ano_l, mes_l + 1, 1) - timedelta(days=1)
        except Exception:
            inicio_padrao_loja = date.today().replace(day=1)
            fim_padrao_loja = date.today()
        with c_inicio_loja:
            data_inicio_lojas = st.date_input(
                "Data inicial",
                value=inicio_padrao_loja,
                key="gestao_unificada_inicio_lojas",
            )
        with c_fim_loja:
            data_fim_lojas = st.date_input(
                "Data final",
                value=fim_padrao_loja,
                key="gestao_unificada_fim_lojas",
            )
        with c_usuario_loja:
            usuario_meta_loja_unificado = st.text_input(
                "Gestor responsável",
                value="Gestor",
                key="gestao_unificada_usuario_lojas",
            )

        df_periodo_lojas = df_lojas_todas[
            df_lojas_todas["periodo_referencia"].astype(str) == str(periodo_gestao_lojas)
        ].copy() if not df_lojas_todas.empty else pd.DataFrame()

        colunas_lojas = [
            "periodo_referencia", "regional_loja", "gerente", "meta_mes",
            "meta_margem_bruta_valor", "meta_margem_bruta_pct",
            "representatividade_entrega_pct", "representatividade_entrega_valor", "status"
        ]
        if df_periodo_lojas.empty:
            df_periodo_lojas = pd.DataFrame(columns=colunas_lojas)
        for col in colunas_lojas:
            if col not in df_periodo_lojas.columns:
                df_periodo_lojas[col] = "" if col in {"periodo_referencia", "regional_loja", "gerente", "status"} else 0.0
        df_periodo_lojas["periodo_referencia"] = str(periodo_gestao_lojas)

        editor_lojas_unificado = st.data_editor(
            df_periodo_lojas[colunas_lojas],
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "periodo_referencia": st.column_config.TextColumn("Período", disabled=True),
                "regional_loja": st.column_config.TextColumn("Regional / Loja", required=True),
                "gerente": st.column_config.TextColumn("Gerente", required=True),
                "meta_mes": st.column_config.NumberColumn("Meta Faturamento (R$)", min_value=0.0, format="R$ %.2f"),
                "meta_margem_bruta_valor": st.column_config.NumberColumn("Meta Margem Bruta (R$)", min_value=0.0, format="R$ %.2f"),
                "meta_margem_bruta_pct": st.column_config.NumberColumn("Meta MB (%)", min_value=0.0, max_value=100.0, format="%.2f%%"),
                "representatividade_entrega_pct": st.column_config.NumberColumn("Representatividade (%)", min_value=0.0, max_value=100.0, format="%.2f%%"),
                "representatividade_entrega_valor": st.column_config.NumberColumn("Representatividade (R$)", min_value=0.0, format="R$ %.2f"),
                "status": st.column_config.SelectboxColumn("Status", options=["Planejada", "Ativa", "Encerrada", "Cancelada"]),
            },
            key=f"gestao_unificada_editor_lojas_{periodo_gestao_lojas}",
        )

        bl1, bl2, bl3 = st.columns(3)
        recalcular_lojas = bl1.button("🧮 Recalcular margem", use_container_width=True, key="gestao_unificada_recalcular_lojas")
        salvar_lojas = bl2.button("💾 Salvar metas de loja", type="primary", use_container_width=True, key="gestao_unificada_salvar_lojas")
        duplicar_lojas = bl3.button("📑 Duplicar para outro período", use_container_width=True, key="gestao_unificada_duplicar_lojas")

        if recalcular_lojas or salvar_lojas:
            if data_fim_lojas < data_inicio_lojas:
                st.error("A data final não pode ser anterior à data inicial.")
            else:
                temp_lojas = editor_lojas_unificado.copy()
                temp_lojas["periodo_referencia"] = str(periodo_gestao_lojas)
                for c in ["meta_mes", "meta_margem_bruta_valor", "meta_margem_bruta_pct", "representatividade_entrega_pct", "representatividade_entrega_valor"]:
                    temp_lojas[c] = pd.to_numeric(temp_lojas[c], errors="coerce").fillna(0.0)
                if recalcular_lojas:
                    temp_lojas["meta_margem_bruta_valor"] = (
                        temp_lojas["meta_mes"] * temp_lojas["meta_margem_bruta_pct"] / 100
                    ).round(2)
                temp_lojas = temp_lojas[temp_lojas["regional_loja"].astype(str).str.strip() != ""].copy()
                restantes_lojas = df_lojas_todas[
                    df_lojas_todas["periodo_referencia"].astype(str) != str(periodo_gestao_lojas)
                ].copy() if not df_lojas_todas.empty else pd.DataFrame()
                consolidado_lojas = pd.concat([restantes_lojas, temp_lojas], ignore_index=True)
                salvar_metas_lojas(consolidado_lojas.to_dict("records"), usuario_meta_loja_unificado)
                st.success("Metas de loja salvas e registradas no histórico.")
                st.rerun()

        if duplicar_lojas:
            st.session_state["mostrar_duplicacao_lojas_unificada"] = True
        if st.session_state.get("mostrar_duplicacao_lojas_unificada"):
            d1, d2 = st.columns([1, 2])
            with d1:
                nova_comp_lojas = st.date_input(
                    "Nova competência",
                    value=inicio_padrao_loja,
                    key="gestao_unificada_nova_comp_lojas",
                ).strftime("%Y-%m")
            with d2:
                st.caption("A cópia preserva o período original e cria uma nova versão editável.")
            if st.button("Confirmar duplicação", type="primary", key="gestao_unificada_confirmar_dup_lojas"):
                origem_lojas = df_lojas_todas[
                    df_lojas_todas["periodo_referencia"].astype(str) == str(periodo_gestao_lojas)
                ].copy()
                if origem_lojas.empty:
                    st.warning("Não existem metas no período selecionado.")
                elif nova_comp_lojas == str(periodo_gestao_lojas):
                    st.warning("Escolha uma competência diferente.")
                else:
                    copia_lojas = origem_lojas.copy()
                    copia_lojas["periodo_referencia"] = nova_comp_lojas
                    base_lojas = pd.concat([df_lojas_todas, copia_lojas], ignore_index=True)
                    base_lojas = base_lojas.drop_duplicates(
                        subset=["periodo_referencia", "regional_loja"], keep="last"
                    )
                    salvar_metas_lojas(base_lojas.to_dict("records"), usuario_meta_loja_unificado)
                    st.session_state["mostrar_duplicacao_lojas_unificada"] = False
                    st.success(f"Metas duplicadas para {nova_comp_lojas}.")
                    st.rerun()

        if not editor_lojas_unificado.empty:
            total_meta_loja = pd.to_numeric(editor_lojas_unificado["meta_mes"], errors="coerce").fillna(0).sum()
            total_mb_loja = pd.to_numeric(editor_lojas_unificado["meta_margem_bruta_valor"], errors="coerce").fillna(0).sum()
            total_rep_loja = pd.to_numeric(editor_lojas_unificado["representatividade_entrega_valor"], errors="coerce").fillna(0).sum()
            rl1, rl2, rl3 = st.columns(3)
            rl1.metric("Meta global das lojas", moeda_real(total_meta_loja))
            rl2.metric("Margem bruta global", moeda_real(total_mb_loja), percentual(total_mb_loja / total_meta_loja * 100 if total_meta_loja else 0))
            rl3.metric("Representatividade", moeda_real(total_rep_loja), percentual(total_rep_loja / total_meta_loja * 100 if total_meta_loja else 0))

    with aba_compradores:
        st.markdown("### Gestão de Metas por Comprador")
        st.caption(
            "Cadastre metas individuais por comprador e período, mantendo o mesmo padrão "
            "de status, vigência, validação e histórico da meta geral."
        )

        historico_metas_unificado = carregar_historico()
        periodos_compradores = list(dict.fromkeys(
            [METAS_GESTOR.get("periodo_referencia", "")] +
            [h.get("periodo_referencia") for h in historico_metas_unificado if h.get("periodo_referencia")]
        ))
        cp1, cp2, cp3 = st.columns([1, 1, 1.3])
        with cp1:
            periodo_meta_comp_unificado = periodo_gestao_unificado
            st.text_input("Período de referência", value=periodo_meta_comp_unificado, disabled=True, key="periodo_compradores_unificado_exibicao")
        garantir_metas_compradores_periodo(periodo_meta_comp_unificado)
        compradores_ativos_unificados = lista_compradores_ativos()
        with cp2:
            comprador_meta_unificado = st.selectbox(
                "Comprador",
                compradores_ativos_unificados,
                key="gestao_unificada_comprador",
            )
        with cp3:
            usuario_meta_comprador = st.text_input(
                "Gestor responsável",
                value="Gestor",
                key="gestao_unificada_usuario_comprador",
            )

        meta_comp_atual = obter_meta_comprador(comprador_meta_unificado, periodo_meta_comp_unificado)
        with st.form("gestao_unificada_form_comprador"):
            gc1, gc2, gc3 = st.columns(3)
            with gc1:
                meta_venda_comp = st.number_input("Meta de Venda (R$)", min_value=0.0, value=float(meta_comp_atual.get("meta_venda", 0)), step=1000.0, format="%.2f")
                participacao_comp = st.number_input("Participação da Venda (%)", 0.0, 100.0, float(meta_comp_atual.get("participacao_venda_pct", 0)), 0.1, format="%.2f")
            with gc2:
                meta_cmv_comp = st.number_input("Meta CMV (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_cmv_pct", 0)), 0.1, format="%.2f")
                fator_cob_comp = st.number_input("Fator de Cobertura", min_value=0.0, value=float(meta_comp_atual.get("fator_cobertura", 0)), step=0.05, format="%.2f")
            with gc3:
                meta_ruptura_comp = st.number_input("Meta de Ruptura (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_ruptura_pct", 0)), 0.1, format="%.2f")
                meta_reposicao_comp = st.number_input("Meta de Reposição (%)", 0.0, 200.0, float(meta_comp_atual.get("meta_reposicao_pct", 0)), 0.1, format="%.2f")

            st.markdown("#### Distribuição do estoque por curva")
            cc1, cc2, cc3, cc4 = st.columns(4)
            with cc1: curva_a_comp = st.number_input("Curva A (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_curva_a_pct", 0)), 1.0, key="unif_curva_a")
            with cc2: curva_b_comp = st.number_input("Curva B (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_curva_b_pct", 0)), 1.0, key="unif_curva_b")
            with cc3: curva_c_comp = st.number_input("Curva C (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_curva_c_pct", 0)), 1.0, key="unif_curva_c")
            with cc4: curva_d_comp = st.number_input("Curva D (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_curva_d_pct", 0)), 1.0, key="unif_curva_d")

            pc1, pc2 = st.columns(2)
            with pc1:
                premio_comp = st.number_input("Valor de prêmio atingível (R$)", min_value=0.0, value=float(meta_comp_atual.get("valor_premio", 0)), step=100.0, format="%.2f")
            with pc2:
                status_comp = st.selectbox(
                    "Status",
                    ["Planejada", "Ativa", "Encerrada", "Cancelada"],
                    index=["Planejada", "Ativa", "Encerrada", "Cancelada"].index(meta_comp_atual.get("status", "Ativa")),
                    key="gestao_unificada_status_comprador",
                )
            salvar_comp = st.form_submit_button("💾 Salvar meta do comprador", use_container_width=True)

            if salvar_comp:
                total_curvas_comp = curva_a_comp + curva_b_comp + curva_c_comp + curva_d_comp
                if abs(total_curvas_comp - 100.0) > 0.01:
                    st.error("A soma das curvas deve ser igual a 100%.")
                else:
                    dados_comp = carregar_metas_por_comprador()
                    nova_meta_comp = {
                        "periodo_referencia": periodo_meta_comp_unificado,
                        "comprador": comprador_meta_unificado,
                        "meta_venda": meta_venda_comp,
                        "participacao_venda_pct": participacao_comp,
                        "meta_cmv_pct": meta_cmv_comp,
                        "meta_cmv_valor": meta_venda_comp * meta_cmv_comp / 100.0,
                        "fator_cobertura": fator_cob_comp,
                        "meta_estoque_total": meta_venda_comp * meta_cmv_comp / 100.0 * fator_cob_comp,
                        "meta_curva_a_pct": curva_a_comp,
                        "meta_curva_b_pct": curva_b_comp,
                        "meta_curva_c_pct": curva_c_comp,
                        "meta_curva_d_pct": curva_d_comp,
                        "meta_ruptura_pct": meta_ruptura_comp,
                        "meta_reposicao_pct": meta_reposicao_comp,
                        "valor_premio": premio_comp,
                        "status": status_comp,
                        "usuario_cadastro": usuario_meta_comprador,
                        "ultima_atualizacao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    }
                    substituiu_comp = False
                    for idx_comp, item_comp in enumerate(dados_comp):
                        if item_comp.get("periodo_referencia") == periodo_meta_comp_unificado and item_comp.get("comprador") == comprador_meta_unificado:
                            dados_comp[idx_comp] = nova_meta_comp
                            substituiu_comp = True
                            break
                    if not substituiu_comp:
                        dados_comp.append(nova_meta_comp)
                    salvar_metas_por_comprador(dados_comp)
                    st.success("Meta do comprador salva com sucesso.")
                    st.rerun()

        resumo_comp = pd.DataFrame([
            item for item in carregar_metas_por_comprador()
            if item.get("periodo_referencia") == periodo_meta_comp_unificado
            and str(item.get("comprador", "")).strip().casefold() in _conjunto_compradores_ativos()
        ])
        if not resumo_comp.empty:
            st.markdown("#### Resumo das metas do período")
            dataframe_br(resumo_comp, use_container_width=True, hide_index=True, height=310)
            soma_vendas_comp = pd.to_numeric(resumo_comp.get("meta_venda", 0), errors="coerce").fillna(0).sum()
            soma_part_comp = pd.to_numeric(resumo_comp.get("participacao_venda_pct", 0), errors="coerce").fillna(0).sum()
            rc1, rc2, rc3 = st.columns(3)
            rc1.metric("Soma das metas de venda", moeda_real(soma_vendas_comp))
            rc2.metric("Meta geral do período", moeda_real(METAS_GESTOR.get('meta_venda_total_mes', 0)))
            rc3.metric("Participação total", percentual(soma_part_comp))

    with aba2:
        historico = carregar_historico()
        if historico:
            hist_df = pd.DataFrame(historico)
            colunas = [
                "id_meta", "periodo_referencia", "data_inicio", "data_fim",
                "status", "descricao", "meta_venda_total_mes",
                "meta_cmv_mes", "valor_premio_total", "usuario_cadastro",
                "ultima_atualizacao"
            ]
            hist_df = hist_df[[c for c in colunas if c in hist_df.columns]].copy()
            if "data_inicio" in hist_df:
                hist_df["data_inicio"] = hist_df["data_inicio"].map(data_br)
            if "data_fim" in hist_df:
                hist_df["data_fim"] = hist_df["data_fim"].map(data_br)
            if "meta_venda_total_mes" in hist_df:
                hist_df["meta_venda_total_mes"] = hist_df["meta_venda_total_mes"].map(moeda)
            if "meta_cmv_mes" in hist_df:
                hist_df["meta_cmv_mes"] = hist_df["meta_cmv_mes"].map(percentual)
            if "valor_premio_total" in hist_df:
                hist_df["valor_premio_total"] = hist_df["valor_premio_total"].map(moeda)
            dataframe_br(hist_df, use_container_width=True, hide_index=True, height=360)

            opcoes = [h.get("id_meta") for h in historico]
            meta_escolhida = st.selectbox("Carregar meta do histórico", opcoes)
            c1, c2 = st.columns(2)
            with c1:
                if st.button("📥 Carregar como meta ativa", use_container_width=True):
                    registro = next(h for h in historico if h.get("id_meta") == meta_escolhida)
                    salvar_metas(registro.copy(), registrar_historico=False)
                    st.success("Meta carregada.")
                    st.rerun()
            with c2:
                if st.button("📑 Duplicar para novo período", use_container_width=True):
                    registro = next(h for h in historico if h.get("id_meta") == meta_escolhida).copy()
                    registro["id_meta"] = f"{registro.get('id_meta','META')}-COPIA"
                    registro["periodo_referencia"] = f"{registro.get('periodo_referencia','')}-copia"
                    registro["status"] = "Planejada"
                    registro["data_cadastro"] = ""
                    salvar_metas(registro)
                    st.success("Meta duplicada. Ajuste o novo período no cadastro.")
                    st.rerun()
        else:
            st.info("Ainda não existem metas no histórico.")


elif visao == "Importar Ruptura":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Ruptura Automática</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Coloque o arquivo oficial na pasta IMPORTAR_RUPTURA. O sistema processa automaticamente.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Automático</div>
    </div>
    """, unsafe_allow_html=True)

    arquivo = _arquivo_mais_recente()
    controle = {}
    if RUPTURA_AUTO_CONTROLE.exists():
        try:
            controle = json.loads(RUPTURA_AUTO_CONTROLE.read_text(encoding="utf-8"))
        except Exception:
            controle = {}
    ultima = controle.get("ultima_importacao", {})

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Período da meta", METAS_GESTOR.get("periodo_referencia", "-"))
    with c2:
        st.metric("Arquivo detectado", arquivo.name if arquivo else "Nenhum")
    with c3:
        st.metric("Última importação", ultima.get("data", "Não realizada"))

    st.caption("Pasta monitorada automaticamente:")
    st.code(str(PASTA_RUPTURA_AUTO.resolve()), language=None)

    b1, b2 = st.columns(2)
    with b1:
        if st.button("🔄 Verificar agora", use_container_width=True):
            try:
                r = processar_ruptura_automatica()
                if r["status"] == "importado":
                    st.success(f"{r['registros']} registros importados para {r['periodo']}.")
                    st.rerun()
                else:
                    st.info(r["mensagem"])
            except Exception as e:
                st.error(str(e))
    with b2:
        if st.button("♻️ Reprocessar período", use_container_width=True):
            try:
                r = processar_ruptura_automatica(forcar=True)
                st.success(f"{r.get('registros', 0)} registros reprocessados.")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    periodos = list(dict.fromkeys(
        [METAS_GESTOR.get("periodo_referencia", "")] +
        [h.get("periodo_referencia") for h in carregar_historico() if h.get("periodo_referencia")]
    ))
    periodo_rup = st.selectbox("Período para análise", periodos)
    base_rup = carregar_ruptura_auto(periodo_rup)

    if not base_rup.empty:
        k1, k2, k3, k4 = st.columns(4)
        with k1: st.metric("Itens", numero_inteiro(len(base_rup)))
        with k2: st.metric("Ruptura total", moeda_real(base_rup['Valor Ruptura'].sum()))
        with k3: st.metric("Necessidade a custo", moeda_real(base_rup['Valor Necessidade Custo'].sum()))
        with k4: st.metric("Não mapeados", numero_inteiro((base_rup["Comprador"] == "Não mapeado").sum()))

        g1, g2 = st.columns(2)
        with g1:
            por_comp = base_rup.groupby("Comprador", as_index=False)["Valor Ruptura"].sum().sort_values("Valor Ruptura")
            fig = px.bar(por_comp, x="Valor Ruptura", y="Comprador", orientation="h", title="Ruptura por comprador")
            fig.update_layout(height=330, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font_color="#f3f7fb")
            plotly_chart_br(fig, use_container_width=True, config={"displayModeBar": False})
        with g2:
            por_curva = base_rup.groupby("Curva Valor", as_index=False)["Valor Ruptura"].sum()
            fig2 = px.pie(por_curva, names="Curva Valor", values="Valor Ruptura", hole=.55, title="Ruptura por curva")
            fig2.update_layout(height=330, paper_bgcolor="rgba(0,0,0,0)", font_color="#f3f7fb")
            plotly_chart_br(fig2, use_container_width=True, config={"displayModeBar": False})

        dataframe_br(base_rup, use_container_width=True, hide_index=True, height=420)
    else:
        st.warning("Ainda não existem dados processados para esse período.")

    st.markdown("### Histórico")
    hist = historico_ruptura_auto()
    if not hist.empty:
        dataframe_br(hist, use_container_width=True, hide_index=True)
    else:
        st.info("Nenhuma importação registrada.")

elif visao == "Banco de Dados":
    status_sql_fontes = status_configuracao_fontes()
    if not status_sql_fontes.get("contas_pagar", {}).get("configurado", False):
        st.error(
            "Contas a Pagar não possui uma consulta SQL real configurada. "
            "A consulta atual é apenas um modelo e sempre retornaria zero. "
            "Abra a seção de edição de SQL e cole a consulta oficial."
        )

    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Banco de Dados e Atualizações Mensais</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Configure o PostgreSQL, salve os scripts e atualize cada mês conforme o período da meta.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Rede Economize Data Center</div>
    </div>
    """, unsafe_allow_html=True)

    aba_conexao, aba_sql, aba_atualizar, aba_historico = st.tabs([
        "🔐 Conexão",
        "🧾 Scripts SQL",
        "🔄 Atualização mensal",
        "📚 Histórico e dados salvos",
    ])

    with aba_conexao:
        st.markdown("### Configuração do PostgreSQL")
        st.caption(
            "Os dados são salvos localmente em `config/database.json`. "
            "Em ambiente online, a senha deverá ser migrada para variáveis de ambiente ou secrets."
        )

        with st.form("form_config_banco"):
            c1, c2 = st.columns([2, 1])
            with c1:
                host = st.text_input("Host", value=CONFIG_BANCO.get("host", ""))
                banco = st.text_input("Banco de dados", value=CONFIG_BANCO.get("banco", ""))
                usuario = st.text_input("Usuário", value=CONFIG_BANCO.get("usuario", ""))
            with c2:
                porta = st.number_input(
                    "Porta",
                    min_value=1,
                    max_value=65535,
                    value=int(CONFIG_BANCO.get("porta", 5432)),
                    step=1
                )
                sslmode = st.selectbox(
                    "SSL Mode",
                    ["prefer", "require", "disable"],
                    index=["prefer", "require", "disable"].index(
                        CONFIG_BANCO.get("sslmode", "prefer")
                    )
                )
                salvar_senha = st.checkbox(
                    "Salvar senha localmente",
                    value=bool(CONFIG_BANCO.get("salvar_senha", True))
                )

            senha = st.text_input(
                "Senha",
                value=CONFIG_BANCO.get("senha", "") if CONFIG_BANCO.get("salvar_senha", True) else "",
                type="password"
            )

            b1, b2 = st.columns(2)
            salvar_cfg = b1.form_submit_button("💾 Salvar configuração", use_container_width=True)
            testar_cfg = b2.form_submit_button("🔌 Testar conexão", use_container_width=True)

            cfg_digitada = {
                "tipo": "PostgreSQL",
                "host": host.strip(),
                "porta": int(porta),
                "banco": banco.strip(),
                "usuario": usuario.strip(),
                "senha": senha if salvar_senha else "",
                "sslmode": sslmode,
                "salvar_senha": salvar_senha,
                "ultima_validacao": CONFIG_BANCO.get("ultima_validacao", ""),
            }

            if salvar_cfg:
                salvar_config_banco(cfg_digitada)
                st.success("Configuração salva.")
                st.rerun()

            if testar_cfg:
                try:
                    cfg_teste = cfg_digitada.copy()
                    cfg_teste["senha"] = senha
                    engine = criar_engine_banco(cfg_teste)
                    with engine.connect() as conn:
                        conn.execute(text("SELECT 1"))
                    engine.dispose()
                    cfg_digitada["ultima_validacao"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                    if salvar_senha:
                        cfg_digitada["senha"] = senha
                    salvar_config_banco(cfg_digitada)
                    st.success("Conexão realizada com sucesso.")
                except Exception as e:
                    st.error(f"Falha na conexão: {e}")

        if CONFIG_BANCO.get("ultima_validacao"):
            st.caption(f"Última conexão validada: {CONFIG_BANCO['ultima_validacao']}")

    with aba_sql:
        cfg_planos_banco = carregar_config_analise_comercial()
        planos_ativos_banco = cfg_planos_banco.get(
            "planos_contas_selecionados",
            [PLANO_CONTAS_PAGAMENTO_PADRAO],
        )
        st.markdown("### Planos atualmente selecionados")
        for plano_banco in planos_ativos_banco:
            st.success(f"✓ {plano_banco}")
        st.caption(
            "A seleção é alterada e salva na tela Análise Comercial. "
            "Todos os planos retornados pela consulta permanecem armazenados no cache."
        )

        st.markdown("### Scripts utilizados nas atualizações")
        st.success(
            "Os scripts enviados anteriormente já estão carregados. "
            "Você pode alterar e salvar qualquer um deles nesta tela."
        )
        st.info(
            "Use os parâmetros `:data_inicio`, `:data_fim` e, quando necessário, "
            "`:periodo_referencia`. O sistema envia automaticamente as datas da meta selecionada."
        )

        fonte_sql = st.selectbox(
            "Fonte",
            list(FONTES_BANCO.keys()),
            format_func=lambda x: FONTES_BANCO[x]["titulo"]
        )
        sql_atual = ler_sql(FONTES_BANCO[fonte_sql]["arquivo_sql"])
        sql_configurado, sql_diagnostico = diagnosticar_sql_fonte(
            fonte_sql, sql_atual
        )
        if not sql_configurado:
            st.warning(sql_diagnostico)
        else:
            st.success("Consulta SQL configurada.")
        st.caption(
            f"Arquivo carregado: {FONTES_BANCO[fonte_sql]['arquivo_sql']} • "
            f"{len(sql_atual.splitlines())} linhas"
        )
        sql_editado = st.text_area(
            f"SQL de {FONTES_BANCO[fonte_sql]['titulo']}",
            value=sql_atual,
            height=420
        )

        if st.button("💾 Salvar script SQL", use_container_width=True):
            salvar_sql(FONTES_BANCO[fonte_sql]["arquivo_sql"], sql_editado)
            st.success("Script salvo.")

    with aba_atualizar:
        st.markdown("### Atualização por período da meta")

        historico_metas = carregar_historico()
        periodos_disponiveis = [
            h.get("periodo_referencia")
            for h in historico_metas
            if h.get("periodo_referencia")
        ]
        periodo_ativo = PERIODO_GLOBAL_SELECIONADO
        periodos_disponiveis = list(
            dict.fromkeys([periodo_ativo] + periodos_disponiveis)
        )

        periodo_selecionado = st.selectbox(
            "Período da meta",
            periodos_disponiveis,
            index=0
        )

        meta_periodo = next(
            (
                h for h in historico_metas
                if h.get("periodo_referencia") == periodo_selecionado
            ),
            METAS_GESTOR
        )

        data_inicio_periodo = meta_periodo.get("data_inicio", METAS_GESTOR.get("data_inicio"))
        data_fim_periodo = meta_periodo.get("data_fim", METAS_GESTOR.get("data_fim"))

        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("Período", periodo_selecionado)
        with c2:
            st.metric("Data inicial", data_br(data_inicio_periodo))
        with c3:
            st.metric("Data final", data_br(data_fim_periodo))

        st.caption(
            "Ao atualizar novamente o mesmo período, o sistema substitui somente os dados "
            "daquele mês e mantém os demais meses, metas, compradores, configurações e "
            "históricos já salvos."
        )

        fontes_selecionadas = st.multiselect(
            "Fontes a atualizar",
            list(FONTES_BANCO.keys()),
            default=list(FONTES_BANCO.keys()),
            format_func=lambda x: FONTES_BANCO[x]["titulo"]
        )

        confirmar = st.checkbox(
            f"Confirmo a atualização do período {periodo_selecionado}"
        )

        if st.button(
            "🧪 Testar somente a consulta de Contas a Pagar",
            use_container_width=True,
            key="testar_sql_contas_pagar_direto",
        ):
            cfg_teste = carregar_config_banco()
            cfg_teste["senha"] = (
                cfg_teste.get("senha")
                or st.session_state.get("senha_banco_execucao", "")
            )
            try:
                with st.spinner("Executando o SQL oficial diretamente no PostgreSQL..."):
                    sql_teste = ler_sql(FONTES_BANCO["contas_pagar"]["arquivo_sql"])
                    df_teste = executar_contas_pagar_psycopg2(
                        cfg_teste, sql_teste, data_inicio_periodo, data_fim_periodo, periodo_selecionado
                    )
                st.success(f"Consulta executada com sucesso: {len(df_teste):,} registros.".replace(",", "."))
                st.caption("Ambiente PostgreSQL: " + " | ".join(map(str, df_teste.attrs.get("ambiente_postgresql", ()))))
                dataframe_br(df_teste.head(20), use_container_width=True)
            except Exception as erro_teste:
                st.error("A consulta foi recusada pelo PostgreSQL.")
                st.code(str(erro_teste), language="text")

        if st.button("🔄 Atualizar fontes selecionadas", use_container_width=True):
            if not confirmar:
                st.warning("Marque a confirmação antes de atualizar.")
            elif not fontes_selecionadas:
                st.warning("Selecione ao menos uma fonte.")
            else:
                cfg_execucao = carregar_config_banco()
                if not cfg_execucao.get("senha"):
                    senha_execucao = st.session_state.get("senha_banco_execucao", "")
                else:
                    senha_execucao = cfg_execucao.get("senha", "")

                if not cfg_execucao.get("host") or not cfg_execucao.get("banco"):
                    st.error("Configure o banco de dados antes da atualização.")
                elif not senha_execucao:
                    st.error(
                        "A senha não está salva. Marque 'Salvar senha localmente' "
                        "ou salve novamente a configuração."
                    )
                else:
                    cfg_execucao["senha"] = senha_execucao

                    total_fontes = len(fontes_selecionadas)
                    total_etapas = total_fontes + 2
                    resultados = []
                    logs = []
                    inicio_geral = time.perf_counter()

                    progresso = st.progress(
                        0,
                        text="Preparando a atualização..."
                    )
                    status_principal = st.empty()
                    metricas_progresso = st.columns(4)
                    log_area = st.empty()

                    def atualizar_painel(etapa, texto, fonte_atual="-", registros=0):
                        decorrido = time.perf_counter() - inicio_geral
                        percentual_etapa = min(etapa / total_etapas, 1.0)
                        progresso.progress(
                            percentual_etapa,
                            text=f"{percentual_etapa * 100:.0f}% • {texto}"
                        )
                        status_principal.info(texto)
                        metricas_progresso[0].metric(
                            "Etapa",
                            f"{etapa}/{total_etapas}"
                        )
                        metricas_progresso[1].metric(
                            "Fonte atual",
                            fonte_atual
                        )
                        metricas_progresso[2].metric(
                            "Registros",
                            numero_inteiro(registros)
                        )
                        metricas_progresso[3].metric(
                            "Tempo decorrido",
                            time.strftime("%H:%M:%S", time.gmtime(decorrido))
                        )
                        log_area.code(
                            "\n".join(logs[-15:]) or "Aguardando início...",
                            language=None
                        )

                    logs.append(
                        f"{datetime.now():%H:%M:%S} • Validando conexão e período {periodo_selecionado}"
                    )
                    atualizar_painel(
                        1,
                        "Validando conexão com o banco e preparando os scripts..."
                    )

                    for indice, fonte in enumerate(fontes_selecionadas, start=1):
                        titulo_fonte = FONTES_BANCO[fonte]["titulo"]
                        inicio_fonte = time.perf_counter()

                        logs.append(
                            f"{datetime.now():%H:%M:%S} • {titulo_fonte}: consulta iniciada"
                        )
                        atualizar_painel(
                            indice,
                            f"Consultando {titulo_fonte} no PostgreSQL...",
                            titulo_fonte,
                            0
                        )

                        try:
                            qtd = executar_atualizacao_fonte(
                                fonte,
                                cfg_execucao,
                                periodo_selecionado,
                                data_inicio_periodo,
                                data_fim_periodo,
                            )
                            tempo_fonte = time.perf_counter() - inicio_fonte
                            logs.append(
                                f"{datetime.now():%H:%M:%S} • {titulo_fonte}: "
                                f"{qtd:,} registros salvos em {tempo_fonte:.1f}s"
                                .replace(",", ".")
                            )
                            resultados.append(
                                {
                                    "Fonte": titulo_fonte,
                                    "Status": "Sucesso",
                                    "Registros": qtd,
                                    "Tempo": time.strftime(
                                        "%H:%M:%S",
                                        time.gmtime(tempo_fonte)
                                    ),
                                    "Mensagem": "Atualizada e salva",
                                }
                            )
                            atualizar_painel(
                                indice + 1,
                                f"{titulo_fonte} concluída e salva.",
                                titulo_fonte,
                                qtd
                            )

                        except Exception as e:
                            tempo_fonte = time.perf_counter() - inicio_fonte
                            logs.append(
                                f"{datetime.now():%H:%M:%S} • {titulo_fonte}: ERRO — {e}"
                            )
                            resultados.append(
                                {
                                    "Fonte": titulo_fonte,
                                    "Status": "Erro",
                                    "Registros": 0,
                                    "Tempo": time.strftime(
                                        "%H:%M:%S",
                                        time.gmtime(tempo_fonte)
                                    ),
                                    "Mensagem": str(e),
                                }
                            )
                            atualizar_painel(
                                indice + 1,
                                f"Erro ao atualizar {titulo_fonte}.",
                                titulo_fonte,
                                0
                            )

                    logs.append(
                        f"{datetime.now():%H:%M:%S} • Recalculando as visões e os KPIs"
                    )
                    atualizar_painel(
                        total_fontes + 1,
                        "Recalculando cards, métricas, gráficos e premiações..."
                    )

                    tempo_total = time.perf_counter() - inicio_geral
                    logs.append(
                        f"{datetime.now():%H:%M:%S} • Atualização finalizada em "
                        f"{time.strftime('%H:%M:%S', time.gmtime(tempo_total))}"
                    )
                    atualizar_painel(
                        total_etapas,
                        "Atualização finalizada.",
                        "Concluído",
                        sum(item["Registros"] for item in resultados)
                    )

                    if any(item["Status"] == "Erro" for item in resultados):
                        status_principal.warning(
                            "Atualização concluída com uma ou mais falhas. "
                            "Consulte o resultado abaixo."
                        )
                    else:
                        status_principal.success(
                            "Todas as fontes selecionadas foram atualizadas e salvas."
                        )

                    resultados_df = pd.DataFrame(resultados)
                    dataframe_br(
                        resultados_df,
                        use_container_width=True,
                        hide_index=True
                    )
                    erros_df = resultados_df[
                        resultados_df["Status"].astype(str).str.casefold().eq("erro")
                    ]
                    if not erros_df.empty:
                        with st.expander(
                            "🔎 Detalhes completos do erro", expanded=True
                        ):
                            for _, linha_erro in erros_df.iterrows():
                                st.code(
                                    f"{linha_erro['Fonte']}\n{linha_erro['Mensagem']}",
                                    language="text",
                                )

    with aba_historico:
        st.markdown("### Limpeza e reconstrução das bases")
        st.caption(
            "Remove somente Vendas, Estoque, Entradas, Contas a Pagar, resumos "
            "e histórico técnico. Metas, compradores, mapas, configurações e SQL "
            "permanecem preservados."
        )
        confirmar_limpeza = st.checkbox(
            "Confirmo que desejo apagar todas as bases e resumos salvos",
            key="confirmar_limpeza_total_bases",
        )
        if st.button(
            "🧹 Limpar tudo para nova atualização",
            type="primary",
            use_container_width=True,
            disabled=not confirmar_limpeza,
            key="limpar_todas_bases_operacionais",
        ):
            limpar_dados_operacionais()
            st.success(
                "Bases e resumos apagados. O projeto está pronto para uma nova carga."
            )
            st.rerun()

        st.markdown("### Importar base anual de Contas a Pagar")
        st.caption(
            "Use o CSV exportado pelo ERP. O sistema separa os registros por "
            "mês de vencimento e substitui apenas o cache de Contas a Pagar."
        )
        csv_contas_pagar = st.file_uploader(
            "Arquivo CSV anual",
            type=["csv", "txt"],
            key="upload_csv_anual_contas_pagar",
        )
        if csv_contas_pagar is not None:
            if st.button(
                "📥 Importar e reconstruir Contas a Pagar",
                use_container_width=True,
                key="importar_csv_anual_contas_pagar",
            ):
                try:
                    with st.spinner("Importando e separando as competências..."):
                        resultado_importacao, planos_importados = (
                            importar_csv_anual_contas_pagar(csv_contas_pagar)
                        )
                        st.cache_data.clear()
                    st.success(
                        "Contas a Pagar importado e salvo por competência."
                    )
                    dataframe_br(
                        resultado_importacao,
                        use_container_width=True,
                        hide_index=True,
                    )
                    if planos_importados:
                        st.caption(
                            f"Planos identificados: {len(planos_importados)}"
                        )
                except Exception as erro_importacao:
                    st.error(str(erro_importacao))

        arquivo_log_cp = LOG_DIR / "contas_pagar_erros.log"
        if arquivo_log_cp.exists():
            with st.expander(
                "🔎 Último erro técnico de Contas a Pagar",
                expanded=False,
            ):
                conteudo_log = arquivo_log_cp.read_text(
                    encoding="utf-8",
                    errors="replace",
                )
                st.code(conteudo_log[-8000:], language="text")

        sql_log_cp = LOG_DIR / "contas_pagar_sql_executado.sql"
        erro_log_cp = LOG_DIR / "contas_pagar_erro.txt"

        if erro_log_cp.exists():
            with st.expander(
                "🔎 Diagnóstico completo de Contas a Pagar",
                expanded=True,
            ):
                st.code(
                    erro_log_cp.read_text(
                        encoding="utf-8",
                        errors="replace",
                    )[-12000:],
                    language="text",
                )

        if sql_log_cp.exists():
            with st.expander(
                "🧾 SQL final executado em Contas a Pagar",
                expanded=False,
            ):
                st.code(
                    sql_log_cp.read_text(
                        encoding="utf-8",
                        errors="replace",
                    ),
                    language="sql",
                )

        st.markdown("### Histórico de atualizações")
        hist = historico_atualizacoes()
        if not hist.empty:
            dataframe_br(hist, use_container_width=True, hide_index=True, height=300)
        else:
            st.info("Nenhuma atualização de banco foi realizada.")

        st.markdown("### Consultar dados mensais salvos")
        c1, c2 = st.columns(2)
        with c1:
            fonte_consulta = st.selectbox(
                "Fonte salva",
                list(FONTES_BANCO.keys()),
                format_func=lambda x: FONTES_BANCO[x]["titulo"],
                key="fonte_consulta_cache"
            )
        with c2:
            periodo_consulta = st.selectbox(
                "Período salvo",
                list(dict.fromkeys(
                    [METAS_GESTOR.get("periodo_referencia", "")] +
                    [
                        h.get("periodo_referencia")
                        for h in carregar_historico()
                        if h.get("periodo_referencia")
                    ]
                )),
                key="periodo_consulta_cache"
            )

        snapshot = carregar_snapshot(fonte_consulta, periodo_consulta)
        if not snapshot.empty:
            st.metric("Registros encontrados", numero_inteiro(len(snapshot)))
            dataframe_br(snapshot, use_container_width=True, hide_index=True, height=400)

            csv_snapshot = snapshot.to_csv(
                index=False,
                sep=";",
                encoding="utf-8-sig"
            ).encode("utf-8-sig")
            st.download_button(
                "📥 Exportar dados salvos",
                data=csv_snapshot,
                file_name=f"{fonte_consulta}_{periodo_consulta}.csv",
                mime="text/csv",
                use_container_width=True
            )
        else:
            st.warning("Não existem dados salvos para essa fonte e período.")



elif visao == "Compradores por Classificação":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Compradores por Classificação Principal</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Altere o comprador responsável por cada classificação sem modificar o código do projeto.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Gestão de Responsabilidades</div>
    </div>
    """, unsafe_allow_html=True)

    st.info(
        "As alterações são salvas e passam a valer automaticamente nas visões, "
        "gráficos, metas, resultados e premiações."
    )

    mapa_df = pd.DataFrame(carregar_mapa_compradores_editavel())
    if mapa_df.empty:
        mapa_df = pd.DataFrame(columns=["Área", "Classificação Principal", "Comprador"])

    compradores_disponiveis = sorted(lista_compradores_ativos(), key=lambda x: x.casefold())

    editado = st.data_editor(
        mapa_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Área": st.column_config.TextColumn("Área", width="medium"),
            "Classificação Principal": st.column_config.TextColumn(
                "Classificação Principal",
                width="large",
                required=True
            ),
            "Comprador": st.column_config.SelectboxColumn(
                "Comprador Responsável",
                options=compradores_disponiveis,
                required=True,
                width="medium"
            ),
        },
        key="editor_mapa_compradores"
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("💾 Salvar alterações", use_container_width=True):
            registros = editado.fillna("").to_dict(orient="records")
            registros = [
                {
                    "Área": str(r.get("Área", "")).strip(),
                    "Classificação Principal": str(r.get("Classificação Principal", "")).strip(),
                    "Comprador": str(r.get("Comprador", "")).strip(),
                }
                for r in registros
                if str(r.get("Classificação Principal", "")).strip()
            ]
            salvar_mapa_compradores_editavel(registros)
            st.success("Responsabilidades atualizadas com sucesso.")
            st.rerun()

    with c2:
        if st.button("↩️ Restaurar padrão inicial", use_container_width=True):
            salvar_mapa_compradores_editavel(MAPA_COMPRADORES_PADRAO)
            st.success("Mapa padrão restaurado.")
            st.rerun()

    with c3:
        csv_mapa = editado.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig")
        st.download_button(
            "📥 Exportar mapa",
            data=csv_mapa,
            file_name="mapa_compradores_classificacao.csv",
            mime="text/csv",
            use_container_width=True
        )

    st.markdown("### Resumo por comprador")
    resumo = (
        editado.groupby("Comprador", dropna=False)
        .size()
        .reset_index(name="Classificações")
        .sort_values("Classificações", ascending=False)
    )
    dataframe_br(resumo, use_container_width=True, hide_index=True)

    st.caption(
        "Após alterar o mapa, atualize ou reabra o painel para recalcular todas as visões "
        "com o novo responsável."
    )



elif visao == "Cadastro de Compradores":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Cadastro de Compradores</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Inclua, altere, renomeie ou inative compradores diretamente pelo sistema.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Gestão de Usuários Responsáveis</div>
    </div>
    """, unsafe_allow_html=True)

    st.info(
        "Ao alterar o nome de um comprador, o sistema também atualiza automaticamente "
        "as classificações vinculadas ao nome anterior."
    )

    cadastro_df = pd.DataFrame(carregar_cadastro_compradores())
    if cadastro_df.empty:
        cadastro_df = pd.DataFrame(columns=["Comprador", "Status"])

    original_df = cadastro_df.copy()

    editado_compradores = st.data_editor(
        cadastro_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Comprador": st.column_config.TextColumn(
                "Nome do Comprador",
                required=True,
                width="large"
            ),
            "Status": st.column_config.SelectboxColumn(
                "Status",
                options=["Ativo", "Inativo"],
                required=True,
                width="medium"
            ),
        },
        key="editor_cadastro_compradores"
    )

    c1, c2 = st.columns(2)

    with c1:
        if st.button("💾 Salvar compradores", use_container_width=True):
            novos = []
            nomes_usados = set()

            for _, row in editado_compradores.fillna("").iterrows():
                nome = str(row.get("Comprador", "")).strip()
                status = str(row.get("Status", "Ativo")).strip() or "Ativo"
                if not nome:
                    continue
                if nome.lower() in nomes_usados:
                    st.error(f"Comprador duplicado: {nome}")
                    st.stop()
                nomes_usados.add(nome.lower())
                novos.append({"Comprador": nome, "Status": status})

            if not novos:
                st.error("Cadastre pelo menos um comprador.")
            else:
                # Detecta renomeações pela posição original e atualiza o mapa.
                mapa = carregar_mapa_compradores_editavel()
                limite = min(len(original_df), len(editado_compradores))

                for i in range(limite):
                    antigo = str(original_df.iloc[i].get("Comprador", "")).strip()
                    novo = str(editado_compradores.iloc[i].get("Comprador", "")).strip()
                    if antigo and novo and antigo != novo:
                        for item in mapa:
                            if str(item.get("Comprador", "")).strip() == antigo:
                                item["Comprador"] = novo
                        atualizar_nome_comprador_metas(antigo, novo)

                salvar_mapa_compradores_editavel(mapa)
                salvar_cadastro_compradores(novos)
                st.success("Cadastro de compradores atualizado.")
                st.rerun()

    with c2:
        if st.button("↩️ Restaurar compradores padrão", use_container_width=True):
            salvar_cadastro_compradores(COMPRADORES_PADRAO)
            st.success("Compradores padrão restaurados.")
            st.rerun()

    st.markdown("### Compradores ativos")
    ativos = pd.DataFrame(
        [{"Comprador": nome} for nome in lista_compradores_ativos()]
    )
    dataframe_br(ativos, use_container_width=True, hide_index=True)

    st.caption(
        "Compradores inativos permanecem no histórico, mas deixam de aparecer "
        "nos filtros e nos novos vínculos de classificação."
    )



elif visao == "Metas por Comprador":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Metas por Comprador</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Defina metas individuais para cada comprador e para cada período.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Gestão Individual de Performance</div>
    </div>
    """, unsafe_allow_html=True)

    st.info(
        "Ao criar um novo comprador, o sistema gera automaticamente uma meta inicial "
        "para o período ativo. Depois, você pode ajustar os valores individualmente."
    )

    historico_metas = carregar_historico()
    periodos = list(dict.fromkeys(
        [METAS_GESTOR.get("periodo_referencia", "")] +
        [
            h.get("periodo_referencia")
            for h in historico_metas
            if h.get("periodo_referencia")
        ]
    ))

    c_periodo, c_comprador = st.columns(2)
    with c_periodo:
        periodo_meta_comprador = st.selectbox(
            "Período da meta",
            periodos,
            key="periodo_meta_comprador"
        )

    metas_individuais = garantir_metas_compradores_periodo(periodo_meta_comprador)
    compradores_ativos = lista_compradores_ativos()

    with c_comprador:
        comprador_meta = st.selectbox(
            "Comprador",
            compradores_ativos,
            key="comprador_meta_individual"
        )

    meta_atual = obter_meta_comprador(comprador_meta, periodo_meta_comprador)

    with st.form("form_meta_individual_comprador"):
        st.markdown(f"### Meta de {comprador_meta}")

        g1, g2, g3 = st.columns(3)
        with g1:
            meta_venda_ind = st.number_input(
                "Meta de Venda (R$)",
                min_value=0.0,
                value=float(meta_atual.get("meta_venda", 0)),
                step=1000.0,
                format="%.2f"
            )
            participacao_ind = st.number_input(
                "Participação da Venda (%)",
                min_value=0.0,
                max_value=100.0,
                value=float(meta_atual.get("participacao_venda_pct", 0)),
                step=0.1,
                format="%.2f"
            )
        with g2:
            meta_cmv_pct_ind = st.number_input(
                "Meta CMV (%)",
                min_value=0.0,
                max_value=100.0,
                value=float(meta_atual.get("meta_cmv_pct", 0)),
                step=0.1,
                format="%.2f"
            )
            fator_cob_ind = st.number_input(
                "Fator de Cobertura",
                min_value=0.0,
                value=float(meta_atual.get("fator_cobertura", 0)),
                step=0.05,
                format="%.2f"
            )
        with g3:
            meta_ruptura_ind = st.number_input(
                "Meta de Ruptura (%)",
                min_value=0.0,
                max_value=100.0,
                value=float(meta_atual.get("meta_ruptura_pct", 0)),
                step=0.1,
                format="%.2f"
            )
            meta_reposicao_ind = st.number_input(
                "Meta de Reposição (%)",
                min_value=0.0,
                max_value=200.0,
                value=float(meta_atual.get("meta_reposicao_pct", 0)),
                step=0.1,
                format="%.2f"
            )

        st.markdown("### Distribuição do estoque por curva")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            curva_a_ind = st.number_input(
                "Curva A (%)", 0.0, 100.0,
                float(meta_atual.get("meta_curva_a_pct", 0)), 1.0
            )
        with c2:
            curva_b_ind = st.number_input(
                "Curva B (%)", 0.0, 100.0,
                float(meta_atual.get("meta_curva_b_pct", 0)), 1.0
            )
        with c3:
            curva_c_ind = st.number_input(
                "Curva C (%)", 0.0, 100.0,
                float(meta_atual.get("meta_curva_c_pct", 0)), 1.0
            )
        with c4:
            curva_d_ind = st.number_input(
                "Curva D (%)", 0.0, 100.0,
                float(meta_atual.get("meta_curva_d_pct", 0)), 1.0
            )

        valor_premio_ind = st.number_input(
            "Valor de Prêmio Atingível (R$)",
            min_value=0.0,
            value=float(meta_atual.get("valor_premio", 0)),
            step=100.0,
            format="%.2f"
        )

        status_ind = st.selectbox(
            "Status da meta individual",
            ["Planejada", "Ativa", "Encerrada", "Cancelada"],
            index=["Planejada", "Ativa", "Encerrada", "Cancelada"].index(
                meta_atual.get("status", "Ativa")
            )
        )

        salvar_meta_ind = st.form_submit_button(
            "💾 Salvar meta do comprador",
            use_container_width=True
        )

        if salvar_meta_ind:
            total_curvas = curva_a_ind + curva_b_ind + curva_c_ind + curva_d_ind
            if abs(total_curvas - 100.0) > 0.01:
                st.error("A soma das curvas deve ser igual a 100%.")
            else:
                dados = carregar_metas_por_comprador()
                nova_meta = {
                    "periodo_referencia": periodo_meta_comprador,
                    "comprador": comprador_meta,
                    "meta_venda": meta_venda_ind,
                    "participacao_venda_pct": participacao_ind,
                    "meta_cmv_pct": meta_cmv_pct_ind,
                    "meta_cmv_valor": meta_venda_ind * meta_cmv_pct_ind / 100.0,
                    "fator_cobertura": fator_cob_ind,
                    "meta_estoque_total": (
                        meta_venda_ind
                        * meta_cmv_pct_ind / 100.0
                        * fator_cob_ind
                    ),
                    "meta_curva_a_pct": curva_a_ind,
                    "meta_curva_b_pct": curva_b_ind,
                    "meta_curva_c_pct": curva_c_ind,
                    "meta_curva_d_pct": curva_d_ind,
                    "meta_ruptura_pct": meta_ruptura_ind,
                    "meta_reposicao_pct": meta_reposicao_ind,
                    "valor_premio": valor_premio_ind,
                    "status": status_ind,
                    "ultima_atualizacao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                }

                atualizou = False
                for i, item in enumerate(dados):
                    if (
                        item.get("periodo_referencia") == periodo_meta_comprador
                        and item.get("comprador") == comprador_meta
                    ):
                        dados[i] = nova_meta
                        atualizou = True
                        break

                if not atualizou:
                    dados.append(nova_meta)

                salvar_metas_por_comprador(dados)
                st.success("Meta individual salva com sucesso.")
                st.rerun()

    st.markdown("### Resumo das metas do período")
    resumo_periodo = pd.DataFrame([
        item
        for item in carregar_metas_por_comprador()
        if item.get("periodo_referencia") == periodo_meta_comprador
    ])

    if not resumo_periodo.empty:
        colunas_resumo = [
            "comprador", "meta_venda", "participacao_venda_pct",
            "meta_cmv_pct", "fator_cobertura",
            "meta_ruptura_pct", "meta_reposicao_pct",
            "valor_premio", "status", "ultima_atualizacao"
        ]
        resumo_periodo = resumo_periodo[
            [c for c in colunas_resumo if c in resumo_periodo.columns]
        ].copy()

        if "meta_venda" in resumo_periodo:
            resumo_periodo["meta_venda"] = resumo_periodo["meta_venda"].map(moeda)
        if "valor_premio" in resumo_periodo:
            resumo_periodo["valor_premio"] = resumo_periodo["valor_premio"].map(moeda)
        for coluna in [
            "participacao_venda_pct", "meta_cmv_pct",
            "meta_ruptura_pct", "meta_reposicao_pct"
        ]:
            if coluna in resumo_periodo:
                resumo_periodo[coluna] = resumo_periodo[coluna].map(percentual)

        dataframe_br(
            resumo_periodo,
            use_container_width=True,
            hide_index=True
        )

        total_meta_venda = sum(
            float(item.get("meta_venda", 0))
            for item in carregar_metas_por_comprador()
            if item.get("periodo_referencia") == periodo_meta_comprador
            and item.get("status") != "Cancelada"
        )
        total_participacao = sum(
            float(item.get("participacao_venda_pct", 0))
            for item in carregar_metas_por_comprador()
            if item.get("periodo_referencia") == periodo_meta_comprador
            and item.get("status") != "Cancelada"
        )

        r1, r2, r3 = st.columns(3)
        with r1:
            st.metric("Soma das metas de venda", moeda_real(total_meta_venda))
        with r2:
            st.metric("Meta geral do período", moeda_real(METAS_GESTOR.get('meta_venda_total_mes', 0)))
        with r3:
            st.metric("Participação total", percentual(total_participacao))

        if abs(total_meta_venda - float(METAS_GESTOR.get("meta_venda_total_mes", 0))) > 0.01:
            st.warning(
                "A soma das metas individuais está diferente da meta geral do período."
            )
        if abs(total_participacao - 100.0) > 0.01:
            st.warning(
                "A soma das participações dos compradores está diferente de 100%."
            )


renderizar_exportacao_tela()

st.markdown("<div class='eirox-footer'>REDE ECONOMIZE • KPI COMMERCIAL</div>", unsafe_allow_html=True)
